import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def multiline_text(value):
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def split_bullets(value, limit=8):
    text = multiline_text(value)
    if not text:
        return []
    text = re.sub(r"[📊🤝🧭✅🎯⚙️🛠️📌]+", "", text)
    parts = []
    for raw in re.split(r"\n+|(?:^|\s)[•\-]\s+", text):
        item = clean_text(raw)
        item = re.sub(r"^(HARD SKILLS|SOFT SKILLS|KPIS?|COMPETENCIAS|ACTIVIDADES|HERRAMIENTAS)\s*:?\s*", "", item, flags=re.I)
        if item and item not in parts:
            parts.append(item)
    return parts[:limit]


def skill_sections(value):
    text = multiline_text(value)
    hard = []
    soft = []
    if not text:
        return hard, soft

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    hard_match = re.search(r"HARD SKILLS\s*:?(.*?)(?:SOFT SKILLS\s*:|$)", normalized, flags=re.I | re.S)
    soft_match = re.search(r"SOFT SKILLS\s*:?(.*)$", normalized, flags=re.I | re.S)
    hard = split_bullets(hard_match.group(1) if hard_match else normalized, 6)
    soft = split_bullets(soft_match.group(1) if soft_match else "", 6)
    if not soft:
        soft = split_bullets(normalized, 6)[-3:]
    return hard, soft


def score_for(text, salt):
    total = sum(ord(ch) for ch in f"{text}|{salt}")
    return 72 + (total % 24)


def build(input_path, output_path):
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    cols = {
        "direccion": 2,
        "area": 3,
        "departamento": 4,
        "puesto": 5,
        "nivel": 6,
        "sueldo": 7,
        "formacion": 8,
        "experiencia": 9,
        "skills": 10,
        "objetivo": 11,
        "actividades": 12,
        "herramientas": 13,
        "kpis": 14,
        "competencias": 15,
        "compliance": 16,
    }

    perfiles = []
    seen = set()
    for sheet in workbook.worksheets:
        empresa = "Furiamotos" if "furia" in sheet.title.lower() else "MaxiMX"
        for row in range(4, sheet.max_row + 1):
            raw = {name: multiline_text(sheet.cell(row, col).value) for name, col in cols.items()}
            if not raw["puesto"] or not raw["departamento"]:
                continue

            key = "|".join(clean_text(raw[k]).upper() for k in ("direccion", "area", "departamento", "puesto"))
            if key in seen:
                continue
            seen.add(key)

            hard, soft = skill_sections(raw["skills"])
            perfil_id = re.sub(r"[^a-z0-9]+", "-", key.lower()).strip("-")
            perfil = {
                "id": perfil_id,
                "empresa": empresa,
                "nombre": clean_text(raw["puesto"]).title(),
                "direccion": clean_text(raw["direccion"]).title(),
                "area": clean_text(raw["area"]).title(),
                "departamento": clean_text(raw["departamento"]).title(),
                "nivel": clean_text(raw["nivel"]),
                "sueldo": clean_text(raw["sueldo"]),
                "formacion": clean_text(raw["formacion"]),
                "experiencia": split_bullets(raw["experiencia"], 5),
                "objetivo": clean_text(raw["objetivo"]),
                "actividades": split_bullets(raw["actividades"], 6),
                "herramientas": split_bullets(raw["herramientas"], 6),
                "hard": [{"nombre": item, "valor": score_for(item, key)} for item in hard],
                "soft": soft,
                "kpis": split_bullets(raw["kpis"], 6),
                "competencias": [{"nombre": item, "valor": score_for(item, key)} for item in split_bullets(raw["competencias"], 6)],
                "compliance": split_bullets(raw["compliance"], 4),
            }
            perfiles.append(perfil)

    perfiles.sort(key=lambda item: (item["direccion"], item["area"], item["departamento"], item["nombre"]))
    payload = {
        "meta": {
            "fuente": Path(input_path).name,
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "estado": "dummy",
            "perfiles": len(perfiles),
            "direcciones": len({p["direccion"] for p in perfiles}),
            "areas": len({p["area"] for p in perfiles}),
            "departamentos": len({p["departamento"] for p in perfiles}),
            "puestos": len({p["nombre"] for p in perfiles}),
        },
        "perfiles": perfiles,
    }
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload["meta"]


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python generar_perfiles_puestos_dummy.py input.xlsx output.json")
    print(json.dumps(build(sys.argv[1], sys.argv[2]), ensure_ascii=False))
