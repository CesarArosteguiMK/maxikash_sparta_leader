from __future__ import annotations

import os
import sys
import tempfile
import unittest
from copy import deepcopy
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch


TEST_DIR = os.path.dirname(os.path.abspath(__file__))
AGENT_ROOT = os.path.normpath(os.path.join(TEST_DIR, ".."))
sys.path.insert(0, os.path.join(AGENT_ROOT, "pydeps"))
sys.path.insert(0, os.path.join(AGENT_ROOT, "scripts"))

import reporte_cobranza as rc
from openpyxl import load_workbook


def registro_cartera(
    *,
    caso: str = "CONTROL",
    cuota: float = 600.0,
    deuda_gc: float = 250.0,
    saldo_a_favor: float = 850.0,
    saldo_aplicable: float = 250.0,
    status: str = "Vigente",
) -> dict:
    """Registro minimo del flujo historico, justo antes de la exclusion puntual."""
    numero_caso = int("".join(ch for ch in caso if ch.isdigit()) or 0)
    return {
        "ID_CREDITO": 9_000_000 + numero_caso,
        "ID_CLIENTE": 8_000_000 + numero_caso,
        "NOMBRE_CLIENTE": "CLIENTE ANONIMIZADO",
        "STATUS_CREDITO": status,
        "CUOTA_SEMANAL": float(cuota),
        "DEUDA_GC": float(deuda_gc),
        "SALDO_A_FAVOR": float(saldo_a_favor),
        "SALDO_APLICABLE_GC": float(saldo_aplicable),
        "FECHA_ULTIMO_ABONO_EFECTIVO": "2026-07-27",
        "ULTIMO_ABONO_EFECTIVO": -float(max(saldo_a_favor, cuota)),
        "MAXI_APP_CONECTO": "No",
        "COMENTARIOS": rc.COMENTARIO_APLICAR,
        "ERROR": "",
        "_MOTIVO_REVISION_CARTERA": "",
    }

def estado_cuenta_con_pagos(
    *,
    lunes_objetivo: date,
    montos: list[float],
    cuota: float = 551.0,
    saldo_vencido: float = 0.0,
    fecha_valor: date = date(2026, 7, 25),
    extemporaneos: list[float] | None = None,
) -> dict:
    """Estado de cuenta mínimo compatible con el algoritmo real de reparto."""
    montos_extemporaneos = extemporaneos or [0.0] * len(montos)
    if len(montos_extemporaneos) != len(montos):
        raise ValueError("extemporaneos debe tener la misma longitud que montos")
    return {
        "statusCredito": "Vigente",
        "cuota": cuota,
        "datosCliente": {"nombreCliente": "CLIENTE PRUEBA"},
        "datosSaldos": {"saldoTotalVencido": saldo_vencido},
        "datosNotasCargos": [],
        "datosCargos": [
            {
                "idCargo": 100,
                "concepto": "CUOTA SEMANAL",
                "monto": cuota,
                "fechaVencimiento": lunes_objetivo.isoformat(),
            },
            {
                "idCargo": 101,
                "concepto": "CUOTA SEMANAL",
                "monto": cuota,
                "fechaVencimiento": (lunes_objetivo + timedelta(days=7)).isoformat(),
            },
        ],
        "datosPagos": [
            {
                "idPago": 9000 + idx,
                "fechaRegistro": f"2026-07-25 0{idx}:00:00",
                "fechaValor": fecha_valor.isoformat(),
                "montoPago": monto + monto_extemporaneo,
                "extemporaneos": monto_extemporaneo,
                "numeroCuotaSemanal": "100,101",
            }
            for idx, (monto, monto_extemporaneo) in enumerate(
                zip(montos, montos_extemporaneos),
                start=1,
            )
        ],
    }


class LunesCuotaObjetivoTest(unittest.TestCase):
    def test_viernes_sabado_domingo_protegen_lunes_siguiente(self) -> None:
        esperado = date(2026, 7, 27)
        for fecha_pago in (date(2026, 7, 24), date(2026, 7, 25), date(2026, 7, 26)):
            with self.subTest(fecha_pago=fecha_pago):
                self.assertEqual(rc.lunes_cuota_objetivo_para_pago(fecha_pago), esperado)

    def test_lunes_a_jueves_protegen_lunes_de_la_misma_semana(self) -> None:
        esperado = date(2026, 7, 27)
        for fecha_pago in (
            date(2026, 7, 27),
            date(2026, 7, 28),
            date(2026, 7, 29),
            date(2026, 7, 30),
        ):
            with self.subTest(fecha_pago=fecha_pago):
                self.assertEqual(rc.lunes_cuota_objetivo_para_pago(fecha_pago), esperado)

    def test_viernes_cambia_al_lunes_de_la_semana_nueva(self) -> None:
        self.assertEqual(
            rc.lunes_cuota_objetivo_para_pago(date(2026, 7, 31)),
            date(2026, 8, 3),
        )


class SaldoFinSemanaTest(unittest.TestCase):
    def test_cuota_incompleta_no_produce_saldo(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[550])
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 0.0)

    def test_cuota_exacta_no_produce_saldo(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[551])
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 0.0)

    def test_pago_750_deja_solo_199_despues_de_cuota(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[750])
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 199.0)

    def test_pago_1279_deja_728_despues_de_cuota(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[1279])
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 728.0)

    def test_pagos_acumulados_cubren_cuota_y_dejan_solo_remanente(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[300, 400])
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 149.0)

    def test_saldo_vencido_bloquea_aplicacion(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(
            lunes_objetivo=lunes,
            montos=[1279],
            saldo_vencido=2,
        )
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 0.0)

    def test_cargo_del_lunes_objetivo_ausente_bloquea_aplicacion(self) -> None:
        lunes = date(2026, 7, 27)
        ec = estado_cuenta_con_pagos(lunes_objetivo=lunes, montos=[1279])
        ec["datosCargos"][0]["fechaVencimiento"] = "2026-07-20"
        self.assertEqual(rc.calcular_saldo_a_favor(ec, lunes), 0.0)

    def test_reporte_domingo_marca_aplicar_tras_cubrir_lunes(self) -> None:
        fecha_pago = date(2026, 7, 25)  # sábado
        lunes = rc.lunes_cuota_objetivo_para_pago(fecha_pago)
        ec = estado_cuenta_con_pagos(
            lunes_objetivo=lunes,
            montos=[1279],
            fecha_valor=fecha_pago,
            extemporaneos=[728],
        )
        row = {
            "id_credito": 1737107,
            "id_cliente": 985311,
            "valor_real": 750.0,
            "fecha_ultimo_pago_efectivo": fecha_pago.isoformat(),
            "monto_abono_efectivo": -1279.0,
        }
        with patch.object(rc, "consultar_s2", return_value={"estadoCuenta": ec}):
            resultado = rc.procesar_registro(
                row,
                fecha_pago.isoformat(),
                lunes,
                fecha_pago,
                date(2026, 7, 26),  # ejecución domingo
            )

        self.assertIsNotNone(resultado)
        self.assertEqual(resultado["SALDO_A_FAVOR"], 728.0)
        self.assertEqual(resultado["SALDO_APLICABLE_GC"], 728.0)
        self.assertIn(rc.COMENTARIO_APLICAR, resultado["COMENTARIOS"])
        self.assertNotIn(rc.COMENTARIO_CUOTA_CUBIERTA, resultado["COMENTARIOS"])


class ExtemporaneosFechaValorTest(unittest.TestCase):
    def test_suma_todos_los_pagos_de_la_fecha_valor(self) -> None:
        ec = {
            "datosPagos": [
                {
                    "idPago": 1,
                    "fechaValor": "2026-07-28T08:00:00",
                    "fechaRegistro": "2026-07-29",
                    "extemporaneos": "125.50",
                },
                {
                    "idPago": 2,
                    "fechaValor": "2026-07-28",
                    "fechaRegistro": "2026-07-27",
                    "extemporaneos": 124.50,
                },
                {
                    "idPago": 3,
                    "fechaValor": "2026-07-27",
                    "extemporaneos": 999,
                },
            ]
        }
        metricas = rc.metricas_extemporaneos_pago_fecha_valor(
            ec,
            date(2026, 7, 28),
        )
        self.assertEqual(metricas["total"], 250.0)
        self.assertEqual(metricas["pagos_fecha"], 2)

    def test_no_usa_fecha_registro_como_sustituto(self) -> None:
        ec = {
            "datosPagos": [
                {
                    "idPago": 1,
                    "fechaRegistro": "2026-07-28",
                    "fechaValor": "2026-07-27",
                    "extemporaneos": 250,
                }
            ]
        }
        self.assertEqual(
            rc.sumar_extemporaneos_pago_fecha_valor(
                ec,
                date(2026, 7, 28),
            ),
            0.0,
        )

    def test_deduplica_id_pago_y_reporta_la_anomalia(self) -> None:
        pago = {
            "idPago": 77,
            "fechaValor": "2026-07-28",
            "extemporaneos": 250,
        }
        metricas = rc.metricas_extemporaneos_pago_fecha_valor(
            {"datosPagos": [pago, deepcopy(pago)]},
            date(2026, 7, 28),
        )
        self.assertEqual(metricas["total"], 250.0)
        self.assertEqual(metricas["ids_duplicados"], 1)


class ReglaGeneralCarteraTest(unittest.TestCase):
    FECHA_NEGOCIO = date(2026, 7, 27)
    LUNES = date(2026, 7, 27)

    @classmethod
    def estado_cuenta(cls, extemporaneos: float) -> dict:
        return estado_cuenta_con_pagos(
            lunes_objetivo=cls.LUNES,
            montos=[801.0],
            fecha_valor=cls.FECHA_NEGOCIO,
            extemporaneos=[extemporaneos],
        )

    @classmethod
    def fila(cls, numero: int) -> dict:
        return {
            "id_credito": 9_100_000 + numero,
            "id_cliente": 8_100_000 + numero,
            "valor_real": 250.0,
            "fecha_ultimo_pago_efectivo": cls.FECHA_NEGOCIO.isoformat(),
            "monto_abono_efectivo": -1051.0,
        }

    def procesar(self, numero: int, extemporaneos: float) -> dict:
        with patch.object(
            rc,
            "consultar_s2",
            return_value={"estadoCuenta": self.estado_cuenta(extemporaneos)},
        ):
            resultado = rc.procesar_registro(
                self.fila(numero),
                self.FECHA_NEGOCIO.isoformat(),
                self.LUNES,
                self.FECHA_NEGOCIO,
                date(2026, 7, 28),
            )
        self.assertIsNotNone(resultado)
        return resultado

    def test_monto_s2_igual_al_aplicable_continua(self) -> None:
        resultado = self.procesar(1, 250.0)
        self.assertEqual(resultado["_EXTEMPORANEOS_S2_DIA"], 250.0)
        self.assertEqual(resultado["_MOTIVO_REVISION_CARTERA"], "")
        self.assertEqual(resultado["SALDO_APLICABLE_GC"], 250.0)

    def test_sin_extemporaneos_s2_se_detiene_para_revision(self) -> None:
        resultado = self.procesar(2, 0.0)
        self.assertTrue(
            resultado["_MOTIVO_REVISION_CARTERA"].startswith(
                "SIN_EXTEMPORANEOS_S2_DIA:"
            )
        )

    def test_extemporaneos_parciales_se_detienen_para_revision(self) -> None:
        resultado = self.procesar(3, 200.0)
        self.assertTrue(
            resultado["_MOTIVO_REVISION_CARTERA"].startswith(
                "EXTEMPORANEOS_INSUFICIENTES_S2_DIA:"
            )
        )

    def test_fecha_valor_ausente_falla_cerrado(self) -> None:
        ec = self.estado_cuenta(250.0)
        ec["datosPagos"][0]["fechaValor"] = ""
        with patch.object(
            rc,
            "consultar_s2",
            return_value={"estadoCuenta": ec},
        ):
            resultado = rc.procesar_registro(
                self.fila(4),
                self.FECHA_NEGOCIO.isoformat(),
                self.LUNES,
                self.FECHA_NEGOCIO,
                date(2026, 7, 28),
            )
        self.assertIsNotNone(resultado)
        self.assertTrue(
            resultado["_MOTIVO_REVISION_CARTERA"].startswith(
                "FECHA_VALOR_S2_NO_DISPONIBLE:"
            )
        )

    def test_simulacion_413_con_ids_nuevos_conserva_364_y_detiene_49(self) -> None:
        registros = [
            self.procesar(
                numero,
                0.0 if numero <= 49 else 250.0,
            )
            for numero in range(1, 414)
        ]
        elegibles, revision, stats = rc.separar_registros_revision_cartera(
            registros
        )

        self.assertEqual(stats["inicial"], 413)
        self.assertEqual(stats["elegibles"], 364)
        self.assertEqual(stats["revision"], 49)
        self.assertTrue(
            all(
                reg["_MOTIVO_REVISION_CARTERA"].startswith(
                    "SIN_EXTEMPORANEOS_S2_DIA:"
                )
                for reg in revision
            )
        )

        salida, stats_pipeline = rc.aplicar_pipeline_final_excel_gc(
            elegibles,
            self.FECHA_NEGOCIO,
        )
        self.assertEqual(len(salida), 364)
        self.assertEqual(stats_pipeline["final"], 364)

    def test_excel_muestra_validacion_s2_en_hoja_revision(self) -> None:
        revision = [self.procesar(numero, 0.0) for numero in range(1, 50)]
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "reporte_gc_prueba.xlsx"
            rc.generar_excel(
                [],
                self.FECHA_NEGOCIO,
                self.FECHA_NEGOCIO,
                str(ruta),
                fecha_generacion_cdmx=date(2026, 7, 28),
                inicio_semana_operativa=date(2026, 7, 28),
                registros_revision=revision,
            )
            wb = load_workbook(ruta, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    ["Reporte Cobranza", "Requiere revisión"],
                )
                encabezados = [
                    cell.value for cell in wb["Requiere revisión"][2]
                ]
                self.assertIn("EXTEMPORANEOS S2 FECHA VALOR", encabezados)
                self.assertIn("PAGOS S2 FECHA VALOR", encabezados)
                self.assertIn("MOTIVO DE REVISION", encabezados)
                self.assertEqual(wb["Requiere revisión"].max_row, 51)
            finally:
                wb.close()

    def test_excel_sin_revision_conserva_una_sola_hoja(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "reporte_gc_sin_revision.xlsx"
            rc.generar_excel(
                [registro_cartera(caso="C01")],
                self.FECHA_NEGOCIO,
                self.FECHA_NEGOCIO,
                str(ruta),
                fecha_generacion_cdmx=date(2026, 7, 29),
                inicio_semana_operativa=date(2026, 7, 28),
                registros_revision=[],
            )
            wb = load_workbook(ruta, read_only=True, data_only=True)
            try:
                self.assertEqual(wb.sheetnames, ["Reporte Cobranza"])
            finally:
                wb.close()


class DescargoValidacionS2Test(unittest.TestCase):
    @staticmethod
    def fila_descargo(numero: int = 1) -> dict:
        return {
            "id_credito": 9_300_000 + numero,
            "Id_cliente": 8_300_000 + numero,
            "nombre": "CLIENTE DESCARGO",
            "tipo_reporte": "Manual",
            "monto_aplicar": 250.0,
            "ultimo_pago_efectivo": "2026-07-27",
            "mensaje": "",
        }

    @staticmethod
    def metricas_s2(extemporaneos: float) -> dict:
        return {
            "s2_ok": True,
            "id_cliente": 8_300_001,
            "nombre_cliente": "CLIENTE DESCARGO",
            "status_credito": "Vigente",
            "cuota_semanal": 551.0,
            "deuda_gc": 250.0,
            "saldo_a_favor": 250.0,
            "metricas_extemporaneos": {
                "total": extemporaneos,
                "fecha_valor_disponible": True,
                "fechas_validas": 1,
                "pagos_fecha": 1,
                "montos_invalidos": 0,
                "ids_duplicados": 0,
            },
        }

    def fusionar(self, extemporaneos: float) -> dict:
        with patch.object(
            rc,
            "_obtener_deuda_gc_bd_por_ids",
            return_value={},
        ), patch.object(
            rc,
            "_obtener_monto_abono_efectivo_por_ids",
            return_value={},
        ), patch.object(
            rc,
            "_metricas_descargo_desde_s2",
            return_value=self.metricas_s2(extemporaneos),
        ):
            resultados = rc.merge_descargo_en_reporte(
                [],
                [self.fila_descargo()],
                fecha_corte="2026-07-27",
                lunes=date(2026, 7, 27),
            )
        self.assertEqual(len(resultados), 1)
        return resultados[0]

    def test_descargo_valido_continua_con_regla_comun(self) -> None:
        registro = self.fusionar(250.0)
        elegibles, revision, _ = rc.separar_registros_revision_cartera(
            [registro]
        )
        self.assertEqual(len(elegibles), 1)
        self.assertEqual(revision, [])

    def test_descargo_sin_extemporaneos_no_reintroduce_credito(self) -> None:
        registro = self.fusionar(0.0)
        elegibles, revision, _ = rc.separar_registros_revision_cartera(
            [registro]
        )
        self.assertEqual(elegibles, [])
        self.assertEqual(len(revision), 1)
        self.assertTrue(
            revision[0]["_MOTIVO_REVISION_CARTERA"].startswith(
                "SIN_EXTEMPORANEOS_S2_DIA:"
            )
        )


class PipelineRegresionTest(unittest.TestCase):
    def setUp(self) -> None:
        self.base = {
            "ID_CREDITO": 1737107,
            "ID_CLIENTE": 985311,
            "NOMBRE_CLIENTE": "CLIENTE PRUEBA",
            "STATUS_CREDITO": "Vigente",
            "CUOTA_SEMANAL": 551.0,
            "DEUDA_GC": 750.0,
            "SALDO_A_FAVOR": 728.0,
            "SALDO_APLICABLE_GC": 728.0,
            "FECHA_ULTIMO_ABONO_EFECTIVO": "2026-07-25 — sábado",
            "ULTIMO_ABONO_EFECTIVO": -1279.0,
            "MAXI_APP_CONECTO": "Sí",
            "COMENTARIOS": rc.COMENTARIO_APLICAR,
            "ERROR": "",
        }

    def test_sobrante_sabado_permanece_en_excel_domingo(self) -> None:
        salida, stats = rc.aplicar_pipeline_final_excel_gc(
            [deepcopy(self.base)],
            date(2026, 7, 25),
        )
        self.assertEqual(stats["final"], 1)
        self.assertEqual(salida[0]["SALDO_APLICABLE_GC"], 728.0)

    def test_regla_maxi_app_no_se_modifica(self) -> None:
        fila = deepcopy(self.base)
        fila["MAXI_APP_CONECTO"] = "No"
        salida, stats = rc.aplicar_pipeline_final_excel_gc([fila], date(2026, 7, 25))
        self.assertEqual(salida, [])
        self.assertEqual(stats["excl_no_rango"], 1)

    def test_limite_por_ultimo_abono_no_se_modifica(self) -> None:
        fila = deepcopy(self.base)
        fila["ULTIMO_ABONO_EFECTIVO"] = -560.0
        salida, stats = rc.aplicar_pipeline_final_excel_gc([fila], date(2026, 7, 25))
        self.assertEqual(stats["ajuste_no_alcanza"], 1)
        self.assertEqual(salida[0]["SALDO_APLICABLE_GC"], 560.0)

    def test_minimo_estricto_mayor_200_no_se_modifica(self) -> None:
        fila = deepcopy(self.base)
        fila["SALDO_A_FAVOR"] = 200.0
        fila["SALDO_APLICABLE_GC"] = 200.0
        salida, stats = rc.aplicar_pipeline_final_excel_gc([fila], date(2026, 7, 25))
        self.assertEqual(salida, [])
        self.assertEqual(stats["excl_saldo_final"], 1)

    def test_marca_legacy_no_aplicar_sigue_excluida(self) -> None:
        fila = deepcopy(self.base)
        fila["COMENTARIOS"] = rc.COMENTARIO_CUOTA_CUBIERTA
        salida, stats = rc.aplicar_pipeline_final_excel_gc([fila], date(2026, 7, 25))
        self.assertEqual(salida, [])
        self.assertEqual(stats["excl_cuota_siguiente_cubierta"], 1)


if __name__ == "__main__":
    unittest.main()
