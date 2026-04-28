#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Descarga incremental desde __SPARTA_SECRET_REDACTED__: filas estatus=3,
genera descargo_estatus3.xlsx (columnas legibles y encabezados formateados) y guia_descargo.json.

Repo: backend/services/gastos-cobranza-agent/scripts/
Agente Node: POST /descargo-estatus3/run (carpeta fija reporte/descargo_estatus3/).

Dependencias: mysql-connector-python, openpyxl, tzdata (ver scripts/requirements.txt).

Ejemplo:
  py -3 descargo_cobranza_gc_estatus3.py --mega-php-defaults --datos-dir "C:\\ruta\\descargo_estatus3"
  py -3 descargo_cobranza_gc_estatus3.py --mega-php-defaults --desde-cero --datos-dir ...
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import math

# Blindaje de dependencias locales: permite importar paquetes instalados en
# <raiz-del-agente>/pydeps aunque el servicio corra con otro usuario/perfil.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_AGENT_ROOT = os.path.normpath(os.path.join(_SCRIPT_DIR, ".."))
_PYDEPS_DIR = os.path.join(_AGENT_ROOT, "pydeps")
if os.path.isdir(_PYDEPS_DIR) and _PYDEPS_DIR not in sys.path:
    sys.path.insert(0, _PYDEPS_DIR)

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MEGA_PHP_DEFAULTS: dict = {
    "host": "__SPARTA_HOST_REDACTED__",
    "port": 3306,
    "database": "__SPARTA_SECRET_REDACTED__",
    "user": "__SPARTA_SECRET_REDACTED__",
    "password": "__SPARTA_PASSWORD_REDACTED__",
}

TABLE_NAME = "cobranza_gc_verificacion_semana"
_DATOS_SUBDIR = "cobranza_gc_estatus3_datos"
_GUIA_LEGACY_CHECKPOINT = "descargo_checkpoint.json"
_DESCARGO_XLSX_BASENAME = "descargo_estatus3.xlsx"


def _guia_descargo_basename() -> str:
    """Mismo criterio que reporte_cobranza.py (checkpoint distinto en localhost vs servidor)."""
    raw = (os.environ.get("REPORTE_COBRANZA_DESCARGO_GUIA_BASENAME") or "").strip()
    if not raw:
        return "guia_descargo.json"
    if "/" in raw or "\\" in raw or ".." in raw or raw.startswith("."):
        return "guia_descargo.json"
    if not raw.endswith(".json"):
        return "guia_descargo.json"
    return raw


def carpeta_datos_descargo(ruta_explicita: Optional[str]) -> Path:
    if ruta_explicita and str(ruta_explicita).strip():
        return Path(str(ruta_explicita).strip()).expanduser().resolve()
    base = Path(__file__).resolve().parent
    if base.name == _DATOS_SUBDIR:
        return base
    return base / _DATOS_SUBDIR


@dataclass
class GuiaDescargo:
    registrado_en_cdmx: str
    id_tabla: int
    id_credito: Optional[int] = None


def _datetime_a_iso_cdmx(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d 00:00:00")
    s = str(val).strip()
    if len(s) >= 19 and s[4] == "-" and s[10] in " T":
        s = s[:19].replace("T", " ")
    return s


def leer_guia_desde_archivo(path: Path) -> Optional[GuiaDescargo]:
    if not path.is_file():
        return None
    try:
        with open(path, encoding="utf-8-sig") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        print(f"Aviso: no se pudo leer {path} ({e}).", file=sys.stderr)
        return None
    reg_raw = data.get("ultimo_registrado_en_cdmx")
    if reg_raw is None or not str(reg_raw).strip():
        return None
    reg = _datetime_a_iso_cdmx(reg_raw)
    if not reg:
        return None

    pk_raw = data.get("ultimo_id_tabla")
    legacy = data.get("ultimo_id")
    ic_raw = data.get("ultimo_id_credito")

    pk_i: Optional[int] = None
    if pk_raw is not None and str(pk_raw).strip() != "":
        try:
            pk_i = int(pk_raw)
        except (TypeError, ValueError):
            print(f"Aviso: ultimo_id_tabla inválido en {path}: {pk_raw!r}", file=sys.stderr)
            return None
    elif legacy is not None and str(legacy).strip() != "":
        try:
            pk_i = int(legacy)
        except (TypeError, ValueError):
            print(f"Aviso: ultimo_id (legado) inválido en {path}: {legacy!r}", file=sys.stderr)
            return None

    if pk_i is None:
        return None

    ic_opt: Optional[int] = None
    if ic_raw is not None and str(ic_raw).strip() != "":
        try:
            ic_opt = int(ic_raw)
        except (TypeError, ValueError):
            print(f"Aviso: ultimo_id_credito inválido en {path}: {ic_raw!r}", file=sys.stderr)
            return None

    return GuiaDescargo(reg, pk_i, ic_opt)


def cargar_lectura_guia(datos_dir: Path) -> Optional[GuiaDescargo]:
    base = _guia_descargo_basename()
    candidatos = [
        datos_dir / base,
        datos_dir / (base + ".bak"),
        datos_dir / _GUIA_LEGACY_CHECKPOINT,
    ]
    vistos: set[Path] = set()
    for p in candidatos:
        try:
            clave = p.resolve()
        except OSError:
            clave = p
        if clave in vistos:
            continue
        vistos.add(clave)
        g = leer_guia_desde_archivo(p)
        if g is not None:
            if p.suffix == ".bak":
                print(f"Nota: checkpoint leído desde respaldo {p.name}.", file=sys.stderr)
            return g
    return None


def col_index_ci(colnames: list[str], want: str) -> int:
    w = want.lower()
    for i, c in enumerate(colnames):
        if str(c).lower() == w:
            return i
    print(
        f"No se encontró la columna {want!r}. Columnas: {colnames}",
        file=sys.stderr,
    )
    sys.exit(2)


def _fila_como_mapa_lc(row: tuple, colnames: list[str]) -> dict:
    """Claves en minúsculas para coincidir columnas BD sin importar mayúsculas."""
    return {str(colnames[i]).lower(): row[i] for i in range(len(row))}


def _valor_presente(v) -> bool:
    if v is None:
        return False
    try:
        if isinstance(v, float) and math.isnan(v):
            return False
    except TypeError:
        pass
    if isinstance(v, str) and v.strip() == "":
        return False
    return True


def _fmt_semana_iso(anio_val, semana_val) -> str:
    try:
        if _valor_presente(anio_val) and _valor_presente(semana_val):
            return f"{int(anio_val)}-W{int(semana_val):02d}"
        if _valor_presente(semana_val):
            return str(int(semana_val))
    except (TypeError, ValueError):
        pass
    return ""


def _ids_credito_desde_rows(rows: list, colnames: list[str]) -> list[int]:
    """id_credito únicos de las filas del descargo (estatus=3)."""
    ix = col_index_ci(colnames, "id_credito")
    out: set[int] = set()
    for row in rows:
        try:
            v = row[ix]
            if v is None or (isinstance(v, str) and not str(v).strip()):
                continue
            out.add(int(v))
        except (TypeError, ValueError, IndexError):
            continue
    return sorted(out)


def _fetch_monto_abono_efectivo_por_ids(cur, ids: list[int]) -> dict[int, Optional[float]]:
    """
    Monto_abono_efectivo por Id_credito en __SPARTA_SECRET_REDACTED__.tbl_segundometro_semana (Dias_mora=0).
    """
    result: dict[int, Optional[float]] = {}
    if not ids:
        return result
    chunk = 500
    for i in range(0, len(ids), chunk):
        part = ids[i : i + chunk]
        ph = ",".join(["%s"] * len(part))
        sql = (
            f"SELECT `Id_credito` AS id_credito, MAX(`Monto_abono_efectivo`) AS monto "
            f"FROM `tbl_segundometro_semana` WHERE `Dias_mora` = 0 AND `Id_credito` IN ({ph}) "
            f"GROUP BY `Id_credito`"
        )
        cur.execute(sql, tuple(part))
        for r in cur.fetchall():
            try:
                ic = int(r[0])
            except (TypeError, ValueError):
                continue
            if r[1] is None:
                result[ic] = None
            else:
                try:
                    result[ic] = float(r[1])
                except (TypeError, ValueError):
                    result[ic] = None
    return result


def construir_tabla_descargo_export(
    rows: list,
    colnames: list[str],
    monto_abono_por_id: Optional[dict[int, Optional[float]]] = None,
) -> tuple[list[str], list[list]]:
    """
    Encabezados legibles + filas para el Excel (misma salida que antes vía pandas).
    `mensaje` en BD → Observaciones.
    """
    headers = [
        "ID crédito",
        "Nombre",
        "Tipo de reporte",
        "Monto a aplicar",
        "Estatus",
        "Id usuario que reportó",
        "Semana ISO",
        "registrado_en_cdmx",
        "Último pago efectivo",
        "ultimo_abono_efectivo",
        "Observaciones",
    ]
    monto_abono_por_id = monto_abono_por_id or {}
    data: list[list] = []
    for row in rows:
        m = _fila_como_mapa_lc(tuple(row), colnames)
        anio = m.get("anio_iso")
        sem = m.get("semana_iso")
        sem_txt = _fmt_semana_iso(anio, sem)
        reg = m.get("registrado_en_cdmx")
        reg_txt = _datetime_a_iso_cdmx(reg) if _valor_presente(reg) else ""
        upe = m.get("ultimo_pago_efectivo")
        upe_txt = _datetime_a_iso_cdmx(upe) if _valor_presente(upe) else ""
        monto_ae = None
        try:
            ic_raw = m.get("id_credito")
            if ic_raw is not None and str(ic_raw).strip() != "":
                ic_int = int(ic_raw)
                if ic_int in monto_abono_por_id:
                    mv = monto_abono_por_id[ic_int]
                    if mv is not None and _valor_presente(mv):
                        monto_ae = float(mv)
        except (TypeError, ValueError):
            monto_ae = None
        fila_out = [
            m.get("id_credito"),
            m.get("nombre"),
            m.get("tipo_reporte"),
            m.get("monto_aplicar"),
            m.get("estatus"),
            m.get("id_usuario_reporte"),
            sem_txt,
            reg_txt,
            upe_txt,
            monto_ae,
            m.get("mensaje"),
        ]
        data.append(fila_out)
    return headers, data


def escribir_excel_descargo_formateado(encabezados: list[str], data_rows: list[list], path: Path) -> None:
    """Libro .xlsx con encabezados con color, bordes, anchos y filtros."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Descargo estatus 3"

    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B4B4B4")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_top = Alignment(vertical="top", wrap_text=True)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_num = Alignment(horizontal="right", vertical="top")

    headers = list(encabezados)
    num_cols = len(headers)
    idx_monto = headers.index("Monto a aplicar") + 1 if "Monto a aplicar" in headers else None
    idx_ult_ab = headers.index("ultimo_abono_efectivo") + 1 if "ultimo_abono_efectivo" in headers else None
    idx_id_cred = headers.index("ID crédito") + 1 if "ID crédito" in headers else None
    idx_est = headers.index("Estatus") + 1 if "Estatus" in headers else None
    idx_usr = headers.index("Id usuario que reportó") + 1 if "Id usuario que reportó" in headers else None

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = cell_border

    monto_fmt = '#,##0.00_ ;-#,##0.00_ ;"-"_ '

    for i, row in enumerate(data_rows, start=2):
        for j, val in enumerate(row, start=1):
            v = val
            if not _valor_presente(v):
                v = None
            c = ws.cell(row=i, column=j, value=v)
            c.border = cell_border
            if j == idx_monto and v is not None:
                try:
                    c.value = float(v)
                    c.number_format = monto_fmt
                    c.alignment = align_num
                except (TypeError, ValueError):
                    c.alignment = align_top
            elif j == idx_ult_ab and v is not None:
                try:
                    c.value = float(v)
                    c.number_format = monto_fmt
                    c.alignment = align_num
                except (TypeError, ValueError):
                    c.alignment = align_top
            elif j in (idx_id_cred, idx_est, idx_usr) and v is not None and j != idx_monto:
                try:
                    if isinstance(v, str) and v.strip() == "":
                        c.value = None
                    else:
                        c.value = int(float(v))
                        c.alignment = Alignment(horizontal="center", vertical="top")
                except (TypeError, ValueError):
                    c.alignment = align_top
            elif j == num_cols:
                c.alignment = align_top
            else:
                c.alignment = align_top

    ws.freeze_panes = "A2"
    last_row = 1 + len(data_rows)
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_row}"

    # Anchos sugeridos (máx. ~55 en observaciones)
    default_widths = {
        "ID crédito": 14,
        "Nombre": 30,
        "Tipo de reporte": 18,
        "Monto a aplicar": 16,
        "Estatus": 10,
        "Id usuario que reportó": 20,
        "Semana ISO": 14,
        "registrado_en_cdmx": 22,
        "Último pago efectivo": 22,
        "ultimo_abono_efectivo": 20,
        "Observaciones": 48,
    }
    for j, h in enumerate(headers, start=1):
        letter = get_column_letter(j)
        w = min(default_widths.get(h, 16), 55)
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 28
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _ahora_cdmx_para_guia() -> str:
    """Marca al escribir la guía (CDMX), para comprobar en UI si hubo corrida reciente."""
    try:
        return datetime.now(ZoneInfo("America/Mexico_City")).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def guardar_guia_descargo(
    path: Path,
    *,
    ultimo_id_credito: int,
    ultimo_id_tabla: int,
    ultimo_registrado_en_cdmx: str,
    table: str,
) -> None:
    payload = {
        "ultimo_id_tabla": int(ultimo_id_tabla),
        "ultimo_registrado_en_cdmx": str(ultimo_registrado_en_cdmx).strip(),
        "ultimo_id_credito": int(ultimo_id_credito),
        "tabla": table,
        "orden_descargo": "registrado_en_cdmx ASC, id ASC",
        "guia_escrita_en_cdmx": _ahora_cdmx_para_guia(),
        "nota": "Checkpoint incremental: la siguiente corrida (sin «desde cero») pide filas con "
        "(registrado_en_cdmx > ultimo_registrado_en_cdmx) O (misma fecha y id de tabla > ultimo_id_tabla). "
        "ultimo_id_credito es solo informativo. guia_escrita_en_cdmx = cuándo se guardó este JSON en el agente.",
    }
    try:
        text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
        json.loads(text)
    except (TypeError, ValueError, json.JSONDecodeError) as e:
        print(f"No se pudo serializar la guía: {e}", file=sys.stderr)
        sys.exit(2)
    tmp = path.with_suffix(path.suffix + ".tmp")
    bak = path.with_suffix(path.suffix + ".bak")
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.is_file() and path.stat().st_size > 0:
            try:
                shutil.copy2(path, bak)
            except OSError as e2:
                print(f"Aviso: no se pudo respaldar guía anterior: {e2}", file=sys.stderr)
        with open(tmp, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
        os.replace(str(tmp), str(path))
    except OSError as e:
        print(f"No se pudo guardar la guía {path}: {e}", file=sys.stderr)
        try:
            if tmp.is_file():
                tmp.unlink()
        except OSError:
            pass
        sys.exit(2)


def descargo_estatus_3_incremental(
    cur,
    *,
    datos_dir: Path,
    desde_cero: bool,
    sin_actualizar_guia: bool = False,
    table: str = TABLE_NAME,
) -> None:
    guia_escritura = datos_dir / _guia_descargo_basename()
    xlsx_path = datos_dir / _DESCARGO_XLSX_BASENAME
    if sin_actualizar_guia:
        print(
            "Modo --sin-actualizar-guía: si hay filas, se genera el Excel pero no se escribe guia_descargo.json.",
            flush=True,
        )
    conds = ["`estatus` = 3"]
    exec_params: list = []

    if not desde_cero:
        guia = cargar_lectura_guia(datos_dir)
        if guia is not None:
            lr = guia.registrado_en_cdmx
            ic_txt = (
                f" (id_credito de esa fila: {guia.id_credito})"
                if guia.id_credito is not None
                else ""
            )
            print(
                f"Vale: me quedé por aquí. Última fila ya procesada: id tabla = {guia.id_tabla}, "
                f"fecha/hora en BD = {lr}{ic_txt}. "
                "Sigo después de ese punto (misma fecha/hora solo filas con id mayor).",
                flush=True,
            )
            conds.append(
                "(`registrado_en_cdmx` > %s OR "
                "(`registrado_en_cdmx` = %s AND `id` > %s))"
            )
            exec_params.extend([lr, lr, guia.id_tabla])
        else:
            print(
                "No hay guía previa: esta vez saco todo lo que cumpla el filtro hasta ahora.",
                flush=True,
            )
    else:
        if sin_actualizar_guia:
            print(
                "Opción --desde-cero + --sin-actualizar-guía: ignoro la guía y pido todo el filtro; "
                "al terminar NO escribo guia_descargo.json (solo prueba).",
                flush=True,
            )
        else:
            print(
                "Opción --desde-cero: ignoro la guía, pido todo el filtro otra vez "
                "y al termino actualizo la guía con el último registro.",
                flush=True,
            )

    sql = (
        f"SELECT * FROM `{table}` WHERE {' AND '.join(conds)} "
        "ORDER BY `registrado_en_cdmx` ASC, `id` ASC"
    )
    cur.execute(sql, tuple(exec_params))
    colnames = [d[0] for d in (cur.description or [])]
    rows = cur.fetchall()
    print(f"Filas nuevas en esta corrida (estatus=3): {len(rows)}", flush=True)

    if rows:
        try:
            datos_dir.mkdir(parents=True, exist_ok=True)
            ids_sm = _ids_credito_desde_rows(rows, colnames)
            monto_map = _fetch_monto_abono_efectivo_por_ids(cur, ids_sm)
            hdrs, data_export = construir_tabla_descargo_export(rows, colnames, monto_map)
            escribir_excel_descargo_formateado(hdrs, data_export, xlsx_path)
        except OSError as e:
            print(f"No se pudo escribir el Excel ({xlsx_path}): {e}", file=sys.stderr)
            sys.exit(2)
        except Exception as e:
            print(f"Error al generar Excel: {e}", file=sys.stderr)
            sys.exit(2)
        ix_pk = col_index_ci(colnames, "id")
        ix_ic = col_index_ci(colnames, "id_credito")
        ix_reg = col_index_ci(colnames, "registrado_en_cdmx")
        ult = rows[-1]
        pk = int(ult[ix_pk])
        ic = int(ult[ix_ic])
        reg_s = _datetime_a_iso_cdmx(ult[ix_reg])
        if sin_actualizar_guia:
            print(
                f"Modo prueba (--sin-actualizar-guía): NO guardé cambios en {guia_escritura.name}.",
                flush=True,
            )
        else:
            guardar_guia_descargo(
                guia_escritura,
                ultimo_id_credito=ic,
                ultimo_id_tabla=pk,
                ultimo_registrado_en_cdmx=reg_s,
                table=table,
            )
            print(
                f"Listo: última fila procesada — id tabla {pk}, fecha/hora BD {reg_s}, "
                f"id_credito {ic} (referencia). Guía actualizada.",
                flush=True,
            )
        print(f"Excel generado: {xlsx_path}", flush=True)
    else:
        print(
            "No hubo filas nuevas: no toco la guía anterior ni sobreescribo el Excel con datos vacíos.",
            flush=True,
        )


def load_env_file(path: str, *, required: bool, override: bool = False) -> None:
    if not os.path.isfile(path):
        if required:
            print(f"No existe el archivo: {path}", file=sys.stderr)
            sys.exit(2)
        return
    try:
        with open(path, encoding="utf-8-sig") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, _, val = line.partition("=")
                key, val = key.strip(), val.strip()
                if len(val) >= 2 and val[0] == val[-1] and val[0] in "\"'":
                    val = val[1:-1]
                if key and (override or key not in os.environ):
                    os.environ[key] = val
    except OSError as e:
        print(f"No se pudo leer {path}: {e}", file=sys.stderr)
        sys.exit(2)


def env_first(*names: str) -> Optional[str]:
    for n in names:
        if n in os.environ:
            return os.environ[n]
    return None


def tiene_host_db_en_entorno() -> bool:
    return bool((env_first("MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR") or "").strip())


def db_config(*, mega_php_defaults: bool) -> dict:
    if mega_php_defaults:
        return {k: v for k, v in _MEGA_PHP_DEFAULTS.items()}

    host = env_first("MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR")
    user = env_first("MEGA_USER", "MYSQL_USER", "DB_USER", "DB_USUARIO")
    password = env_first("MEGA_PASSWORD", "MYSQL_PASSWORD", "DB_PASSWORD", "DB_PASS")
    database = env_first("MEGA_DATABASE", "MYSQL_DATABASE", "DB_NAME", "DB_ESQUEMA")
    port_s = env_first("MEGA_PORT", "MYSQL_PORT", "DB_PUERTO") or "3306"

    if password is None:
        password = ""
    if not (host or "").strip() or not (user or "").strip() or not (database or "").strip():
        print(
            "Faltan MEGA_HOST/MEGA_USER/MEGA_DATABASE (o MYSQL_*/DB_*). "
            "Quita --no-mega-php-defaults o define esas variables.",
            file=sys.stderr,
        )
        sys.exit(2)
    try:
        port_i = int(str(port_s).strip())
    except ValueError:
        print("Puerto debe ser entero.", file=sys.stderr)
        sys.exit(2)

    return {
        "host": host.strip(),
        "user": user.strip(),
        "password": password,
        "database": database.strip(),
        "port": port_i,
    }


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Descarga Excel estatus=3 + guía JSON (sin subir Excel a la BD).",
    )
    ap.add_argument(
        "--mega-php-defaults",
        action="store_true",
        help="Credenciales embebidas (DatabaseSegundometro.php) aunque exista MEGA_HOST.",
    )
    ap.add_argument(
        "--no-mega-php-defaults",
        action="store_true",
        help="Exige MEGA_HOST/MEGA_USER/MEGA_DATABASE (o MYSQL_*/DB_*).",
    )
    ap.add_argument("--env-file", default=None, help="Archivo KEY=valor UTF-8")
    ap.add_argument(
        "--env-file-overrides",
        action="store_true",
        help="Con --env-file, pisar variables ya definidas",
    )
    ap.add_argument("--test-db", action="store_true", help="Solo SELECT 1 y salir")
    ap.add_argument(
        "--datos-dir",
        default=None,
        metavar="CARPETA",
        help=f"Carpeta de guía y Excel (default: carpeta {_DATOS_SUBDIR!r} junto a este script).",
    )
    ap.add_argument(
        "--desde-cero",
        action="store_true",
        help="Ignora la guía, vuelca todo el filtro y al final guarda guía nueva.",
    )
    ap.add_argument(
        "--sin-actualizar-guia",
        action="store_true",
        help="Tras generar el Excel no escribe guia_descargo.json (pruebas sin avanzar el checkpoint).",
    )
    args = ap.parse_args()

    if args.env_file:
        load_env_file(args.env_file, required=True, override=bool(args.env_file_overrides))

    if args.mega_php_defaults and args.no_mega_php_defaults:
        print("No combines --mega-php-defaults con --no-mega-php-defaults.", file=sys.stderr)
        sys.exit(2)

    use_embedded = bool(args.mega_php_defaults)
    if not use_embedded and not args.no_mega_php_defaults and not tiene_host_db_en_entorno():
        use_embedded = True
        print(
            "Nota: sin MEGA_HOST en el entorno; usando credenciales embebidas.",
            file=sys.stderr,
        )
    cfg = db_config(mega_php_defaults=use_embedded)

    conn_kw = {
        "host": cfg["host"],
        "port": cfg["port"],
        "user": cfg["user"],
        "password": cfg["password"],
        "database": cfg["database"],
        "charset": "utf8mb4",
        "use_unicode": True,
    }

    try:
        conn = mysql.connector.connect(**conn_kw)
    except mysql.connector.Error as e:
        print(f"Error de conexión MySQL: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        cur = conn.cursor()
        if args.test_db:
            cur.execute("SELECT 1 AS ok")
            print(cur.fetchone())
            return

        datos_dir = carpeta_datos_descargo(args.datos_dir)
        try:
            datos_dir.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            print(f"No se pudo crear la carpeta {datos_dir}: {e}", file=sys.stderr)
            sys.exit(2)
        print(f"Carpeta de trabajo: {datos_dir}", flush=True)
        descargo_estatus_3_incremental(
            cur,
            datos_dir=datos_dir,
            desde_cero=bool(args.desde_cero),
            sin_actualizar_guia=bool(args.sin_actualizar_guia),
        )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
