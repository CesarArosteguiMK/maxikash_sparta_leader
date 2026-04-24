"""
REPORTE GASTOS DE COBRANZA + SALDO A FAVOR
==========================================
Columnas finales:
  ID CREDITO | ID CLIENTE | NOMBRE CLIENTE | STATUS CREDITO | CUOTA SEMANAL |
  DEUDA GC PENDIENTE | SALDO A FAVOR DEL CLIENTE | SALDO APLICABLE A GC |
  fecha_ultimo_abono_efectivo | ultimo_abono_efectivo (Monto_abono_efectivo, tbl_segundometro_semana) |
  MAXI APP (última conexión CDMX + ¿se conectó?)

Flujo:
  1. MySQL trae créditos elegibles (gastos_cobranza + tbl_segundometro_semana, Dias_mora=0,
     sin despacho, valor_real > 1) y Fecha_ultimo_pago_efectivo = **fecha de negocio** (ayer CDMX).
     Excluye id_credito con cobranza_gc_verificacion_semana estatus=3 en la misma inicio_semana (semana actual).
  2. Se consulta la lista negra (cobranza_gc_verificacion_semana) en una query aparte.
  3. En Python se descartan los id_credito que ya están en lista negra para esa semana
     (inicio_semana = martes del periodo, s2_exitoso = 1).
  4. Los que quedan van a S2 → cálculo → Excel.
  5. El Excel se genera siempre (aunque queden 0 registros después de los filtros).
  6. Reglas de salida del .xlsx (pipeline final):
     - No exportar filas con STATUS CREDITO = ERROR.
     - Conjunto NO (MAXI APP — ¿SE CONECTÓ? = NO): conservar solo SALDO APLICABLE A GC entre 200 y 300 (incluyente).
     - Reintegrar todas las filas SI + las NO válidas.
     - Validación interna (no visible en Excel): ABS(ultimo_abono_efectivo) >= SALDO APLICABLE A GC.
       Si no alcanza, SALDO APLICABLE A GC = ABS(ultimo_abono_efectivo).
     - Filtro final: SALDO APLICABLE A GC > 200 (sin tope superior).
  8. Columnas Maxi app (__SPARTA_SECRET_REDACTED__.ubicacion, idCliente = ID CLIENTE): ventana en calendario CDMX —
     vie–dom → lunes de esa semana hasta hoy CDMX; lun–jue → hoy CDMX y 4 días anteriores (5 días en total).
     Última conexión mostrada en hora CDMX (no depender del huso del servidor del script).
     ultimo_abono_efectivo va después de fecha_ultimo_abono_efectivo y antes de las columnas Maxi app.

En el .xlsx, el color de fila se calcula internamente por reglas de negocio.

Salida: reporte/reporte_cobranza_DD-MM-YYYY.xlsx (fecha calendario CDMX al generar; guiones en lugar de /).

Reintentos: tras la primera pasada S2 en paralelo, se pueden ejecutar pasadas extra solo sobre filas con
columna ERROR (timeout / sin respuesta S2). Un solo Excel final. Variables de entorno opcionales:
  REPORTE_COBRANZA_PASADAS_REINTENTO (default 2 = dos rondas de reintento tras la masiva; 0 = desactivar)
  REPORTE_COBRANZA_REINTENTO_MAX_WORKERS (default 16)
  REPORTE_COBRANZA_REINTENTO_PAUSA_S (default 1.5 segundos entre anuncio y reintento)

Conexión __SPARTA_SECRET_REDACTED__ (columnas Maxi app / tabla ubicacion); si no se definen, se usan los mismos defaults que DatabaseAWS:
  REPORTE_COBRANZA_AWS_HOST, REPORTE_COBRANZA_AWS_PORT (default 3306),
  REPORTE_COBRANZA_AWS_USER, REPORTE_COBRANZA_AWS_PASSWORD, REPORTE_COBRANZA_AWS_DATABASE

Pruebas / descargo:
  REPORTE_COBRANZA_NO_GUARDAR_GUIA_DESCARGO=1 — fusiona descargo igual pero NO escribe guia_descargo.json
  REPORTE_COBRANZA_DESCARGO_GUIA_BASENAME=guia_descargo.local.json — otro nombre de guía (p. ej. localhost vs servidor).
    (el checkpoint no avanza; otra corrida puede volver a traer las mismas filas incrementales).
  REPORTE_COBRANZA_SIN_DESCARGO=1 — omite por completo consulta y merge de descargo (no toca la guía).

Pruebas / Excel sin tocar el reporte oficial del día:
  REPORTE_COBRANZA_MODO_PRUEBA_EXCEL=1 — guarda reporte_cobranza_DD-MM-AAAA_PRUEBA.xlsx (no comprueba ni pisa el .xlsx oficial).
  REPORTE_COBRANZA_REGENERAR=1 — omite la salida temprana si ya existe el .xlsx del día (p. ej. tras renombrar copia desde el agente).

SALDO A FAVOR: lunes de la semana calendario (lun–dom) de la fecha de negocio.

Fecha de negocio: siempre **ayer** respecto al calendario **America/Mexico_City** (hora CDMX real),
para alinear con ejecución programada el día previo al día operativo a las 08:00 CDMX.
"""

from __future__ import annotations

import json
import logging
import os
import re
import shutil
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

import pymysql
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Carpeta compartida con el agente Node: excel semanales y log de ejecuciones Python.
REPORTE_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, "..", "reporte"))
os.makedirs(REPORTE_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(
            os.path.join(REPORTE_DIR, "reporte_cobranza_py.log"),
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────────
DB_CONFIG = {
    "host":            "__SPARTA_HOST_REDACTED__",
    "port":            3306,
    "user":            "__SPARTA_SECRET_REDACTED__",
    "password":        "__SPARTA_PASSWORD_REDACTED__",
    "database":        "__SPARTA_SECRET_REDACTED__",
    "charset":         "utf8mb4",
    "connect_timeout": 30,
    "read_timeout":    180,
    "write_timeout":   180,
    "cursorclass":     pymysql.cursors.DictCursor,
}

# Defaults alineados con Core\DatabaseAWS (__SPARTA_SECRET_REDACTED__). Sobreescribir vía REPORTE_COBRANZA_AWS_* (ver docstring).
_DEFAULT_AWS_MOVIL_HOST = "__SPARTA_HOST_REDACTED__"
_DEFAULT_AWS_MOVIL_USER = "__SPARTA_SECRET_REDACTED__"
_DEFAULT_AWS_MOVIL_PASSWORD = "__SPARTA_PASSWORD_REDACTED__"
_DEFAULT_AWS_MOVIL_DATABASE = "__SPARTA_SECRET_REDACTED__"

S2_URL     = "https://servicios.s2movil.net/s2__SPARTA_SECRET_REDACTED__/estadocuenta"
S2_TOKEN   = "__SPARTA_TOKEN_REDACTED__"
S2_HEADERS = {"Content-Type": "application/json", "Token": S2_TOKEN}

TZ_CDMX = ZoneInfo("America/Mexico_City")

# Google Chat (espacio Gastos Cobranza). Sobreescribir con GASTOS_COBRANZA_GCHAT_URL en el entorno.
GCHAT_URL = os.environ.get(
    "GASTOS_COBRANZA_GCHAT_URL",
    "__SPARTA_WEBHOOK_REDACTED__",
)

FECHA_CORTE_GASTOS = date(2026, 1, 28)
CONCEPTO_GC        = "NOTA DE DE CARGO GASTOS DE COBRANZA"

MAX_WORKERS    = 40
MAX_REINTENTOS = 3
TIMEOUT_S2     = 45
LOG_CADA       = 500
MAXI_APP_IN_CHUNK = 450  # tamaño de IN (...) por consulta a ubicacion (__SPARTA_SECRET_REDACTED__)


def _env_int(name: str, default: int) -> int:
    try:
        return max(0, int(str(os.environ.get(name, str(default))).strip()))
    except ValueError:
        return default


def _env_float(name: str, default: float) -> float:
    try:
        return float(str(os.environ.get(name, str(default))).strip())
    except ValueError:
        return default


def _env_str(name: str, default: str) -> str:
    raw = os.environ.get(name)
    if raw is None:
        return default
    s = str(raw).strip()
    return s if s else default


def _db_config_aws_movil() -> dict:
    port = _env_int("REPORTE_COBRANZA_AWS_PORT", 3306)
    if port < 1 or port > 65535:
        port = 3306
    return {
        "host":            _env_str("REPORTE_COBRANZA_AWS_HOST", _DEFAULT_AWS_MOVIL_HOST),
        "port":            port,
        "user":            _env_str("REPORTE_COBRANZA_AWS_USER", _DEFAULT_AWS_MOVIL_USER),
        "password":        _env_str("REPORTE_COBRANZA_AWS_PASSWORD", _DEFAULT_AWS_MOVIL_PASSWORD),
        "database":        _env_str("REPORTE_COBRANZA_AWS_DATABASE", _DEFAULT_AWS_MOVIL_DATABASE),
        "charset":         "utf8mb4",
        "connect_timeout": 30,
        "read_timeout":    120,
        "write_timeout":   120,
        "cursorclass":     pymysql.cursors.DictCursor,
    }


DB_CONFIG_AWS_MOVIL = _db_config_aws_movil()


# Rondas extra solo para filas con ERROR de S2 (timeout, sin respuesta, etc.). 0 = desactivado.
# Default 2: tercera “pasada” efectiva (1 masiva + 2 reintentos sobre quienes sigan con ERROR).
# Una sola Excel final con todo mezclado (aciertos + errores que sigan fallando).
PASADAS_REINTENTO_ERRORES = _env_int("REPORTE_COBRANZA_PASADAS_REINTENTO", 2)
REINTENTO_MAX_WORKERS = max(1, _env_int("REPORTE_COBRANZA_REINTENTO_MAX_WORKERS", 16))
REINTENTO_PAUSA_S = _env_float("REPORTE_COBRANZA_REINTENTO_PAUSA_S", 1.5)
# Pruebas: 1 = no escribir guia_descargo.json tras fusionar descargo.
NO_GUARDAR_GUIA_DESCARGO = str(os.environ.get("REPORTE_COBRANZA_NO_GUARDAR_GUIA_DESCARGO", "")).strip() == "1"
# Pruebas: 1 = no consultar ni fusionar descargo (no modifica guia_descargo.json).
SIN_DESCARGO = str(os.environ.get("REPORTE_COBRANZA_SIN_DESCARGO", "")).strip() == "1"
# Pruebas: 1 = nombre ..._PRUEBA.xlsx; no bloquea si ya existe el Excel oficial del día.
MODO_PRUEBA_EXCEL = str(os.environ.get("REPORTE_COBRANZA_MODO_PRUEBA_EXCEL", "")).strip() == "1"
REGENERAR_REPORTE = str(os.environ.get("REPORTE_COBRANZA_REGENERAR", "")).strip().lower() in ("1", "true", "yes", "si", "sí")
# Textos de negocio en COMENTARIOS (día laboral vs fin de semana / lunes).
# Color en Excel: CUOTA_CUBIERTA → fila rojiza; APLICAR → fila verde (ver relleno_fila_por_reglas).
# Importante: CUOTA_CUBIERTA contiene "NO APLICAR"; hay que evaluarla ANTES que "APLICAR in com"
# o toda esa fila se pintaría mal en verde.
COMENTARIO_CUOTA_CUBIERTA = "CUOTA SIGUIENTE CUBIERTA - NO APLICAR"
COMENTARIO_APLICAR        = "APLICAR"
COMENTARIO_SIN_REGLA      = "Sin Regla"

# Días en que se usa APLICAR: martes(1) a viernes(4). Sábado(5)-domingo(6)-lunes(0) → CUOTA CUBIERTA.
_DIAS_APLICAR = {1, 2, 3, 4}

# Textos en columna COMENTARIOS (fácil de editar)
ETIQUETA_REGLA_GRIS = "Regla aplicable GC < 200 (gris)"
ETIQUETA_REGLA_HOMERO = "Regla de Homero"
ETIQUETA_REGLA_PORCENTAJE = "Regla porcentaje"
ETIQUETA_HOMERO_SALDO_FAVOR = "Homero saldo a favor (= 250)"

# Mínimo de «SALDO A FAVOR DEL CLIENTE» para incluir la fila en el Excel (pesos MXN).
MIN_SALDO_FAVOR_EXCEL = 200.0

# ─────────────────────────────────────────────
# DESCARGO ESTATUS 3 — MERGE CON EL REPORTE
# ─────────────────────────────────────────────
# Directorio donde viven guia_descargo.json y descargo_estatus3.xlsx.
DESCARGO_ESTATUS3_DIR = os.path.normpath(os.path.join(REPORTE_DIR, "descargo_estatus3"))
TABLE_DESCARGO        = "cobranza_gc_verificacion_semana"


def _guia_descargo_basename_sanitized() -> str:
    """
    Nombre del JSON de checkpoint (solo basename). Permite otro archivo en localhost
    vía REPORTE_COBRANZA_DESCARGO_GUIA_BASENAME=guia_descargo.local.json para no pisar el del servidor.
    """
    raw = (os.environ.get("REPORTE_COBRANZA_DESCARGO_GUIA_BASENAME") or "").strip()
    if not raw:
        return "guia_descargo.json"
    if "/" in raw or "\\" in raw or ".." in raw or raw.startswith("."):
        return "guia_descargo.json"
    if not raw.endswith(".json"):
        return "guia_descargo.json"
    return raw


def _guia_descargo_path() -> str:
    return os.path.join(DESCARGO_ESTATUS3_DIR, _guia_descargo_basename_sanitized())


def _guia_descargo_backup_path() -> str:
    return _guia_descargo_path() + ".bak"


def _leer_guia_descargo() -> dict | None:
    orden = (_guia_descargo_path(), _guia_descargo_backup_path())
    for p in orden:
        if not os.path.isfile(p):
            continue
        try:
            with open(p, encoding="utf-8-sig") as f:
                data = json.load(f)
            if p != _guia_descargo_path():
                log.warning(
                    "Guía descargo principal ausente o ilegible; se usó respaldo: %s",
                    os.path.basename(p),
                )
            return data
        except (OSError, json.JSONDecodeError) as e:
            log.warning("No se pudo leer %s: %s", os.path.basename(p), e)
    return None


def _guardar_guia_descargo(
    *,
    ultimo_id_credito: int,
    ultimo_id_tabla: int,
    ultimo_registrado_en_cdmx: str,
) -> None:
    os.makedirs(DESCARGO_ESTATUS3_DIR, exist_ok=True)
    payload = {
        "ultimo_id_tabla": ultimo_id_tabla,
        "ultimo_registrado_en_cdmx": ultimo_registrado_en_cdmx,
        "ultimo_id_credito": ultimo_id_credito,
        "tabla": TABLE_DESCARGO,
        "orden_descargo": "registrado_en_cdmx ASC, id ASC",
        "guia_escrita_en_cdmx": datetime.now(TZ_CDMX).strftime("%Y-%m-%d %H:%M:%S"),
        "nota": (
            "Checkpoint incremental: la siguiente corrida pide filas con "
            "(registrado_en_cdmx > ultimo_registrado_en_cdmx) O "
            "(misma fecha y id de tabla > ultimo_id_tabla). "
            "Escrito por reporte_cobranza.py al fusionar el reporte unificado."
        ),
    }
    p = _guia_descargo_path()
    tmp = p + ".tmp"
    try:
        text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
        json.loads(text)
    except (TypeError, ValueError, json.JSONDecodeError) as e:
        log.warning("Serialización JSON de guía inválida; no se escribe: %s", e)
        return
    try:
        if os.path.isfile(p) and os.path.getsize(p) > 0:
            try:
                shutil.copy2(p, _guia_descargo_backup_path())
            except OSError as e2:
                log.warning("No se pudo respaldar guía anterior (.bak): %s", e2)
        with open(tmp, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
        os.replace(tmp, p)
        log.info("Guía descargo actualizada: %s", os.path.basename(p))
    except OSError as e:
        log.warning("No se pudo guardar la guía descargo: %s", e)
        try:
            if os.path.isfile(tmp):
                os.unlink(tmp)
        except OSError:
            pass


def _dt_descargo_a_str(val) -> str:
    """Datetime / str del descargo → cadena legible para la columna fecha_ultimo_abono_efectivo."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d 00:00:00")
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-":
        return s[:19].replace("T", " ")
    return s


def obtener_descargo_incremental(inicio_semana: date) -> list[dict]:
    """
    Consulta incrementalmente cobranza_gc_verificacion_semana (estatus=3)
    **solo** para la semana operativa dada (mismo inicio_semana que lista negra),
    usando la guia_descargo.json como checkpoint.
    Devuelve lista de dicts (pymysql DictCursor).
    """
    guia = _leer_guia_descargo()
    ins = inicio_semana.isoformat()
    conds: list[str] = ["`estatus` = 3", "`inicio_semana` = %s"]
    params: list = [ins]

    if guia is not None:
        reg_raw = guia.get("ultimo_registrado_en_cdmx", "")
        pk_raw  = guia.get("ultimo_id_tabla")
        if reg_raw and pk_raw is not None:
            try:
                lr   = str(reg_raw).strip()
                pk_i = int(pk_raw)
                log.info("Descargo checkpoint: id_tabla=%s, fecha=%s", pk_i, lr)
                conds.append(
                    "(`registrado_en_cdmx` > %s OR "
                    "(`registrado_en_cdmx` = %s AND `id` > %s))"
                )
                params.extend([lr, lr, pk_i])
            except (TypeError, ValueError) as e:
                log.warning(
                    "Guía descargo con checkpoint inválido (%s): se traen todas las filas.", e
                )
        else:
            log.info("Guía descargo sin checkpoint válido: traer todas las filas estatus=3.")
    else:
        log.info("Sin guía descargo previa: traer todas las filas estatus=3.")

    sql = (
        f"SELECT * FROM `{TABLE_DESCARGO}` "
        f"WHERE {' AND '.join(conds)} "
        "ORDER BY `registrado_en_cdmx` ASC, `id` ASC"
    )
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
        log.info(
            "Descargo filas nuevas en esta corrida: %d (inicio_semana=%s)",
            len(rows),
            ins,
        )
        return list(rows)
    finally:
        conn.close()


def _relleno_descargo_obs(reg: dict) -> "PatternFill | None":
    """
    Color de fila para registros del descargo que llevan texto de Observaciones.
    Si el mensaje ya trae APLICAR o CUOTA SIGUIENTE CUBIERTA, mismo verde/rojo que el reporte.
    Si no, solo reglas numéricas (SALDO_APLICABLE_GC).
    """
    com = str(reg.get("COMENTARIOS") or "")
    if COMENTARIO_CUOTA_CUBIERTA in com:
        return FILL_RED
    if COMENTARIO_APLICAR in com:
        return FILL_ROW_VERDE
    aplicable = _float_reg(reg, "SALDO_APLICABLE_GC")
    if aplicable < 200:
        return FILL_ROW_GRIS
    return FILL_ROW_NARANJA


def merge_descargo_en_reporte(
    resultados: list[dict],
    rows_descargo: list[dict],
    *,
    fecha_corte: str,
    lunes: date,
) -> list[dict]:
    """
    Agrega al listado del reporte las filas del descargo que no estén ya presentes.

    Reglas de negocio:
    · id_credito del descargo YA en resultados  → saltar (no modificar la fila del reporte).
    · id_credito del descargo NO en resultados  → insertar nueva fila con campos mapeados:
        monto_aplicar       → SALDO_APLICABLE_GC (solo esta columna)
        tipo_reporte        → STATUS_CREDITO
        ultimo_pago_efectivo → FECHA_ULTIMO_ABONO_EFECTIVO
        Monto_abono_efectivo (tbl_segundometro_semana) → ULTIMO_ABONO_EFECTIVO
        id_credito          → ID_CREDITO
        Id_cliente (si existe en fila descargo) → ID CLIENTE
        nombre              → NOMBRE_CLIENTE
        mensaje (Observaciones):
          · Con texto → COMENTARIOS = ese texto; color = APLICAR/CUOTA si aplica, si no reglas numéricas.
          · Vacío     → COMENTARIOS calculado por armar_comentarios_fila; color = relleno_fila_por_reglas.
      Columnas del descargo sin equivalente en el reporte se descartan.
      Columnas del reporte sin datos del descargo quedan en 0 / cadena vacía.
    """
    ids_en_reporte: set[str] = {str(r.get("ID_CREDITO", "")) for r in resultados}
    filas_nuevas  = 0
    filas_saltadas = 0
    ids_descargo = {str(r.get("id_credito", "") or "").strip() for r in rows_descargo}
    deuda_gc_bd_por_credito = _obtener_deuda_gc_bd_por_ids(ids_descargo)
    monto_abono_por_credito = _obtener_monto_abono_efectivo_por_ids(ids_descargo)

    for row in rows_descargo:
        id_c = str(row.get("id_credito", "") or "")
        if id_c in ids_en_reporte:
            filas_saltadas += 1
            continue

        obs = str(row.get("mensaje", "") or "").strip()

        reg: dict = {
            "ID_CREDITO":                   row.get("id_credito"),
            "ID_CLIENTE":                   _id_cliente_desde_row(row),
            "MAXI_APP_ULTIMA_CDMX":         "",
            "MAXI_APP_CONECTO":             "No",
            "NOMBRE_CLIENTE":               str(row.get("nombre", "") or ""),
            "STATUS_CREDITO":               str(row.get("tipo_reporte", "") or ""),
            "CUOTA_SEMANAL":                0.0,
            "DEUDA_GC":                     0.0,
            "SALDO_A_FAVOR":                0.0,
            "SALDO_APLICABLE_GC":           0.0,
            "FECHA_ULTIMO_ABONO_EFECTIVO":  _dt_descargo_a_str(
                row.get("ultimo_pago_efectivo") or row.get("registrado_en_cdmx")
            ),
            "ULTIMO_ABONO_EFECTIVO":        monto_abono_efectivo_para_excel(
                monto_abono_por_credito.get(id_c)
            ),
            "COMENTARIOS":                  "",
            "ERROR":                        "",
        }

        monto_raw = row.get("monto_aplicar")
        if monto_raw is not None and monto_raw != "":
            try:
                reg["SALDO_APLICABLE_GC"] = float(monto_raw)
            except (TypeError, ValueError):
                pass
        # Para estatus=3: completar métricas con S2; SALDO_APLICABLE_GC se respeta desde monto_aplicar.
        metricas_s2 = _metricas_descargo_desde_s2(
            row,
            fecha_corte,
            lunes,
            deuda_gc_bd_fallback=deuda_gc_bd_por_credito.get(id_c, 0.0),
        )
        if metricas_s2:
            idc_s2 = metricas_s2.get("id_cliente")
            if idc_s2 not in ("", None):
                reg["ID_CLIENTE"] = idc_s2
            nombre_s2 = str(metricas_s2.get("nombre_cliente") or "").strip()
            if nombre_s2:
                reg["NOMBRE_CLIENTE"] = nombre_s2
            reg["CUOTA_SEMANAL"] = float(metricas_s2.get("cuota_semanal") or 0)
            reg["DEUDA_GC"] = float(metricas_s2.get("deuda_gc") or 0)
            reg["SALDO_A_FAVOR"] = float(metricas_s2.get("saldo_a_favor") or 0)

        if obs:
            reg["COMENTARIOS"]    = obs
            reg["_fill_override"] = _relleno_descargo_obs(reg)
        else:
            reg["COMENTARIOS"] = armar_comentarios_fila("", reg)

        resultados.append(reg)
        ids_en_reporte.add(id_c)
        filas_nuevas += 1

    log.info(
        "Merge descargo: %d fila(s) nueva(s) insertada(s), %d saltada(s) (id ya en reporte).",
        filas_nuevas,
        filas_saltadas,
    )
    return resultados


# ─── QUERY PRINCIPAL: dinámica — filtro fecha = fecha de negocio (ayer CDMX el día de la corrida)

def sql_query_ids_principal(fecha_filtro: date, inicio_semana_op: date) -> str:
    """fecha_filtro = DATE(Fecha_ultimo_pago_efectivo); inicio_semana_op = martes semana operativa (lista negra)."""
    fs = fecha_filtro.isoformat()
    ins = inicio_semana_op.isoformat()
    return f"""
SELECT
    g.`Id_credito`  AS id_credito,
    MAX(s.`Id_cliente`) AS id_cliente,
    COUNT(*)         AS cuotas,
    MAX(s.`Fecha_ultimo_pago_efectivo`) AS fecha_ultimo_pago_efectivo,
    MAX(s.`Monto_abono_efectivo`)      AS monto_abono_efectivo,
    ROUND(
        SUM(
            CASE
                WHEN g.`estatus_pago` = 2              THEN 0
                WHEN IFNULL(g.`condonado`, 0) = 1      THEN 0
                WHEN g.`estatus_pago` = 1
                    THEN (g.`monto_valor` - IFNULL(g.`condonacion_parcial_monto`, 0))
                       - IFNULL(g.`monto_parcial_pagado`, 0)
                ELSE g.`monto_valor` - IFNULL(g.`condonacion_parcial_monto`, 0)
            END
        ), 2
    ) AS valor_real
FROM `gastos_cobranza` AS g
INNER JOIN `tbl_segundometro_semana` AS s
    ON s.`Id_credito` = g.`Id_credito`
   AND s.`Dias_mora` = 0
WHERE g.`condonado` = 0
  AND g.`estatus_pago` = 0
  AND s.`Fecha_ultimo_pago_efectivo` IS NOT NULL
  AND DATE(s.`Fecha_ultimo_pago_efectivo`) = '{fs}'
  AND NOT EXISTS (
      SELECT 1
      FROM `__SPARTA_SECRET_REDACTED__`.`asigna_creditos_despacho` AS d
      WHERE d.`id_credito` = g.`Id_credito`
        AND d.`estatus`    = '1'
  )
  AND NOT EXISTS (
      SELECT 1
      FROM `cobranza_gc_verificacion_semana` AS cgc
      WHERE cgc.`id_credito` = g.`Id_credito`
        AND cgc.`estatus` = 3
        AND cgc.`inicio_semana` = '{ins}'
  )
GROUP BY g.`Id_credito`
HAVING valor_real > 1
ORDER BY g.`Id_credito`;
"""


def fecha_negocio_y_reloj_cdmx() -> tuple[date, datetime]:
    """
    Reloj CDMX real y fecha de negocio = calendario CDMX **ayer**.
    La corrida programada (p. ej. día previo al operativo, 08:00 CDMX) usa ese día para S2 y filtros.
    """
    ahora_cdmx = datetime.now(TZ_CDMX)
    hoy_cal = ahora_cdmx.date()
    ayer = hoy_cal - timedelta(days=1)
    return ayer, ahora_cdmx


def notificar_google_chat(texto: str) -> None:
    """Aviso al webhook Google Chat; no interrumpe el reporte si falla."""
    u = (GCHAT_URL or "").strip()
    if not u.startswith("http"):
        return
    try:
        r = requests.post(
            u,
            json={"text": texto[:8000]},
            headers={"Content-Type": "application/json; charset=UTF-8"},
            timeout=20,
        )
        if r.status_code >= 400:
            log.warning("Google Chat HTTP %s: %s", r.status_code, (r.text or "")[:200])
    except Exception as e:
        log.warning("Google Chat error: %s", e)


# ─── QUERY LISTA NEGRA: IDs ya verificados esta semana ───────────────────────
QUERY_LISTA_NEGRA = """
SELECT DISTINCT `id_credito`
FROM `cobranza_gc_verificacion_semana`
WHERE `inicio_semana` = %s
  AND `s2_exitoso`    = 1;
"""


def inicio_semana_operativa_martes(d: date) -> date:
    """Martes que abre el periodo operativo mar–dom."""
    return d - timedelta(days=(d.weekday() - 1) % 7)


# Python weekday(): mismo orden que MySQL WEEKDAY (lunes=0 … domingo=6)
_NOMBRE_DIA_SEMANA_ES = (
    "lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo",
)


def _parse_fecha_desde_mysql(raw) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _id_cliente_desde_row(row: dict) -> int | str:
    """Id_cliente desde MySQL/descargo (varias claves posibles en dict). Vacío si no hay."""
    for k in ("id_cliente", "Id_cliente", "ID_CLIENTE"):
        if k not in row:
            continue
        v = row.get(k)
        if v is None or v == "":
            continue
        try:
            return int(v)
        except (TypeError, ValueError):
            continue
    return ""


def _id_cliente_desde_s2(ec: dict) -> int | str:
    """Id_cliente desde estado de cuenta S2 (si viene en datosCliente)."""
    cliente = ec.get("datosCliente", {}) or {}
    for k in ("idCliente", "idcliente", "id_cliente", "Id_cliente", "ID_CLIENTE"):
        if k not in cliente:
            continue
        v = cliente.get(k)
        if v is None or v == "":
            continue
        try:
            return int(v)
        except (TypeError, ValueError):
            continue
    return ""


def _obtener_deuda_gc_bd_por_ids(ids_credito: set[str]) -> dict[str, float]:
    """
    Fallback para estatus=3: deuda GC pendiente por id_credito desde __SPARTA_SECRET_REDACTED__.gastos_cobranza.
    """
    out: dict[str, float] = {}
    ids_limpios: list[int] = []
    for x in ids_credito:
        sx = str(x or "").strip()
        if not sx:
            continue
        try:
            xi = int(sx)
        except (TypeError, ValueError):
            continue
        if xi > 0:
            ids_limpios.append(xi)
    if not ids_limpios:
        return out

    sql_tpl = """
SELECT
    g.`Id_credito` AS id_credito,
    ROUND(
        SUM(
            CASE
                WHEN g.`estatus_pago` = 2         THEN 0
                WHEN IFNULL(g.`condonado`, 0) = 1 THEN 0
                WHEN g.`estatus_pago` = 1
                    THEN (g.`monto_valor` - IFNULL(g.`condonacion_parcial_monto`, 0))
                       - IFNULL(g.`monto_parcial_pagado`, 0)
                ELSE g.`monto_valor` - IFNULL(g.`condonacion_parcial_monto`, 0)
            END
        ),
        2
    ) AS valor_real
FROM `gastos_cobranza` AS g
WHERE g.`Id_credito` IN ({ph})
  AND IFNULL(g.`condonado`, 0) = 0
  AND g.`estatus_pago` IN (0, 1)
GROUP BY g.`Id_credito`
"""
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            chunk = 500
            ids_sorted = sorted(set(ids_limpios))
            for i in range(0, len(ids_sorted), chunk):
                part = ids_sorted[i : i + chunk]
                ph = ",".join(["%s"] * len(part))
                cur.execute(sql_tpl.format(ph=ph), tuple(part))
                for row in cur.fetchall():
                    cid = str(row.get("id_credito") or "").strip()
                    if not cid:
                        continue
                    try:
                        out[cid] = float(row.get("valor_real") or 0)
                    except (TypeError, ValueError):
                        out[cid] = 0.0
    finally:
        conn.close()
    return out


def _obtener_monto_abono_efectivo_por_ids(ids_credito: set[str]) -> dict[str, float | None]:
    """
    Monto_abono_efectivo por id_credito desde __SPARTA_SECRET_REDACTED__.tbl_segundometro_semana (Dias_mora=0).
    """
    out: dict[str, float | None] = {}
    ids_limpios: list[int] = []
    for x in ids_credito:
        sx = str(x or "").strip()
        if not sx:
            continue
        try:
            xi = int(sx)
        except (TypeError, ValueError):
            continue
        if xi > 0:
            ids_limpios.append(xi)
    if not ids_limpios:
        return out

    sql_tpl = """
SELECT
    s.`Id_credito` AS id_credito,
    MAX(s.`Monto_abono_efectivo`) AS monto_abono_efectivo
FROM `tbl_segundometro_semana` AS s
WHERE s.`Dias_mora` = 0
  AND s.`Id_credito` IN ({ph})
GROUP BY s.`Id_credito`
"""
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            chunk = 500
            ids_sorted = sorted(set(ids_limpios))
            for i in range(0, len(ids_sorted), chunk):
                part = ids_sorted[i : i + chunk]
                ph = ",".join(["%s"] * len(part))
                cur.execute(sql_tpl.format(ph=ph), tuple(part))
                for row in cur.fetchall():
                    cid = str(row.get("id_credito") or "").strip()
                    if not cid:
                        continue
                    raw = row.get("monto_abono_efectivo")
                    if raw is None:
                        out[cid] = None
                        continue
                    try:
                        out[cid] = float(raw)
                    except (TypeError, ValueError):
                        out[cid] = None
    finally:
        conn.close()
    return out


def _metricas_descargo_desde_s2(
    row_descargo: dict,
    fecha_corte: str,
    lunes: date,
    *,
    deuda_gc_bd_fallback: float = 0.0,
) -> dict:
    """
    Para filas estatus=3 del descargo, completa columnas igual que procesar_registro:
    - nombre/status/cuota desde S2
    - deuda_gc: calcular_gc(ec, FECHA_CORTE_GASTOS); si <=1, fallback BD por id_credito
    - saldo_a_favor: calcular_saldo_a_favor(ec, lunes)
    SALDO_APLICABLE_GC no se calcula aquí: viene de monto_aplicar.
    """
    id_credito = row_descargo.get("id_credito")
    try:
        deuda_gc = float(deuda_gc_bd_fallback or 0)
    except (TypeError, ValueError):
        deuda_gc = 0.0
    if deuda_gc <= 1:
        try:
            deuda_gc = float(row_descargo.get("valor_real") or 0)
        except (TypeError, ValueError):
            deuda_gc = 0.0
    if deuda_gc <= 1:
        deuda_gc = 0.0

    data = consultar_s2(id_credito, fecha_corte)
    if not data:
        return {
            "id_cliente": "",
            "nombre_cliente": "",
            "status_credito": "",
            "cuota_semanal": 0.0,
            "deuda_gc": deuda_gc,
            "saldo_a_favor": 0.0,
        }
    ec = data.get("estadoCuenta", {}) or {}
    cliente = ec.get("datosCliente", {}) or {}
    status = ec.get("statusCredito", "") or ""
    cuota = float(ec.get("cuota") or 0)

    gc_desde_s2 = float(calcular_gc(ec, FECHA_CORTE_GASTOS) or 0)
    if gc_desde_s2 > 1:
        deuda_gc = gc_desde_s2

    saldo_favor = float(calcular_saldo_a_favor(ec, lunes) or 0)

    return {
        "id_cliente": _id_cliente_desde_s2(ec),
        "nombre_cliente": str(cliente.get("nombreCliente", "") or ""),
        "status_credito": str(status),
        "cuota_semanal": cuota,
        "deuda_gc": deuda_gc,
        "saldo_a_favor": saldo_favor,
    }


def fecha_abono_efectivo_para_excel(row: dict, hoy: date) -> str:
    """
    Valor de columna Excel: misma fecha que MAX(Fecha_ultimo_pago_efectivo) del SELECT MySQL
    (tbl_segundometro_semana), con nombre del día en español. Antes solo se rellenaba si el pago
    era lunes, lo que dejaba la celda vacía para el resto de días.
    """
    _ = hoy  # firma conservada (procesar_registro / filas de error)
    raw = row.get("fecha_ultimo_pago_efectivo")
    d_pago = _parse_fecha_desde_mysql(raw)
    if d_pago is None:
        return ""
    nombre = _NOMBRE_DIA_SEMANA_ES[d_pago.weekday()]
    return f"{d_pago.isoformat()} — {nombre}"


def monto_abono_efectivo_para_excel(raw) -> float | str:
    """Monto_abono_efectivo (`tbl_segundometro_semana`, __SPARTA_SECRET_REDACTED__) para columna ultimo_abono_efectivo."""
    if raw is None:
        return ""
    if isinstance(raw, str) and not raw.strip():
        return ""
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return ""
    if v != v:  # NaN
        return ""
    return round(v, 2)


def obtener_ids_mysql(fecha_filtro_pago_efectivo: date, inicio_semana_op: date) -> list[dict]:
    """Trae créditos elegibles desde MySQL (sin filtro lista negra; excluye estatus 3 misma semana)."""
    log.info(
        "Conectando a MySQL para créditos (filtro fecha pago efectivo=%s, excluye GC estatus=3 inicio_semana=%s)...",
        fecha_filtro_pago_efectivo,
        inicio_semana_op,
    )
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(sql_query_ids_principal(fecha_filtro_pago_efectivo, inicio_semana_op))
            rows = cur.fetchall()
        log.info(f"  -> {len(rows):,} registros en MySQL")
        return list(rows)
    finally:
        conn.close()


def obtener_lista_negra(inicio_semana: date) -> set[str]:
    """Devuelve el conjunto de id_credito (como string) que ya están en lista negra esta semana."""
    log.info(f"Consultando lista negra (inicio_semana={inicio_semana})...")
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(QUERY_LISTA_NEGRA, (inicio_semana,))
            rows = cur.fetchall()
        ids = {str(r["id_credito"]) for r in rows}
        log.info(f"  -> {len(ids):,} IDs en lista negra")
        return ids
    finally:
        conn.close()


def _orden_id_credito(val: str):
    s = str(val).strip()
    if s.isdigit():
        return (0, int(s))
    return (1, s)


def guardar_ids_negra_fuera_pool_elegible(
    fecha_generacion_cdmx: date,
    inicio_semana: date,
    fecha_filtro_pago: date,
    ids_fuera: set[str],
) -> str:
    """
    IDs que SÍ están en lista negra (s2_exitoso=1) pero no salieron en el SELECT principal
    del día (DATE último pago efectivo ≠ fecha_filtro_pago u otros filtros). Explica 45 vs 28.
    """
    nombre = (
        f"lista_negra_en_bd_fuera_pool_{fecha_generacion_cdmx.isoformat()}_"
        f"inicio_{inicio_semana.isoformat()}_pagoef_{fecha_filtro_pago.isoformat()}.txt"
    )
    ruta = os.path.join(REPORTE_DIR, nombre)
    try:
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(
                "# En lista negra BD (s2_exitoso / inicio_semana) pero NO en pool elegibles del día\n"
                "# (no cumplen el query principal: último pago efectivo y demás filtros).\n"
                f"# fecha_generacion_cdmx={fecha_generacion_cdmx}\n"
                f"# inicio_semana={inicio_semana}\n"
                f"# filtro_fecha_ultimo_pago_efectivo_principal_QUERY={fecha_filtro_pago}\n"
                f"# total_ids={len(ids_fuera)}\n\n"
            )
            for _id in sorted(ids_fuera, key=_orden_id_credito):
                f.write(f"{_id}\n")
        return ruta
    except Exception as e:
        log.warning("No se pudo guardar archivo lista_negra fuera de pool: %s", e)
        return ""


def guardar_ids_descartados_lista_negra(
    fecha_generacion_cdmx: date,
    inicio_semana: date,
    ids_descartados: set[str],
) -> str:
    """Guarda en reporte/ el detalle de IDs realmente descartados por lista negra en la corrida."""
    nombre = (
        f"lista_negra_descartados_{fecha_generacion_cdmx.isoformat()}_"
        f"inicio_{inicio_semana.isoformat()}.txt"
    )
    ruta = os.path.join(REPORTE_DIR, nombre)
    try:
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(
                "# IDs realmente descartados por lista negra\n"
                f"# fecha_generacion_cdmx={fecha_generacion_cdmx}\n"
                f"# inicio_semana={inicio_semana}\n"
                f"# total_ids_descartados={len(ids_descartados)}\n\n"
            )
            for _id in sorted(ids_descartados, key=_orden_id_credito):
                f.write(f"{_id}\n")
        return ruta
    except Exception as e:
        log.warning("No se pudo guardar archivo de IDs descartados por lista negra: %s", e)
        return ""


def rango_fechas_ventana_maxi_app_cdmx(fecha_ref_cdmx: date) -> tuple[date, date]:
    """
    Ventana inclusiva según calendario CDMX (fecha_ref = «hoy» CDMX al generar el reporte).

    - Viernes, sábado o domingo (equivalente a DAYOFWEEK MySQL 6,7,1): del lunes de esa semana al día ref.
    - Lunes a jueves: 5 días — desde fecha_ref - 4 días hasta fecha_ref (incluye hoy + 4 anteriores).
    """
    wd = fecha_ref_cdmx.weekday()  # lunes=0 … domingo=6 (igual que MySQL WEEKDAY)
    if wd in (4, 5, 6):  # vie, sáb, dom
        inicio = fecha_ref_cdmx - timedelta(days=wd)
    else:
        inicio = fecha_ref_cdmx - timedelta(days=4)
    return inicio, fecha_ref_cdmx


def _parse_mysql_fecha_creacion(raw) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime.combine(raw, datetime.min.time())
    s = str(raw).strip()
    if not s:
        return None
    s0 = s.replace("Z", "").split(".")[0]
    if len(s0) >= 19:
        try:
            return datetime.fromisoformat(s0[:19])
        except ValueError:
            pass
    if len(s0) >= 10:
        try:
            return datetime.combine(date.fromisoformat(s0[:10]), datetime.min.time())
        except ValueError:
            pass
    return None


def formatear_ultima_conexion_maxi_cdmx(dt_raw) -> str:
    """Interpreta fecha_creacion de MySQL como hora local CDMX si viene naive; muestra en CDMX."""
    dt = _parse_mysql_fecha_creacion(dt_raw)
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TZ_CDMX)
    else:
        dt = dt.astimezone(TZ_CDMX)
    nombre = _NOMBRE_DIA_SEMANA_ES[dt.weekday()]
    return f"{dt.strftime('%Y-%m-%d %H:%M:%S')} ({nombre}) — CDMX"


def obtener_mapa_ultima_conexion_maxi_app(
    ids_cliente: set[int],
    fecha_ref_cdmx: date,
) -> dict[int, datetime]:
    """
    Por idCliente, MAX(fecha_creacion) en ubicacion dentro de DATE(fecha_creacion) entre inicio y fin
    (límites según rango_fechas_ventana_maxi_app_cdmx). Requiere que DATE en BD sea comparable al
    calendario CDMX (típico si los registros son hora México).
    """
    out: dict[int, datetime] = {}
    if not ids_cliente:
        return out
    ini, fin = rango_fechas_ventana_maxi_app_cdmx(fecha_ref_cdmx)
    ids_limpios = sorted(i for i in ids_cliente if isinstance(i, int) and i > 0)
    if not ids_limpios:
        return out
    sql_tpl = (
        "SELECT idCliente AS id_cliente, MAX(fecha_creacion) AS fecha "
        "FROM ubicacion WHERE idCliente IN ({ph}) "
        "AND DATE(fecha_creacion) BETWEEN %s AND %s GROUP BY idCliente"
    )
    conn = pymysql.connect(**DB_CONFIG_AWS_MOVIL)
    try:
        with conn.cursor() as cur:
            for i in range(0, len(ids_limpios), MAXI_APP_IN_CHUNK):
                chunk = ids_limpios[i : i + MAXI_APP_IN_CHUNK]
                ph = ",".join(["%s"] * len(chunk))
                cur.execute(sql_tpl.format(ph=ph), (*chunk, ini.isoformat(), fin.isoformat()))
                for row in cur.fetchall():
                    ic = int(row["id_cliente"])
                    parsed = _parse_mysql_fecha_creacion(row.get("fecha"))
                    if parsed is not None:
                        out[ic] = parsed
        log.info(
            "  Maxi app (__SPARTA_SECRET_REDACTED__): ventana CDMX %s .. %s — %s idCliente consultados, %s con conexión",
            ini,
            fin,
            f"{len(ids_limpios):,}",
            f"{len(out):,}",
        )
    except Exception as e:
        log.warning("  Maxi app: no se pudo consultar __SPARTA_SECRET_REDACTED__/ubicacion: %s", e)
    finally:
        conn.close()
    return out


def enriquecer_columnas_maxi_app(resultados: list[dict], fecha_ref_cdmx: date) -> None:
    """Rellena MAXI_APP_ULTIMA_CDMX y MAXI_APP_CONECTO en cada fila (idCliente = ID_CLIENTE)."""
    ids: set[int] = set()
    for r in resultados:
        ic = r.get("ID_CLIENTE")
        if ic == "" or ic is None:
            continue
        try:
            ids.add(int(ic))
        except (TypeError, ValueError):
            continue
    m = obtener_mapa_ultima_conexion_maxi_app(ids, fecha_ref_cdmx)
    for r in resultados:
        ic = r.get("ID_CLIENTE")
        try:
            ki = int(ic)
        except (TypeError, ValueError):
            r["MAXI_APP_ULTIMA_CDMX"] = ""
            r["MAXI_APP_CONECTO"] = "No"
            continue
        fdt = m.get(ki)
        if fdt is None:
            r["MAXI_APP_ULTIMA_CDMX"] = ""
            r["MAXI_APP_CONECTO"] = "No"
        else:
            r["MAXI_APP_ULTIMA_CDMX"] = formatear_ultima_conexion_maxi_cdmx(fdt)
            r["MAXI_APP_CONECTO"] = "Sí"


# ─────────────────────────────────────────────
# SESSION POOL
# ─────────────────────────────────────────────
_thread_local = threading.local()


def get_session():
    if not hasattr(_thread_local, "session"):
        s = requests.Session()
        s.headers.update(S2_HEADERS)
        s.mount("https://", requests.adapters.HTTPAdapter(
            pool_connections=1, pool_maxsize=1, max_retries=0))
        _thread_local.session = s
    return _thread_local.session


def consultar_s2(id_credito, fecha_corte):
    payload = {"idCredito": id_credito, "fechaCorte": fecha_corte}
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            resp = get_session().post(S2_URL, json=payload, timeout=TIMEOUT_S2)
            resp.raise_for_status()
            data = resp.json()
            if data.get("http") == 200:
                return data
            return None
        except requests.exceptions.Timeout:
            log.warning(f"[{id_credito}] Timeout intento {intento}")
        except requests.exceptions.ConnectionError:
            log.warning(f"[{id_credito}] ConexionError intento {intento}")
        except Exception as e:
            log.error(f"[{id_credito}] Error: {e}")
            return None
        if intento < MAX_REINTENTOS:
            time.sleep(2.0 * intento)
    return None


def calcular_gc(ec, fecha_limite):
    total_gc = 0.0
    for nota in (ec.get("datosNotasCargos") or []):
        if nota.get("concepto") != CONCEPTO_GC:
            continue
        try:
            fm = date.fromisoformat(str(nota["fechaMovimiento"])[:10])
        except Exception:
            continue
        if fm > fecha_limite:
            total_gc += float(nota.get("extemporaneos") or 0)

    if total_gc <= 0:
        return 0.0

    pagado = sum(float(p.get("extemporaneos") or 0)
                 for p in (ec.get("datosPagos") or []))

    return max(round(total_gc - pagado, 2), 0.0)


def calcular_saldo_a_favor(ec, lunes: date):
    saldos = ec.get("datosSaldos") or {}
    vencido = float(saldos.get("saldoTotalVencido") or 0)
    if vencido > 1:
        return 0.0

    lunes_str = str(lunes)

    cargos = ec.get("datosCargos") or []
    if not isinstance(cargos, list) or not cargos:
        return 0.0

    cargos_sorted = sorted(
        cargos,
        key=lambda c: int(c.get("idCargo") or 0)
    )

    monto_por_idcargo = {}
    es_anticipo_por_idcargo = {}

    target_idcargo = None
    target_monto = 0.0

    for cargo in cargos_sorted:
        idc = int(cargo.get("idCargo") or 0)
        if idc <= 0:
            continue

        concepto_upper = str(cargo.get("concepto") or "").upper()
        monto_cargo = float(cargo.get("monto") or 0)
        monto_por_idcargo[idc] = monto_cargo
        es_anticipo_por_idcargo[idc] = ("ANTICIPO A CAPITAL" in concepto_upper)

        fv = str(cargo.get("fechaVencimiento") or "")[:10]
        if fv == lunes_str and ("CUOTA SEMANAL" in concepto_upper) and target_idcargo is None:
            target_idcargo = idc
            target_monto = monto_cargo

    if target_idcargo is None or target_monto <= 0:
        return 0.0

    pagos_ordenados = sorted(
        ec.get("datosPagos") or [],
        key=lambda p: (str(p.get("fechaRegistro") or ""), p.get("idPago", 0))
    )

    abonado_por_idcargo = {}
    tol = 0.02

    for pago in pagos_ordenados:
        nums = [int(x) for x in re.findall(r"\d+", str(pago.get("numeroCuotaSemanal") or ""))]
        if not nums:
            continue

        monto_real = float(pago.get("montoPago") or 0) - float(pago.get("extemporaneos") or 0)
        monto_real = round(monto_real, 2)
        if monto_real <= 0:
            continue

        pago_real_total = monto_real
        restante = monto_real

        for idc in nums:
            if restante <= 0:
                break
            if idc not in monto_por_idcargo:
                continue

            monto_cargo = monto_por_idcargo[idc]
            ya = float(abonado_por_idcargo.get(idc, 0.0) or 0.0)
            falta = round(monto_cargo - ya, 2)
            if falta <= 0:
                continue

            es_sobrante_remaining = (restante + tol) < pago_real_total
            if es_anticipo_por_idcargo.get(idc, False) and es_sobrante_remaining:
                continue

            aplicar = min(restante, falta)
            aplicar = round(aplicar, 2)
            if aplicar <= 0:
                continue

            abonado_por_idcargo[idc] = round(ya + aplicar, 2)
            restante = round(restante - aplicar, 2)

        if restante > 0:
            sig = max(nums) + 1
            while es_anticipo_por_idcargo.get(sig, False):
                sig += 1
            abonado_por_idcargo[sig] = round(float(abonado_por_idcargo.get(sig, 0.0) or 0.0) + restante, 2)

    abonado_actual = float(abonado_por_idcargo.get(target_idcargo, 0.0) or 0.0)
    if abonado_actual + 0.009 < target_monto:
        return 0.0

    total_sf = 0.0
    excedente_actual = round(abonado_actual - target_monto, 2)
    if excedente_actual > 0:
        total_sf += excedente_actual

    for idc, ab in abonado_por_idcargo.items():
        if int(idc) > int(target_idcargo):
            total_sf += float(ab or 0.0)

    return round(total_sf, 2)


def procesar_registro(row, fecha_corte_str, lunes: date, hoy: date, fecha_comentario_cdmx: date):
    """Igual que reporte_cobranza (Downloads): dict con datos o error, None si se excluye."""
    id_credito = row["id_credito"]
    base = {
        "ID_CREDITO": id_credito,
        "ID_CLIENTE": _id_cliente_desde_row(row),
        "MAXI_APP_ULTIMA_CDMX": "",
        "MAXI_APP_CONECTO": "No",
        "NOMBRE_CLIENTE": "",
        "STATUS_CREDITO": "",
        "CUOTA_SEMANAL": 0.0,
        "DEUDA_GC": float(row["valor_real"] or 0),
        "SALDO_A_FAVOR": 0.0,
        "SALDO_APLICABLE_GC": 0.0,
        "FECHA_ULTIMO_ABONO_EFECTIVO": fecha_abono_efectivo_para_excel(row, hoy),
        "ULTIMO_ABONO_EFECTIVO": monto_abono_efectivo_para_excel(row.get("monto_abono_efectivo")),
        "COMENTARIOS": "",
        "ERROR": "",
    }

    data = consultar_s2(id_credito, fecha_corte_str)
    if not data:
        base["ERROR"] = "Sin respuesta S2 tras reintentos"
        return base

    ec      = data.get("estadoCuenta", {}) or {}
    cliente = ec.get("datosCliente",   {}) or {}
    status  = ec.get("statusCredito",  "") or ""

    base["NOMBRE_CLIENTE"] = cliente.get("nombreCliente", "")
    base["STATUS_CREDITO"] = status
    base["CUOTA_SEMANAL"]  = float(ec.get("cuota") or 0)

    if str(status).strip().lower() != "vigente":
        return None

    gc_desde_s2 = calcular_gc(ec, FECHA_CORTE_GASTOS)

    if gc_desde_s2 > 1:
        deuda_gc = gc_desde_s2
    else:
        deuda_gc = float(row.get("valor_real") or 0)

    if deuda_gc <= 1:
        return None

    sf = calcular_saldo_a_favor(ec, lunes)
    if sf <= 0:
        return None

    saldo_aplicable = round(min(sf, deuda_gc), 2)
    # sf > 0 garantizado (filtramos arriba). Etiqueta según día calendario REAL CDMX de ejecución:
    # Martes–Viernes → APLICAR (verde); Sábado–Lunes → CUOTA SIGUIENTE CUBIERTA (rojo).
    comentario = (
        COMENTARIO_APLICAR
        if fecha_comentario_cdmx.weekday() in _DIAS_APLICAR
        else COMENTARIO_CUOTA_CUBIERTA
    )

    base.update({
        "DEUDA_GC": deuda_gc,
        "SALDO_A_FAVOR": sf,
        "SALDO_APLICABLE_GC": saldo_aplicable,
    })
    base["COMENTARIOS"] = armar_comentarios_fila(comentario, base)
    return base


COLUMNAS = [
    ("ID_CREDITO", "ID CREDITO"),
    ("ID_CLIENTE", "ID CLIENTE"),
    ("NOMBRE_CLIENTE", "NOMBRE CLIENTE"),
    ("STATUS_CREDITO", "STATUS CREDITO"),
    ("CUOTA_SEMANAL", "CUOTA SEMANAL"),
    ("DEUDA_GC", "DEUDA GC PENDIENTE"),
    ("SALDO_A_FAVOR", "SALDO A FAVOR DEL CLIENTE"),
    ("SALDO_APLICABLE_GC", "SALDO APLICABLE A GC"),
    ("FECHA_ULTIMO_ABONO_EFECTIVO", "fecha_ultimo_abono_efectivo"),
    ("ULTIMO_ABONO_EFECTIVO", "ultimo_abono_efectivo"),
    ("MAXI_APP_ULTIMA_CDMX", "MAXI APP — ÚLTIMA CONEXIÓN (CDMX)"),
    ("MAXI_APP_CONECTO", "MAXI APP — ¿SE CONECTÓ?"),
]
ANCHOS = {
    "ID_CREDITO": 14,
    "ID_CLIENTE": 14,
    "NOMBRE_CLIENTE": 34,
    "STATUS_CREDITO": 14,
    "CUOTA_SEMANAL": 14,
    "DEUDA_GC": 20,
    "SALDO_A_FAVOR": 24,
    "SALDO_APLICABLE_GC": 22,
    "FECHA_ULTIMO_ABONO_EFECTIVO": 32,
    "ULTIMO_ABONO_EFECTIVO": 20,
    "MAXI_APP_ULTIMA_CDMX": 36,
    "MAXI_APP_CONECTO": 22,
}
MONEY_COLS = {"CUOTA_SEMANAL", "DEUDA_GC", "SALDO_A_FAVOR", "SALDO_APLICABLE_GC", "ULTIMO_ABONO_EFECTIVO"}
SF_COLS    = {"SALDO_A_FAVOR", "SALDO_APLICABLE_GC"}
DEUDA_COLS = {"DEUDA_GC"}

FILL_RED   = PatternFill("solid", start_color="FFCDD2")  # rojo suave: CUOTA SIGUIENTE CUBIERTA - NO APLICAR
FILL_SF    = PatternFill("solid", start_color="DDEBF7")
FILL_DEUDA = PatternFill("solid", start_color="FFEBEE")
FILL_ALT   = PatternFill("solid", start_color="F2F2F2")

# Reglas de fila (orden estricto; la primera que aplique pinta toda la fila)
FILL_ROW_VERDE  = PatternFill("solid", start_color="C8E6C9")     # verde: COMENTARIOS con APLICAR (mar–vie)
FILL_ROW_NARANJA = PatternFill("solid", start_color="FFE0B2")    # naranja: Sin Regla
FILL_ROW_GRIS   = PatternFill("solid", start_color="BDBDBD")     # gris: aplicable < 200
FILL_ROW_HOMERO = PatternFill("solid", start_color="B3E5FC")     # azul cielo: deuda y SF en [200,300]
FILL_ROW_PCT    = PatternFill("solid", start_color="E1BEE7")     # morado claro: SF/cuota en 0%–25%
FILL_ROW_SF250  = PatternFill("solid", start_color="FFF9C4")     # amarillo claro: SF == 250


def _float_reg(reg: dict, key: str) -> float:
    try:
        return float(reg.get(key) or 0)
    except (TypeError, ValueError):
        return 0.0


def _fila_excel_excluida_cuota_siguiente_cubierta(reg: dict) -> bool:
    """True = no incluir en .xlsx: comentario base «CUOTA SIGUIENTE CUBIERTA - NO APLICAR» (sáb–lun)."""
    return COMENTARIO_CUOTA_CUBIERTA in str(reg.get("COMENTARIOS") or "")


def _normalizar_texto_si_no(v: object) -> str:
    t = str(v or "").strip().lower()
    if t in {"si", "sí", "s", "yes", "y", "1", "true"}:
        return "si"
    if t in {"no", "n", "0", "false"}:
        return "no"
    return ""


def _es_status_error_excel(reg: dict) -> bool:
    return str(reg.get("STATUS_CREDITO") or "").strip().lower() == "error"


def _extraer_fecha_abono_excel(reg: dict) -> date | None:
    """
    Extrae fecha (YYYY-MM-DD) desde FECHA_ULTIMO_ABONO_EFECTIVO.
    Soporta formatos del script:
      - "YYYY-MM-DD — martes"
      - "YYYY-MM-DD HH:MM:SS"
      - "YYYY-MM-DD"
    """
    raw = str(reg.get("FECHA_ULTIMO_ABONO_EFECTIVO") or "").strip()
    if len(raw) < 10:
        return None
    head = raw[:10]
    if len(head) == 10 and head[4] == "-" and head[7] == "-":
        try:
            return date.fromisoformat(head)
        except ValueError:
            return None
    return None


def aplicar_pipeline_final_excel_gc(registros: list[dict], fecha_permitida: date) -> tuple[list[dict], dict]:
    """
    Reglas de negocio para el Excel final:
      1) Conservar solo FECHA_ULTIMO_ABONO_EFECTIVO == fecha_permitida (ayer CDMX).
      2) Excluir STATUS_CREDITO=ERROR solo en salida.
      3) Sobre MAXI_APP_CONECTO=NO, conservar saldo aplicable en [200,300].
      4) Reintegrar todos los SI + NO válidos.
      5) Validación interna ABS(ultimo_abono_efectivo) >= saldo aplicable.
      6) Si no alcanza, saldo aplicable = ABS(ultimo_abono_efectivo).
      7) Filtro final saldo aplicable > 200 (sin tope superior).
    """
    # Paso 1: fecha de último abono estricta al día de negocio (ayer CDMX).
    con_fecha_permitida = [
        r for r in registros if _extraer_fecha_abono_excel(r) == fecha_permitida
    ]
    excl_fecha_no_permitida = len(registros) - len(con_fecha_permitida)

    # Paso 2: STATUS=ERROR fuera del Excel final.
    sin_error = [r for r in con_fecha_permitida if not _es_status_error_excel(r)]
    excl_status_error = len(con_fecha_permitida) - len(sin_error)

    subset_si: list[dict] = []
    subset_no: list[dict] = []
    subset_otros: list[dict] = []
    for r in sin_error:
        con = _normalizar_texto_si_no(r.get("MAXI_APP_CONECTO"))
        if con == "si":
            subset_si.append(r)
        elif con == "no":
            subset_no.append(r)
        else:
            subset_otros.append(r)

    # Paso 3: NO con saldo aplicable entre 200 y 300 (incluyente).
    subset_no_validos = [
        r for r in subset_no if 200.0 <= _float_reg(r, "SALDO_APLICABLE_GC") <= 300.0
    ]
    excl_no_rango = len(subset_no) - len(subset_no_validos)

    # Paso 4: restaurar SI + NO válidos + valores indeterminados (no SI/no NO).
    restaurado = subset_si + subset_no_validos + subset_otros

    # Paso 5 + 6: validación interna y ajuste saldo aplicable.
    ajuste_no_alcanza = 0
    for r in restaurado:
        saldo_aplicable = _float_reg(r, "SALDO_APLICABLE_GC")
        ultimo_abono_abs = abs(_float_reg(r, "ULTIMO_ABONO_EFECTIVO"))
        alcanza = ultimo_abono_abs >= saldo_aplicable
        r["_ALCANZA_GC_INTERNO"] = "LE ALCANZA" if alcanza else "NO LE ALCANZA"
        if not alcanza:
            r["SALDO_APLICABLE_GC"] = round(ultimo_abono_abs, 2)
            ajuste_no_alcanza += 1

    # Paso 7: filtro final saldo aplicable > 200.
    filtrado_final = [r for r in restaurado if _float_reg(r, "SALDO_APLICABLE_GC") > 200.0]
    excl_saldo_final = len(restaurado) - len(filtrado_final)

    stats = {
        "inicial": len(registros),
        "con_fecha_permitida": len(con_fecha_permitida),
        "excl_fecha_no_permitida": excl_fecha_no_permitida,
        "sin_error": len(sin_error),
        "excl_status_error": excl_status_error,
        "si_total": len(subset_si),
        "no_total": len(subset_no),
        "no_validos_200_300": len(subset_no_validos),
        "excl_no_rango": excl_no_rango,
        "otros_conecto": len(subset_otros),
        "ajuste_no_alcanza": ajuste_no_alcanza,
        "excl_saldo_final": excl_saldo_final,
        "final": len(filtrado_final),
    }
    return filtrado_final, stats


def etiquetas_reglas_desde_reg(reg: dict) -> list[str]:
    """Etiquetas de negocio para COMENTARIOS (pueden combinarse varias)."""
    out: list[str] = []
    aplicable = _float_reg(reg, "SALDO_APLICABLE_GC")
    if aplicable < 200:
        out.append(ETIQUETA_REGLA_GRIS)

    deuda = _float_reg(reg, "DEUDA_GC")
    sf = _float_reg(reg, "SALDO_A_FAVOR")
    if 200 <= deuda <= 300 and 200 <= sf <= 300:
        out.append(ETIQUETA_REGLA_HOMERO)

    cuota = _float_reg(reg, "CUOTA_SEMANAL")
    if cuota > 1e-9:
        pct = round(100.0 * sf / cuota, 2)
        if 0 <= pct <= 25:
            out.append(ETIQUETA_REGLA_PORCENTAJE)

    if round(sf, 2) == 250.0:
        out.append(ETIQUETA_HOMERO_SALDO_FAVOR)

    return out


def armar_comentarios_fila(comentario_cuota: str, reg: dict) -> str:
    partes: list[str] = []
    if comentario_cuota:
        partes.append(comentario_cuota)
    partes.extend(etiquetas_reglas_desde_reg(reg))
    if not partes:
        partes.append(COMENTARIO_SIN_REGLA)
    return " | ".join(partes)


def relleno_fila_por_reglas(reg: dict) -> PatternFill | None:
    """
    Primera regla que cumpla gana (no se combinan).
    0) CUOTA SIGUIENTE CUBIERTA (sáb–lun)  → rojo (antes que APLICAR: "NO APLICAR" contiene "APLICAR")
    1) APLICAR (mar–vie)                    → verde
    2) SALDO_APLICABLE_GC < 200            → gris
    3) Regla de Homero: deuda y SF en [200,300] → azul
    4) Regla porcentaje: SF/cuota en [0,25]%    → morado
    5) Homero SF == 250                         → amarillo
    6) Sin Regla (fallback)                     → naranja
    """
    com = str(reg.get("COMENTARIOS") or "")

    if COMENTARIO_CUOTA_CUBIERTA in com:
        return FILL_RED

    if COMENTARIO_APLICAR in com:
        return FILL_ROW_VERDE

    aplicable = _float_reg(reg, "SALDO_APLICABLE_GC")
    if aplicable < 200:
        return FILL_ROW_GRIS

    deuda = _float_reg(reg, "DEUDA_GC")
    sf = _float_reg(reg, "SALDO_A_FAVOR")
    if 200 <= deuda <= 300 and 200 <= sf <= 300:
        return FILL_ROW_HOMERO

    cuota = _float_reg(reg, "CUOTA_SEMANAL")
    if cuota > 1e-9:
        pct = round(100.0 * sf / cuota, 2)
        if 0 <= pct <= 25:
            return FILL_ROW_PCT

    if round(sf, 2) == 250.0:
        return FILL_ROW_SF250

    if COMENTARIO_SIN_REGLA in com:
        return FILL_ROW_NARANJA

    return None


def generar_excel(
    registros,
    lunes: date,
    fecha_filtro_ultimo_pago: date,
    ruta_salida: str,
    *,
    fecha_generacion_cdmx: date,
    inicio_semana_operativa: date,
):
    log.info(f"Construyendo Excel ({len(registros):,} filas)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Cobranza"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols  = len(COLUMNAS)

    iy, iw, _iw = fecha_generacion_cdmx.isocalendar()
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = (
        f"Reporte Gastos Cobranza  |  Archivo / generación (CDMX): {fecha_generacion_cdmx}  |  "
        f"ISO {iy}-W{iw:02d}  |  "
        f"Filtro DATE(último pago efectivo) = {fecha_filtro_ultimo_pago}  |  "
        f"Semana operativa inicio (martes): {inicio_semana_operativa}  |  "
        f"Lunes cal. fecha negocio: {lunes}  |  GC desde: {FECHA_CORTE_GASTOS}  |  "
        f"Total registros: {len(registros):,}"
    )
    ws["A1"].font      = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18

    for ci, (key, label) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        if key in SF_COLS:
            c.fill = PatternFill("solid", start_color="1565C0")
        elif key in DEUDA_COLS:
            c.fill = PatternFill("solid", start_color="B71C1C")
        else:
            c.fill = PatternFill("solid", start_color="1F3864")
    ws.row_dimensions[2].height = 30

    f_data = Font(name="Arial", size=9)
    for ri, reg in enumerate(registros, 3):
        alt = (ri % 2 == 0)
        row_fill = reg.get("_fill_override")
        if row_fill is None:
            row_fill = relleno_fila_por_reglas(reg)
        for ci, (key, _) in enumerate(COLUMNAS, 1):
            val = reg.get(key, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font      = f_data
            c.border    = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            if row_fill is not None:
                c.fill = row_fill
            elif key in SF_COLS:
                c.fill = FILL_SF
            elif key in DEUDA_COLS:
                c.fill = FILL_DEUDA
            elif alt:
                c.fill = FILL_ALT
            if key in ("COMENTARIOS", "MAXI_APP_ULTIMA_CDMX"):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if key in MONEY_COLS and isinstance(val, (int, float)):
                c.number_format = "$#,##0.00"
            elif key in ("ID_CREDITO", "ID_CLIENTE") and isinstance(val, (int, float)):
                c.number_format = "#,##0"

    for ci, (key, _) in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = ANCHOS.get(key, 14)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

    wb.save(ruta_salida)
    log.info(f"Excel guardado: {ruta_salida}")


def main() -> None:
    t0 = time.time()
    hoy, ahora_cdmx = fecha_negocio_y_reloj_cdmx()
    fecha_calendario_cdmx = ahora_cdmx.date()
    lunes = hoy - timedelta(days=hoy.weekday())
    # Lista negra por semana operativa vigente del calendario REAL CDMX (no "ayer").
    inicio_semana = inicio_semana_operativa_martes(fecha_calendario_cdmx)
    fecha_corte = str(hoy)

    os.makedirs(REPORTE_DIR, exist_ok=True)
    # Nombre: día de generación en CDMX, formato tipo 31-03-2026 (DD-MM-AAAA; sin / por rutas Windows).
    fecha_generacion_cdmx = fecha_calendario_cdmx
    sufijo = "_PRUEBA" if MODO_PRUEBA_EXCEL else ""
    nombre_excel = f"reporte_cobranza_{fecha_generacion_cdmx.strftime('%d-%m-%Y')}{sufijo}.xlsx"
    ruta_excel = os.path.join(REPORTE_DIR, nombre_excel)

    # Si el reporte oficial de hoy ya fue generado, avisar y salir (no aplica en modo prueba ni regeneración forzada).
    if not MODO_PRUEBA_EXCEL and os.path.isfile(ruta_excel) and not REGENERAR_REPORTE:
        aviso = (
            f"\n{'=' * 65}\n"
            f"  AVISO: El reporte del día ya existe\n"
            f"  Archivo : {nombre_excel}\n"
            f"  Este reporte ya fue generado hoy. Por favor descárguelo\n"
            f"  directamente desde la tabla de reportes y úselo.\n"
            f"  Si necesita regenerarlo, elimine el archivo existente\n"
            f"  y vuelva a ejecutar el agente.\n"
            f"{'=' * 65}"
        )
        print(aviso, flush=True)
        log.warning("Reporte del día ya existe: %s — proceso terminado sin regenerar.", ruta_excel)
        return

    if MODO_PRUEBA_EXCEL:
        log.warning(
            "MODO PRUEBA EXCEL: salida %s (el reporte oficial del día no se comprueba ni se sobrescribe).",
            nombre_excel,
        )
        print(
            f"\n*** MODO PRUEBA EXCEL ***  Archivo: {nombre_excel}  (oficial del día intacto)\n",
            flush=True,
        )

    log.info("=" * 65)
    log.info(f"  Reloj CDMX (inicio) : {ahora_cdmx.isoformat()}")
    log.info(f"  Fecha negocio (ayer): {hoy}")
    log.info("  Calendario CDMX al arrancar el proceso : %s", fecha_calendario_cdmx)
    log.info(
        "  Comentarios APLICAR/NO APLICAR según calendario CDMX real: %s (%s)",
        fecha_calendario_cdmx,
        _NOMBRE_DIA_SEMANA_ES[fecha_calendario_cdmx.weekday()],
    )
    log.info(f"  Lunes semana negocio : {lunes}")
    log.info(f"  Inicio semana negra : {inicio_semana}  (martes del periodo)")
    log.info(f"  fechaCorte -> S2    : {fecha_corte}")
    log.info(f"  Filtro ultimo pago  : DATE(...) = {hoy}")
    log.info(f"  Excel guardado en   : {os.path.abspath(ruta_excel)}")
    log.info(f"  GC desde (limite)   : {FECHA_CORTE_GASTOS}")
    log.info(f"  Hilos paralelos     : {MAX_WORKERS}")
    log.info("=" * 65)

    notificar_google_chat(
        f"📊 **Reporte Cobranza** — inicio\n"
        f"Hora CDMX: `{ahora_cdmx.strftime('%Y-%m-%d %H:%M')}`\n"
        f"Fecha de negocio (ayer CDMX): `{hoy}`\n"
        f"→ MySQL filt. último pago efectivo = `{hoy}` · S2 `fechaCorte` = `{fecha_corte}`"
    )

    # PASO 1: traer TODOS los elegibles desde MySQL
    todos_rows = obtener_ids_mysql(hoy, inicio_semana)
    if not todos_rows:
        log.warning("Sin registros MySQL. Se generará Excel vacío.")
    notificar_google_chat(
        f"📥 **MySQL** listo: **{len(todos_rows):,}** créditos elegibles (filtro fecha `{hoy}`)."
    )

    # PASO 2: traer lista negra y descartar en Python
    ids_negra = obtener_lista_negra(inicio_semana)
    ids_todos = {str(r["id_credito"]) for r in todos_rows}
    ids_descartados_negra = ids_todos & ids_negra
    ids_negra_sin_pool = ids_negra - ids_todos
    ids_rows = [r for r in todos_rows if str(r["id_credito"]) not in ids_descartados_negra]
    descartados_negra = len(ids_descartados_negra)
    if len(ids_todos) != len(todos_rows):
        log.warning(
            "  Aviso: MySQL devolvió %s filas pero solo %s IDs únicos (hay repetidos en el origen).",
            len(todos_rows),
            len(ids_todos),
        )
    ruta_ids_desc = guardar_ids_descartados_lista_negra(
        fecha_generacion_cdmx, inicio_semana, ids_descartados_negra
    )
    ruta_fuera_pool = guardar_ids_negra_fuera_pool_elegible(
        fecha_generacion_cdmx, inicio_semana, hoy, ids_negra_sin_pool
    )
    log.info(f"  Descartados por lista negra (reales, IDs únicos en cruce) : {descartados_negra:,}")
    log.info(f"  IDs únicos lista negra semana                              : {len(ids_negra):,}")
    log.info(
        "  Lista negra en BD pero NO en pool del día (no consultan S2 esta corrida; "
        "fecha ult. pago efectivo del query = %s): %s",
        hoy,
        f"{len(ids_negra_sin_pool):,}",
    )
    if ruta_fuera_pool:
        log.info("  Archivo: negra BD fuera de pool elegible                 : %s", ruta_fuera_pool)
    if ids_negra_sin_pool:
        m2 = ", ".join(sorted(ids_negra_sin_pool, key=_orden_id_credito)[:40])
        log.info("  Muestra negra sin pool (hasta 40)                         : %s", m2)
    if ruta_ids_desc:
        log.info("  Archivo IDs descartados lista negra                       : %s", ruta_ids_desc)
    if ids_descartados_negra:
        muestra = ", ".join(sorted(ids_descartados_negra, key=_orden_id_credito)[:40])
        log.info("  Muestra IDs descartados (hasta 40)                        : %s", muestra)
    log.info(f"  Quedan para consultar S2    : {len(ids_rows):,}")
    notificar_google_chat(
        f"🚫 **Lista negra** semana `{inicio_semana}`: se quitan **{descartados_negra:,}** · "
        f"**{len(ids_rows):,}** van a S2."
    )

    total = len(ids_rows)
    resultados_raw: list = [None] * total
    contador = {"n": 0}
    lock = threading.Lock()
    chat_progreso_lock = threading.Lock()
    ultimo_pct_chat = -1

    def quizas_chat_progreso(n: int) -> None:
        nonlocal ultimo_pct_chat
        if total <= 0:
            return
        pct = int(100 * n / total)
        hitos = (10, 25, 50, 75, 90, 100)
        with chat_progreso_lock:
            for h in hitos:
                if pct >= h and ultimo_pct_chat < h:
                    ultimo_pct_chat = h
                    el = time.time() - t0
                    vel = n / el if el > 0 else 1
                    notificar_google_chat(
                        f"⚙️ **S2** avance **{h}%** (`{n:,}` / `{total:,}`) · ~`{vel:.1f}` req/s"
                    )
                    break

    if total > 0:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_idx = {
                executor.submit(
                    procesar_registro, row, fecha_corte, lunes, hoy, fecha_calendario_cdmx
                ): i
                for i, row in enumerate(ids_rows)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    res = future.result()
                except Exception as e:
                    row = ids_rows[idx]
                    res = {
                        "ID_CREDITO": row["id_credito"],
                        "ID_CLIENTE": _id_cliente_desde_row(row),
                        "MAXI_APP_ULTIMA_CDMX": "",
                        "MAXI_APP_CONECTO": "No",
                        "NOMBRE_CLIENTE": "",
                        "STATUS_CREDITO": "",
                        "CUOTA_SEMANAL": 0.0,
                        "DEUDA_GC": float(row["valor_real"] or 0),
                        "SALDO_A_FAVOR": 0.0,
                        "SALDO_APLICABLE_GC": 0.0,
                        "FECHA_ULTIMO_ABONO_EFECTIVO": fecha_abono_efectivo_para_excel(row, hoy),
                        "ULTIMO_ABONO_EFECTIVO": monto_abono_efectivo_para_excel(row.get("monto_abono_efectivo")),
                        "COMENTARIOS": "",
                        "ERROR": f"Excepcion: {e}",
                    }
                resultados_raw[idx] = res

                with lock:
                    contador["n"] += 1
                    n = contador["n"]
                quizas_chat_progreso(n)
                if n % LOG_CADA == 0 or n == total:
                    el = time.time() - t0
                    vel = n / el if el > 0 else 1
                    log.info(
                        f"  [{n:>6,}/{total:,}] {n/total*100:5.1f}%"
                        f" | {vel:5.1f} req/s"
                        f" | ETA: {(total-n)/vel/60:.1f} min"
                    )
    else:
        notificar_google_chat("⚙️ **S2** sin consultas (0 créditos tras filtros).")

    def _fila_error_s2_reintentable(r) -> bool:
        if r is None:
            return False
        return bool(str(r.get("ERROR") or "").strip())

    def _reintento_fusionar(idx: int, nuevo) -> None:
        """Si el reintento devuelve None (excluido negocio), conserva la fila con error anterior."""
        if nuevo is None:
            return
        resultados_raw[idx] = nuevo

    for num_pasada in range(PASADAS_REINTENTO_ERRORES):
        pendientes_idx = [
            i
            for i in range(total)
            if resultados_raw[i] is not None and _fila_error_s2_reintentable(resultados_raw[i])
        ]
        if not pendientes_idx:
            break
        log.info(
            "  Reintento errores S2 — pasada %s de %s — %s crédito(s)",
            num_pasada + 1,
            PASADAS_REINTENTO_ERRORES,
            len(pendientes_idx),
        )
        notificar_google_chat(
            f"🔁 **Reintento S2** (pasada {num_pasada + 1}/{PASADAS_REINTENTO_ERRORES}): "
            f"**{len(pendientes_idx):,}** crédito(s) con error en la pasada anterior."
        )
        if REINTENTO_PAUSA_S > 0:
            time.sleep(REINTENTO_PAUSA_S)
        w_retry = min(MAX_WORKERS, REINTENTO_MAX_WORKERS)
        with ThreadPoolExecutor(max_workers=w_retry) as executor:
            future_to_idx = {
                executor.submit(
                    procesar_registro, ids_rows[i], fecha_corte, lunes, hoy, fecha_calendario_cdmx
                ): i
                for i in pendientes_idx
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    nuevo = future.result()
                except Exception as e:
                    row = ids_rows[idx]
                    nuevo = {
                        "ID_CREDITO": row["id_credito"],
                        "ID_CLIENTE": _id_cliente_desde_row(row),
                        "MAXI_APP_ULTIMA_CDMX": "",
                        "MAXI_APP_CONECTO": "No",
                        "NOMBRE_CLIENTE": "",
                        "STATUS_CREDITO": "",
                        "CUOTA_SEMANAL": 0.0,
                        "DEUDA_GC": float(row["valor_real"] or 0),
                        "SALDO_A_FAVOR": 0.0,
                        "SALDO_APLICABLE_GC": 0.0,
                        "FECHA_ULTIMO_ABONO_EFECTIVO": fecha_abono_efectivo_para_excel(row, hoy),
                        "ULTIMO_ABONO_EFECTIVO": monto_abono_efectivo_para_excel(row.get("monto_abono_efectivo")),
                        "COMENTARIOS": "",
                        "ERROR": f"Excepcion (reintento): {e}",
                    }
                _reintento_fusionar(idx, nuevo)
        corregidos = sum(
            1
            for i in pendientes_idx
            if resultados_raw[i] is not None and not _fila_error_s2_reintentable(resultados_raw[i])
        )
        log.info("  Tras reintento: %s de %s corregidos en esta pasada", corregidos, len(pendientes_idx))
        for i in pendientes_idx:
            r = resultados_raw[i]
            if r is not None and _fila_error_s2_reintentable(r):
                err_txt = str(r.get("ERROR") or "").strip() or "(sin texto ERROR)"
                log.warning(
                    "  Sigue con error S2 tras reintento: id_credito=%s — %s",
                    r.get("ID_CREDITO"),
                    err_txt[:300],
                )
        notificar_google_chat(
            f"🔁 **Reintento** pasada {num_pasada + 1} terminada: "
            f"**{corregidos:,}** recuperados de **{len(pendientes_idx):,}**."
        )

    resultados = [r for r in resultados_raw if r is not None]
    excluidos_s2 = total - len(resultados)
    log.info(f"  Excluidos post-S2 (no vigente / sin GC / sin saldo) : {excluidos_s2:,}")
    notificar_google_chat(
        f"✅ **S2** terminado. En reporte final: **{len(resultados):,}** filas "
        f"(excl. post-S2: **{excluidos_s2:,}**). Fusionando descargo…"
    )

    # PASO 3: Descargo incremental (estatus=3) + merge con el reporte
    if SIN_DESCARGO:
        log.info("Descargo: omitido (REPORTE_COBRANZA_SIN_DESCARGO=1). guia_descargo.json no se consulta ni se escribe.")
        print("--- descargo incremental: OMITIDO (prueba SIN_DESCARGO) ---", flush=True)
        notificar_google_chat("ℹ️ **Descargo estatus-3**: omitido (`REPORTE_COBRANZA_SIN_DESCARGO=1`).")
    else:
        log.info("Consultando descargo incremental (estatus=3)...")
        print("--- descargo incremental ---", flush=True)
        rows_descargo = obtener_descargo_incremental(inicio_semana)
        if rows_descargo:
            resultados = merge_descargo_en_reporte(
                resultados,
                rows_descargo,
                fecha_corte=fecha_corte,
                lunes=lunes,
            )
            ult = rows_descargo[-1]
            if NO_GUARDAR_GUIA_DESCARGO:
                log.warning(
                    "  guia_descargo.json NO actualizada (REPORTE_COBRANZA_NO_GUARDAR_GUIA_DESCARGO=1). "
                    "La próxima corrida volverá a ver el mismo checkpoint."
                )
                print(
                    f"Descargo: {len(rows_descargo)} fila(s) fusionadas. "
                    "Guía NO guardada (modo prueba).",
                    flush=True,
                )
                notificar_google_chat(
                    f"🔀 **Descargo estatus-3**: **{len(rows_descargo):,}** fila(s) fusionadas — "
                    "**guía no guardada** (prueba)."
                )
            else:
                _guardar_guia_descargo(
                    ultimo_id_credito=int(ult.get("id_credito", 0) or 0),
                    ultimo_id_tabla=int(ult.get("id", 0) or 0),
                    ultimo_registrado_en_cdmx=_dt_descargo_a_str(ult.get("registrado_en_cdmx")),
                )
                print(
                    f"Descargo: {len(rows_descargo)} fila(s) nueva(s) en esta corrida. "
                    "Guía actualizada.",
                    flush=True,
                )
                notificar_google_chat(
                    f"🔀 **Descargo estatus-3**: **{len(rows_descargo):,}** fila(s) nueva(s) fusionadas al reporte."
                )
        else:
            log.info("Descargo: sin filas nuevas en esta corrida. La guía no se actualiza.")
            print(
                "Descargo incremental: sin filas nuevas en esta corrida. La guía no se actualiza.",
                flush=True,
            )
            notificar_google_chat("ℹ️ **Descargo estatus-3**: sin filas nuevas en esta corrida.")

    fecha_ref_maxi_cdmx = datetime.now(TZ_CDMX).date()
    log.info("  Maxi app: fecha referencia CDMX (ventana) = %s", fecha_ref_maxi_cdmx)
    enriquecer_columnas_maxi_app(resultados, fecha_ref_maxi_cdmx)

    resultados, st_gc_excel = aplicar_pipeline_final_excel_gc(resultados, hoy)
    log.info(
        "  Pipeline Excel GC (fecha permitida=%s): inicial=%s · fecha_ok=%s · sin_status_error=%s · NO_validos_200_300=%s/%s · ajustes_no_le_alcanza=%s · final=%s",
        hoy.isoformat(),
        f"{st_gc_excel['inicial']:,}",
        f"{st_gc_excel['con_fecha_permitida']:,}",
        f"{st_gc_excel['sin_error']:,}",
        f"{st_gc_excel['no_validos_200_300']:,}",
        f"{st_gc_excel['no_total']:,}",
        f"{st_gc_excel['ajuste_no_alcanza']:,}",
        f"{st_gc_excel['final']:,}",
    )
    if (
        st_gc_excel["excl_fecha_no_permitida"]
        or st_gc_excel["excl_status_error"]
        or st_gc_excel["excl_no_rango"]
        or st_gc_excel["excl_saldo_final"]
    ):
        notificar_google_chat(
            "📎 **Filtro Excel GC** aplicado\n"
            f"- Fecha último abono ≠ `{hoy.isoformat()}`: **{st_gc_excel['excl_fecha_no_permitida']:,}**\n"
            f"- STATUS=ERROR omitidas: **{st_gc_excel['excl_status_error']:,}**\n"
            f"- `MAXI APP=NO` fuera de [200,300]: **{st_gc_excel['excl_no_rango']:,}**\n"
            f"- Ajustadas por `NO LE ALCANZA`: **{st_gc_excel['ajuste_no_alcanza']:,}**\n"
            f"- Omitidas por saldo final <= 200: **{st_gc_excel['excl_saldo_final']:,}**\n"
            f"- Filas finales Excel: **{st_gc_excel['final']:,}**"
        )

    # PASO 4: generar Excel unificado (siempre, aunque queden 0 registros)
    generar_excel(
        resultados,
        lunes,
        hoy,
        ruta_excel,
        fecha_generacion_cdmx=fecha_generacion_cdmx,
        inicio_semana_operativa=inicio_semana,
    )

    elapsed = time.time() - t0
    log.info("=" * 65)
    log.info(f"  TIEMPO              : {elapsed/60:.1f} min")
    log.info(f"  MySQL total         : {len(todos_rows):,}")
    log.info(f"  Lista negra quitó   : {descartados_negra:,}")
    log.info(f"  Procesados con S2   : {total:,}")
    log.info(f"  En reporte Excel    : {len(resultados):,}")
    log.info(f"  Con advertencia     : {sum(1 for r in resultados if r.get('COMENTARIOS')):,}")
    log.info(f"  Con error S2        : {sum(1 for r in resultados if r.get('ERROR')):,}")
    log.info(f"  DEUDA GC            : ${sum(r.get('DEUDA_GC', 0) for r in resultados):>15,.2f}")
    log.info(f"  SALDO A FAVOR       : ${sum(r.get('SALDO_A_FAVOR', 0) for r in resultados):>15,.2f}")
    log.info(f"  SALDO APLICABLE     : ${sum(r.get('SALDO_APLICABLE_GC', 0) for r in resultados):>15,.2f}")
    log.info("=" * 65)

    notificar_google_chat(
        f"🏁 **Reporte Cobranza** — **OK**\n"
        f"Tiempo: `{elapsed/60:.1f}` min · Archivo: `{os.path.basename(ruta_excel)}`\n"
        f"Filas Excel: **{len(resultados):,}**"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as _e:
        logging.exception("Fallo reporte cobranza")
        try:
            notificar_google_chat(f"❌ **Reporte Cobranza** — **ERROR**\n```{_e!s}```")
        except Exception:
            pass
        raise
