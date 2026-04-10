#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Carga masiva: lee un Excel con columna ID crédito e inserta en cobranza_gc_verificacion_semana
(__SPARTA_SECRET_REDACTED__). Si existe la columna «SALDO APLICABLE A GC», su valor por fila se guarda
en monto_aplicar. La columna «COMENTARIOS» (si existe) se guarda en `celula` (texto de reglas /
motivos); si no hay columna reconocible o no se puede leer, `celula` queda NULL en insert y update.
Si el id ya existe para la misma inicio_semana con estatus=3, no se inserta duplicado:
se actualiza a estatus=2 y `celula` pasa al valor de COMENTARIOS o NULL si no aplica. No hace descargo; para eso
usa descargo_cobranza_gc_estatus3.py

Ubicación en el repo: backend/services/gastos-cobranza-agent/scripts/
Invocado por el agente Node: POST /carga-verificacion-semana/run (Excel en reporte/ec-uploads).

Todo en este archivo (sin módulos auxiliares). Lectura Excel con openpyxl (mismo stack que reporte_cobranza / descargo estatus 3). Dependencias:
  pip install -r scripts/requirements.txt

Ejemplo:
  py -3 carga_cobranza_gc_verificacion_semana_desde_excel.py --no-gui archivo.xlsx
"""

from __future__ import annotations

import argparse
import math
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

import mysql.connector
from openpyxl import load_workbook

_MEGA_PHP_DEFAULTS: dict = {
    "host": "__SPARTA_HOST_REDACTED__",
    "port": 3306,
    "database": "__SPARTA_SECRET_REDACTED__",
    "user": "__SPARTA_SECRET_REDACTED__",
    "password": "__SPARTA_PASSWORD_REDACTED__",
}

CDMX = ZoneInfo("America/Mexico_City")
TABLE_NAME = "cobranza_gc_verificacion_semana"


def log_lista_negra_resumen_ids(
    excel_path: Path,
    ids_tocados: List[int],
    *,
    dry_run: bool = False,
) -> None:
    """
    Línea explícita para el log del agente: qué id_credito entraron a lista negra (insert o 3→2).
    """
    nom = excel_path.name
    pref = "[lista-negra-gc] (dry-run) " if dry_run else "[lista-negra-gc] "
    if not ids_tocados:
        print(
            f"{pref}Archivo {nom!r}: ningun id con INSERT ni UPDATE (estatus 3->2) en esta corrida.",
            flush=True,
        )
        return
    srt = sorted(ids_tocados)
    cuerpo = ", ".join(str(x) for x in srt)
    print(
        f"{pref}Archivo {nom!r}: ids aplicados a {TABLE_NAME} "
        f"(insert nuevos o estatus 3->2), total {len(srt)}: {cuerpo}",
        flush=True,
    )


def pick_excel_path_gui() -> Optional[str]:
    try:
        import tkinter as tk
        from tkinter import filedialog
        from tkinter import TclError as TkTclError
    except Exception as e:
        print(f"No se pudo cargar tkinter ({e}). Usa --excel con la ruta.", file=sys.stderr)
        return None
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except TkTclError:
        pass
    try:
        path = filedialog.askopenfilename(
            title="Carga cobranza GC - selecciona el Excel",
            filetypes=[
                ("Libro Excel", "*.xlsx *.xlsm"),
                ("Todos los archivos", "*.*"),
            ],
        )
    except TkTclError as e:
        print(f"No se pudo abrir el diálogo de archivos ({e}). Usa --excel.", file=sys.stderr)
        path = ""
    finally:
        try:
            root.destroy()
        except TkTclError:
            pass
    return str(path).strip() if path else None


def load_env_file(path: str, *, required: bool, override: bool = False) -> None:
    if not os.path.isfile(path):
        if required:
            print(f"No existe el archivo: {path}", file=sys.stderr)
            sys.exit(2)
        return
    try:
        with open(path, encoding="utf-8-sig") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, _, val = line.partition("=")
                key, val = key.strip(), val.strip()
                if len(val) >= 2 and val[0] == val[-1] and val[0] in "\"'":
                    val = val[1:-1]
                if key and (override or key not in os.environ):
                    os.environ[key] = val
    except OSError as e:
        print(f"No se pudo leer {path}: {e}", file=sys.stderr)
        sys.exit(2)


def env_first(*names: str) -> Optional[str]:
    for n in names:
        if n in os.environ:
            return os.environ[n]
    return None


def tiene_host_db_en_entorno() -> bool:
    return bool((env_first("MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR") or "").strip())


def db_config(*, mega_php_defaults: bool) -> dict:
    if mega_php_defaults:
        return {k: v for k, v in _MEGA_PHP_DEFAULTS.items()}

    host = env_first("MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR")
    user = env_first("MEGA_USER", "MYSQL_USER", "DB_USER", "DB_USUARIO")
    password = env_first("MEGA_PASSWORD", "MYSQL_PASSWORD", "DB_PASSWORD", "DB_PASS")
    database = env_first("MEGA_DATABASE", "MYSQL_DATABASE", "DB_NAME", "DB_ESQUEMA")
    port_s = env_first("MEGA_PORT", "MYSQL_PORT", "DB_PUERTO") or "3306"

    if password is None:
        password = ""
    if not (host or "").strip() or not (user or "").strip() or not (database or "").strip():
        print(
            "Faltan MEGA_HOST/MEGA_USER/MEGA_DATABASE (o MYSQL_*/DB_*). "
            "Quita --no-mega-php-defaults o define esas variables.",
            file=sys.stderr,
        )
        sys.exit(2)
    try:
        port_i = int(str(port_s).strip())
    except ValueError:
        print("Puerto debe ser entero.", file=sys.stderr)
        sys.exit(2)

    return {
        "host": host.strip(),
        "user": user.strip(),
        "password": password,
        "database": database.strip(),
        "port": port_i,
    }


def normalize_header(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_áéíóúñü]", "", s)
    return s


def is_missing_val(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def resolve_worksheet(wb, sheet) -> Any:
    if isinstance(sheet, str) and not str(sheet).isdigit():
        return wb[str(sheet)]
    idx = int(sheet)
    if idx < 0 or idx >= len(wb.worksheets):
        raise ValueError(f"Índice de hoja inválido: {sheet!r}")
    return wb.worksheets[idx]


def read_sheet_preview_rows(excel_path: Path, sheet, max_scan: int) -> List[Tuple[Any, ...]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = resolve_worksheet(wb, sheet)
        out: List[Tuple[Any, ...]] = []
        for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
            out.append(tuple(row))
        return out
    finally:
        wb.close()


def uniquify_column_names(header_vals: Sequence[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    columns: List[str] = []
    for i, v in enumerate(header_vals):
        raw = "" if is_missing_val(v) else str(v).strip()
        base = raw if raw else f"Unnamed: {i}"
        n = seen.get(base, 0)
        seen[base] = n + 1
        columns.append(base if n == 0 else f"{base}.{n}")
    return columns


def read_rows_openpyxl(
    excel_path: Path, sheet, header_row_0: int
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """header_row_0: índice 0-based de la fila de encabezados (fila 2 en Excel → 1)."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = resolve_worksheet(wb, sheet)
        hr1 = header_row_0 + 1
        it = ws.iter_rows(min_row=hr1, max_row=hr1, values_only=True)
        header_row = next(it, None)
        if header_row is None:
            return [], []
        columns = uniquify_column_names(header_row)
        data: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=hr1 + 1, values_only=True):
            tup = tuple(row)
            d: Dict[str, Any] = {}
            for j, c in enumerate(columns):
                d[c] = tup[j] if j < len(tup) else None
            data.append(d)
        return columns, data
    finally:
        wb.close()


def find_id_credito_column_name(columns: Sequence[str]) -> Optional[str]:
    candidates = {}
    for c in columns:
        n = normalize_header(c)
        candidates[n] = c

    for key in ("id_credito", "idcredito", "id_crédito"):
        if key in candidates:
            return candidates[key]

    for c in columns:
        n = normalize_header(c)
        if "id" in n and "credito" in n.replace("crédito", "credito"):
            return c
    return None


def find_id_credito_column(columns: Sequence[str]) -> str:
    col = find_id_credito_column_name(columns)
    if col is None:
        print(
            "No se encontró columna de id_credito. Encabezados: "
            + ", ".join(repr(c) for c in columns),
            file=sys.stderr,
        )
        sys.exit(2)
    return col


def headers_mostly_unnamed(columns: Sequence[str]) -> bool:
    if not columns:
        return True
    unnamed = sum(1 for c in columns if str(c).startswith("Unnamed"))
    return unnamed >= max(1, len(columns) - 1)


def _fila_parece_encabezado_id_credito(values: Sequence[Any]) -> int:
    """
    Puntuación de una fila sin encabezado: ¿parece la fila de títulos con ID crédito?
    100 = coincidencia fuerte; 80 = contiene id y credito en la misma celda.
    """
    best = 0
    non_empty = 0
    for v in values:
        if is_missing_val(v):
            continue
        non_empty += 1
        try:
            s = normalize_header(str(v).strip())
        except Exception:
            continue
        if s in ("id_credito", "idcredito", "id_crédito"):
            best = max(best, 100)
        elif "id" in s and "credito" in s.replace("crédito", "credito"):
            best = max(best, 80)
    if best >= 80 and non_empty >= 2:
        return best
    return 0


def sniff_header_row_candidates(
    excel_path: Path,
    sheet,
    *,
    max_scan: int = 45,
) -> List[int]:
    """
    Filas (índice 0-based, misma convención que --header-row) donde una celda parece «ID CREDITO».
    Prioriza reporte_cobranza.xlsx (título fila 1, encabezados fila 2 → header-row=1).
    """
    try:
        raw = read_sheet_preview_rows(excel_path, sheet, max_scan)
    except Exception:
        return []
    if not raw:
        return []
    scored: List[Tuple[int, int]] = []
    for r, rowvals in enumerate(raw):
        sc = _fila_parece_encabezado_id_credito(rowvals)
        if sc > 0:
            scored.append((sc, r))
    scored.sort(key=lambda t: (-t[0], t[1]))
    out: List[int] = []
    seen = set()
    for _, r in scored:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def read_excel_for_carga(
    excel_path: Path,
    sheet,
    preferred_header_row: int,
    *,
    max_scan: int = 40,
) -> Tuple[List[str], List[Dict[str, Any]], int]:
    """
    Lee el .xlsx probando fila de encabezados (índice 0-based).
    Muchos reportes traen título en fila 1 y nombres de columna en fila 2 → --header-row 1.
    Primero se prueban filas detectadas automáticamente (celda tipo ID CREDITO).
    Devuelve (nombres_columna, filas_como_dict, fila_encabezado_0based).
    """
    sniffed = sniff_header_row_candidates(excel_path, sheet, max_scan=max_scan)
    order: List[int] = []
    for x in sniffed + [preferred_header_row] + list(range(0, max_scan)):
        if x in order or x < 0:
            continue
        order.append(x)

    last_cols: Optional[List[str]] = None
    last_mostly_unnamed = False
    for hr in order:
        try:
            cols, data = read_rows_openpyxl(excel_path, sheet, hr)
        except Exception as e:
            print(f"Aviso: no se pudo leer con fila de encabezado índice {hr}: {e}", file=sys.stderr)
            continue
        if not cols:
            continue
        last_cols = list(cols)
        last_mostly_unnamed = headers_mostly_unnamed(cols)
        if last_mostly_unnamed:
            continue
        if find_id_credito_column_name(cols) is None:
            continue
        if hr != preferred_header_row:
            print(
                f"Nota: encabezados útiles en fila Excel {hr + 1} "
                f"(índice --header-row={hr}); la preferida era fila {preferred_header_row + 1}.",
                flush=True,
            )
        return cols, data, hr

    if last_mostly_unnamed and last_cols is not None:
        print(
            "Casi todos los encabezados son 'Unnamed'. Revisa --header-row "
            "(en Excel: fila donde está el texto ID CREDITO; en el agente JSON: headerRow = número de esa fila, 1-based).",
            file=sys.stderr,
        )
    print(
        "No se encontraron encabezados reconocibles (p. ej. id_credito). "
        "Si hay filas de título arriba, use --header-row N con N = fila de títulos menos 1 "
        "(ej. títulos en fila 3 de Excel → --header-row 2). "
        "Reporte cobranza del agente: encabezados en fila 2 → --header-row 1 o headerRow: 2 en JSON.",
        file=sys.stderr,
    )
    if last_cols is not None:
        print(f"Columnas leídas: {last_cols}", file=sys.stderr)
    sys.exit(2)


def find_monto_aplicar_column(columns: Sequence[str]) -> Optional[str]:
    """Encabezado típico del reporte: SALDO APLICABLE A GC → monto_aplicar en BD."""
    candidates = {normalize_header(c): c for c in columns}
    for key in (
        "saldo_aplicable_a_gc",
        "saldo_aplicable_gc",
        "saldo_aplicable_agc",
    ):
        if key in candidates:
            return candidates[key]
    for c in columns:
        n = normalize_header(c)
        if "saldo" in n and "aplicable" in n:
            return c
    return None


def parse_monto_aplicar(val) -> Optional[float]:
    if is_missing_val(val):
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return float(val)
    if isinstance(val, float):
        return float(val)
    s = str(val).strip()
    if not s or s in ("—", "-"):
        return None
    low = s.lower()
    if low in ("n/a", "na", "null", "none"):
        return None
    s = s.replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = parts[0] + "." + parts[1]
        else:
            s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def find_comentarios_column(columns: Sequence[str]) -> Optional[str]:
    """Encabezado COMENTARIOS del reporte → columna `celula` en BD (texto)."""
    candidates = {normalize_header(c): c for c in columns}
    for key in ("comentarios",):
        if key in candidates:
            return candidates[key]
    for c in columns:
        n = normalize_header(c)
        if "comentario" in n:
            return c
    return None


def comentarios_para_celula(val) -> Optional[str]:
    if is_missing_val(val):
        return None
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r"\s+", " ", s)
    max_len = 5000
    if len(s) > max_len:
        s = s[:max_len]
    return s


def celula_desde_fila_comentarios(row: Dict[str, Any], col_comentarios: Optional[str]) -> Optional[str]:
    """Si no hay columna COMENTARIOS o falla la lectura, devuelve None (celula NULL en BD)."""
    if col_comentarios is None:
        return None
    raw = row.get(col_comentarios)
    return comentarios_para_celula(raw)


def find_nombre_cliente_column(columns: Sequence[str]) -> Optional[str]:
    candidates = {normalize_header(c): c for c in columns}
    for key in (
        "nombre_cliente",
        "nombre_del_cliente",
        "cliente_nombre",
        "nombrecompleto",
    ):
        if key in candidates:
            return candidates[key]
    for c in columns:
        n = normalize_header(c)
        if "nombre" in n and "cliente" in n:
            return c
    return None


def nombre_desde_celda(val) -> str:
    if is_missing_val(val):
        return "—"
    s = str(val).strip()
    if not s:
        return "—"
    s = re.sub(r"\s+", " ", s)
    if len(s) > 255:
        s = s[:255]
    return s


def parse_id_credito(val) -> Optional[int]:
    if is_missing_val(val):
        return None
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return None
        v = re.sub(r"[^\d]", "", v)
        if not v:
            return None
        return int(v)
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def iso_year_week(d: date) -> tuple[int, int]:
    y, w, _ = d.isocalendar()
    return int(y), int(w)


def inicio_semana_operativa_martes(d: date) -> date:
    return d - timedelta(days=(d.weekday() - 1) % 7)


DIA_SEMANA_ES = (
    "lunes",
    "martes",
    "miércoles",
    "jueves",
    "viernes",
    "sábado",
    "domingo",
)


def parse_date_arg(s: str) -> date:
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no válida: {s!r} (use YYYY-MM-DD o DD/MM/YYYY)")


def parse_registrado_arg(s: Optional[str]) -> datetime:
    if s:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                dt = datetime.strptime(s.strip(), fmt)
                return dt.replace(tzinfo=CDMX)
            except ValueError:
                continue
        print(f"--registrado-cdmx no válido: {s!r}", file=sys.stderr)
        sys.exit(2)
    return datetime.now(CDMX)


def main() -> None:
    # Windows suele usar cp1252 en consola; UTF-8 evita UnicodeEncodeError en mensajes al agente/log.
    if sys.platform == "win32":
        for _stream in (sys.stdout, sys.stderr):
            if hasattr(_stream, "reconfigure"):
                try:
                    _stream.reconfigure(encoding="utf-8", errors="replace")
                except Exception:
                    pass

    ap = argparse.ArgumentParser(
        description="Excel -> INSERT en cobranza_gc_verificacion_semana (carga masiva).",
        epilog="Ruta al .xlsx al final, --excel, o sin ruta -> ventana para elegir archivo.",
    )
    ap.add_argument("--excel", "-e", dest="excel_opt", default=None, metavar="RUTA")
    ap.add_argument(
        "excel_positional",
        nargs="?",
        default=None,
        metavar="archivo.xlsx",
    )
    ap.add_argument("--sheet", default=0)
    ap.add_argument(
        "--header-row",
        type=int,
        default=0,
        metavar="N",
        help="Fila de encabezados: índice 0-based (0=primera fila del libro). "
        "Por defecto también se detecta la fila que contiene «ID CREDITO». "
        "Reporte reporte_cobranza_*.xlsx: encabezados en fila 2 del Excel → N=1.",
    )
    ap.add_argument("--inicio-semana", default=None, metavar="FECHA")
    ap.add_argument("--sin-normalizar-martes", action="store_true")
    ap.add_argument("--mensaje", "-m", default=None)
    ap.add_argument("--anio-iso", type=int, default=None)
    ap.add_argument("--semana-iso", type=int, default=None)
    ap.add_argument("--registrado-cdmx", default=None)
    ap.add_argument("--estatus", type=int, default=2, choices=(0, 1, 2))
    ap.add_argument("--nombre-lote", default="")
    ap.add_argument("--ignorar-nombre-excel", action="store_true")
    ap.add_argument(
        "--tipo-reporte",
        default="falta_aplicar",
        choices=("error", "falta_aplicar"),
        help="Valor en INSERT (omitir con --tipo-reporte-nulo).",
    )
    ap.add_argument(
        "--tipo-reporte-nulo",
        action="store_true",
        help="INSERT con tipo_reporte = NULL (carga masiva lista negra desde Excel / agente).",
    )
    ap.add_argument("--dedupe", action="store_true")
    ap.add_argument("--no-dedupe", action="store_true")
    ap.add_argument("--ignore-duplicates", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--mega-php-defaults", action="store_true")
    ap.add_argument("--no-mega-php-defaults", action="store_true")
    ap.add_argument("--env-file", default=None)
    ap.add_argument("--env-file-overrides", action="store_true")
    ap.add_argument("--test-db", action="store_true")
    ap.add_argument("--no-gui", action="store_true")
    args = ap.parse_args()

    if args.env_file:
        load_env_file(args.env_file, required=True, override=bool(args.env_file_overrides))

    if args.mega_php_defaults and args.no_mega_php_defaults:
        print("No combines --mega-php-defaults con --no-mega-php-defaults.", file=sys.stderr)
        sys.exit(2)

    use_embedded = bool(args.mega_php_defaults)
    if not use_embedded and not args.no_mega_php_defaults and not tiene_host_db_en_entorno():
        use_embedded = True
        print(
            "Nota: sin MEGA_HOST en el entorno; usando credenciales embebidas.",
            file=sys.stderr,
        )
    cfg = db_config(mega_php_defaults=use_embedded)

    if args.sin_normalizar_martes and not args.inicio_semana:
        print(
            "--sin-normalizar-martes solo tiene sentido con --inicio-semana explícita.",
            file=sys.stderr,
        )
        sys.exit(2)

    if args.inicio_semana:
        try:
            inicio_raw = parse_date_arg(args.inicio_semana)
        except ValueError as e:
            print(str(e), file=sys.stderr)
            sys.exit(2)
        modo_fecha = "manual"
    else:
        inicio_raw = datetime.now(CDMX).date()
        modo_fecha = "automático (hoy CDMX)"

    inicio_martes = inicio_semana_operativa_martes(inicio_raw)
    if args.sin_normalizar_martes:
        if inicio_raw.weekday() != 1:
            print(
                f"--sin-normalizar-martes: la fecha debe ser martes. "
                f"Recibido {inicio_raw} ({DIA_SEMANA_ES[inicio_raw.weekday()]})",
                file=sys.stderr,
            )
            sys.exit(2)
        inicio = inicio_raw
        print(
            f"inicio_semana: {inicio} ({DIA_SEMANA_ES[inicio.weekday()]}) "
            f"[sin normalizar; ya es martes]  (fecha {modo_fecha})"
        )
    else:
        inicio = inicio_martes
        if modo_fecha == "automático (hoy CDMX)":
            print(
                f"inicio_semana automático: hoy CDMX {inicio_raw} "
                f"({DIA_SEMANA_ES[inicio_raw.weekday()]}) -> martes del periodo {inicio} "
                f"({DIA_SEMANA_ES[inicio.weekday()]})"
            )
        elif inicio_raw != inicio:
            print(
                f"Fecha indicada: {inicio_raw} ({DIA_SEMANA_ES[inicio_raw.weekday()]}) "
                f"-> inicio_semana operativo (martes): {inicio} ({DIA_SEMANA_ES[inicio.weekday()]})"
            )
        else:
            print(
                f"inicio_semana (martes del periodo): {inicio} ({DIA_SEMANA_ES[inicio.weekday()]})"
            )

    if args.mensaje is None:
        tag = datetime.now(CDMX).strftime("%Y%m%d_%H%M%S")
        args.mensaje = f"COBRANZA_GC_EXCEL_AUTO_{inicio.isoformat()}_{tag}"
        print(f"mensaje (automático): {args.mensaje!r}")

    if args.anio_iso is not None and args.semana_iso is not None:
        anio_iso, semana_iso = args.anio_iso, args.semana_iso
    elif args.anio_iso is not None or args.semana_iso is not None:
        print("Use ambos --anio-iso y --semana-iso o ninguno.", file=sys.stderr)
        sys.exit(2)
    else:
        anio_iso, semana_iso = iso_year_week(inicio)

    registrado = parse_registrado_arg(args.registrado_cdmx)
    registrado_naive = registrado.replace(tzinfo=None)

    conn_kw = {
        "host": cfg["host"],
        "port": cfg["port"],
        "user": cfg["user"],
        "password": cfg["password"],
        "database": cfg["database"],
        "charset": "utf8mb4",
        "use_unicode": True,
    }

    try:
        conn = mysql.connector.connect(**conn_kw)
    except mysql.connector.Error as e:
        print(f"Error de conexión MySQL: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        cur = conn.cursor()
        if args.test_db:
            cur.execute("SELECT 1 AS ok")
            print(cur.fetchone())
            return

        excel_arg = args.excel_opt or args.excel_positional
        if not excel_arg or not str(excel_arg).strip():
            if args.no_gui:
                print(
                    "Falta el Excel: pasa archivo.xlsx, --excel o quita --no-gui.",
                    file=sys.stderr,
                )
                sys.exit(2)
            print("Abriendo ventana para elegir el archivo...", flush=True)
            picked = pick_excel_path_gui()
            if not picked:
                print("No se selecciono ningun archivo.", file=sys.stderr)
                sys.exit(2)
            excel_arg = picked
            print(f"Excel seleccionado: {excel_arg}")

        excel_path = Path(str(excel_arg).strip())
        if not excel_path.is_file():
            print(f"No existe el Excel: {excel_path}", file=sys.stderr)
            sys.exit(2)

        columnas, filas, _header_usado = read_excel_for_carga(
            excel_path,
            args.sheet,
            int(args.header_row),
        )

        col_id = find_id_credito_column(columnas)
        col_nom = None if args.ignorar_nombre_excel else find_nombre_cliente_column(columnas)
        col_monto = find_monto_aplicar_column(columnas)
        col_comentarios = find_comentarios_column(columnas)
        nombre_fijo = (args.nombre_lote or "").strip()
        if nombre_fijo:
            if len(nombre_fijo) > 255:
                nombre_fijo = nombre_fijo[:255]
        else:
            nombre_fijo = ""

        print(f"Columna id_credito: {col_id!r}")
        if col_nom:
            print(f"Columna nombre (por fila): {col_nom!r}")
        elif not args.ignorar_nombre_excel:
            print("No hay columna de nombre cliente reconocible -> nombre = '-' en cada fila.")
        if col_monto:
            print(f"Columna monto_aplicar (Excel): {col_monto!r}")
        else:
            print('No hay columna tipo "SALDO APLICABLE A GC" -> monto_aplicar = NULL en cada fila.')
        if col_comentarios:
            print(f"Columna COMENTARIOS -> celula (BD): {col_comentarios!r}")
        else:
            print(
                "No hay columna COMENTARIOS reconocible -> celula = NULL en todas las filas "
                "(insert y update 3->2)."
            )

        filas_hoja = len(filas)
        filas_sin_id = 0
        registros: List[Tuple[int, str, Optional[float], Optional[str]]] = []
        for row in filas:
            i = parse_id_credito(row.get(col_id))
            if i is None or i <= 0:
                filas_sin_id += 1
                continue
            if nombre_fijo:
                nm = nombre_fijo
            elif col_nom is not None:
                nm = nombre_desde_celda(row.get(col_nom))
            else:
                nm = "—"
            mo = parse_monto_aplicar(row.get(col_monto)) if col_monto is not None else None
            ct = celula_desde_fila_comentarios(row, col_comentarios)
            registros.append((i, nm, mo, ct))

        dedupe = not args.no_dedupe
        dup_en_excel = 0
        if dedupe:
            seen: set[int] = set()
            unique: List[Tuple[int, str, Optional[float], Optional[str]]] = []
            for i, nm, mo, ct in registros:
                if i in seen:
                    dup_en_excel += 1
                    continue
                seen.add(i)
                unique.append((i, nm, mo, ct))
            registros = unique

        ids = [i for i, _, _, _ in registros]

        existentes_por_credito: dict[int, list[tuple[int, int]]] = {}
        if ids:
            paso = 500
            for off in range(0, len(ids), paso):
                chunk = ids[off : off + paso]
                ph = ",".join(["%s"] * len(chunk))
                cur.execute(
                    f"SELECT `id`, `id_credito`, `estatus` FROM `{TABLE_NAME}` "
                    f"WHERE `inicio_semana` = %s AND `id_credito` IN ({ph})",
                    (inicio,) + tuple(chunk),
                )
                for row in cur.fetchall():
                    pk, ic, est = int(row[0]), int(row[1]), int(row[2])
                    existentes_por_credito.setdefault(ic, []).append((pk, est))

        ids_en_bd = set(existentes_por_credito.keys())
        a_insertar = [(i, nm, mo, ct) for i, nm, mo, ct in registros if i not in ids_en_bd]
        ids_con_estatus_3 = {
            i
            for i, nm, mo, ct in registros
            if i in existentes_por_credito
            and any(est == 3 for _, est in existentes_por_credito[i])
        }
        omitidos_sin_actualizar = len(ids) - len(a_insertar) - len(ids_con_estatus_3)

        print(
            f"Filas en hoja: {filas_hoja}  |  sin id válido: {filas_sin_id}  |  "
            f"con id válido: {len(ids)}"
            + (f"  |  duplicados en Excel omitidos: {dup_en_excel}" if dedupe else "")
        )
        print(
            f"IDs a considerar: {len(ids)}  |  nuevos (insert): {len(a_insertar)}  "
            f"|  ya en BD con estatus 3 (pasar a 2 + celula): {len(ids_con_estatus_3)}  "
            f"|  ya en BD sin estatus 3 (sin cambio): {omitidos_sin_actualizar}"
        )
        print(
            f"BD={cfg['database']!r}  tabla={TABLE_NAME}  inicio_semana={inicio}  "
            f"anio_iso={anio_iso}  semana_iso={semana_iso}  registrado_en_cdmx={registrado_naive}"
        )

        if args.dry_run:
            print("(dry-run) Primeros 10 a insertar (id, nombre, monto, celula):", a_insertar[:10])
            muestra_3 = sorted(ids_con_estatus_3)[:15]
            if muestra_3:
                print(f"(dry-run) IDs que se actualizarian (estatus 3->2 + celula): {muestra_3}")
            if omitidos_sin_actualizar:
                print(
                    f"(dry-run) Sin insert ni update (misma semana, sin fila estatus 3): "
                    f"{omitidos_sin_actualizar} id(s)"
                )
            ids_sim = sorted({i for i, _, _, _ in a_insertar} | set(ids_con_estatus_3))
            log_lista_negra_resumen_ids(excel_path, ids_sim, dry_run=True)
            return

        if not a_insertar and not ids_con_estatus_3:
            print(
                "Nada que hacer: no hay filas nuevas ni filas en estatus 3 para esta inicio_semana."
            )
            log_lista_negra_resumen_ids(excel_path, [])
            return

        tipo_rep = None if args.tipo_reporte_nulo else args.tipo_reporte
        if args.tipo_reporte_nulo:
            print("tipo_reporte en INSERT: NULL (carga lista negra / agente)")
        s2_ok = 1
        incluido = 1

        # id_usuario_reporte: NULL en carga masiva desde Excel (en PHP/EstadoCuenta va el usuario de sesión).
        sql = (
            f"INSERT {'IGNORE ' if args.ignore_duplicates else ''}INTO `{TABLE_NAME}` "
            "(id_credito, inicio_semana, anio_iso, semana_iso, registrado_en_cdmx, "
            "ultimo_pago_efectivo, s2_exitoso, incluido_reporte, mensaje, nombre, tipo_reporte, "
            "monto_aplicar, estatus, celula, id_usuario_reporte) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        )
        sql_update_e3 = (
            f"UPDATE `{TABLE_NAME}` SET `estatus` = %s, `celula` = %s "
            f"WHERE `id_credito` = %s AND `inicio_semana` = %s AND `estatus` = 3"
        )
        up_params = [
            (2, ct, i, inicio)
            for i, nm, mo, ct in registros
            if i in ids_con_estatus_3
        ]
        if up_params:
            try:
                uchunk = 500
                filas_up = 0
                for uoff in range(0, len(up_params), uchunk):
                    part_u = up_params[uoff : uoff + uchunk]
                    cur.executemany(sql_update_e3, part_u)
                    conn.commit()
                    rc = cur.rowcount
                    filas_up += int(rc) if rc is not None else len(part_u)
                print(
                    f"Filas actualizadas (estatus 3 -> 2, celula = COMENTARIOS o NULL si no hubo columna): {filas_up}"
                )
            except mysql.connector.Error as e:
                conn.rollback()
                print(f"Error MySQL al actualizar estatus 3 -> 2: {e}", file=sys.stderr)
                if "celula" in str(e).lower() or "1366" in str(e):
                    print(
                        "Si `celula` es numérica en MySQL, hace falta VARCHAR/TEXT para los comentarios.",
                        file=sys.stderr,
                    )
                sys.exit(1)

        batch: list[tuple] = [
            (
                i,
                inicio.isoformat(),
                anio_iso,
                semana_iso,
                registrado_naive,
                None,
                s2_ok,
                incluido,
                args.mensaje,
                nm,
                tipo_rep,
                mo,
                args.estatus,
                ct,
                None,
            )
            for i, nm, mo, ct in a_insertar
        ]

        chunk = 500
        total_ins = 0
        for off in range(0, len(batch), chunk):
            part = batch[off : off + chunk]
            try:
                cur.executemany(sql, part)
                conn.commit()
            except mysql.connector.Error as e:
                conn.rollback()
                print(f"Error MySQL al insertar: {e}", file=sys.stderr)
                err_s = str(e)
                if "tipo_reporte" in err_s or "1265" in err_s:
                    print(
                        "Si usa --tipo-reporte-nulo, la columna tipo_reporte en MySQL debe permitir NULL "
                        "(o dejar de usar ese flag).",
                        file=sys.stderr,
                    )
                if "celula" in err_s.lower() or "1366" in err_s:
                    print(
                        "Si `celula` es numérica en MySQL, hace falta VARCHAR/TEXT para guardar "
                        "los comentarios del Excel.",
                        file=sys.stderr,
                    )
                sys.exit(1)
            total_ins += cur.rowcount if cur.rowcount is not None else len(part)
        if total_ins:
            print(f"Insertadas (filas afectadas reportadas): {total_ins}")
        if omitidos_sin_actualizar:
            print(
                f"Sin cambio (misma inicio_semana, ya en BD y sin fila en estatus 3): "
                f"{omitidos_sin_actualizar}"
            )

        ids_tocados = sorted({i for i, _, _, _ in a_insertar} | set(ids_con_estatus_3))
        log_lista_negra_resumen_ids(excel_path, ids_tocados)

    finally:
        conn.close()


if __name__ == "__main__":
    main()
