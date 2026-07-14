#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera Excel de clientes en Cierre_Actual = 'a) Current' con gastos de cobranza
pendientes y saldo mayor a 0 en S2 (estado de cuenta).

Fuente BD:
  __SPARTA_SECRET_REDACTED__.tbl_segundometro_semana
  __SPARTA_SECRET_REDACTED__.gastos_cobranza

Fuente S2:
  POST https://servicios.s2movil.net/s2__SPARTA_SECRET_REDACTED__/estadocuenta

Ejemplo:
  C:\\xampp\\htdocs\\sparta___SPARTA_SECRET_REDACTED__\\backend\\API\\tools\\PythonPortable\\python.exe ^
    scripts\\reporte_current_gc_pendiente_saldo_s2.py --mega-php-defaults

Prueba chica:
  ...\\python.exe scripts\\reporte_current_gc_pendiente_saldo_s2.py --mega-php-defaults --limit 50
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
PYDEPS = REPO_ROOT / "backend" / "services" / "gastos-cobranza-agent" / "pydeps"
if PYDEPS.is_dir():
    sys.path.insert(0, str(PYDEPS))

import mysql.connector
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_DB = {
    "host": "",
    "port": 3306,
    "database": "",
    "user": "",
    "password": "",
}
S2_URL = os.getenv("S2_ESTADO_CUENTA_URL") or "https://servicios.s2movil.net/s2__SPARTA_SECRET_REDACTED__/estadocuenta"
S2_TOKEN = os.getenv("S2_ESTADO_CUENTA_TOKEN") or ""
OUTPUT_DIR = REPO_ROOT / "backend" / "storage" / "reportes"
SQL_FILE = REPO_ROOT / "scripts" / "consulta_current_gastos_cobranza_pendiente.sql"
MONEY_RE = re.compile(r"[\s$,]")
PRINT_LOCK = threading.Lock()
THREAD_LOCAL = threading.local()
CACHE_VERSION = 2


def env_first(*names: str) -> str | None:
    for name in names:
        value = os.getenv(name)
        if value not in (None, ""):
            return value
    return None


def load_env_file(path: str = r"C:\xampp\secure\sparta___SPARTA_SECRET_REDACTED__.env") -> None:
    env_path = Path(os.getenv("SPARTA_ENV_FILE") or path)
    if not env_path.is_file():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def db_config(use_defaults: bool) -> dict[str, Any]:
    if use_defaults:
        load_env_file()
        return {
            "host": env_first("SEGUNDOMETRO_DB_HOST", "MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR") or "",
            "port": int(env_first("SEGUNDOMETRO_DB_PORT", "MEGA_PORT", "MYSQL_PORT", "DB_PUERTO") or "3306"),
            "database": env_first("SEGUNDOMETRO_DB_NAME", "MEGA_DATABASE", "MYSQL_DATABASE", "DB_NAME", "DB_ESQUEMA") or "",
            "user": env_first("SEGUNDOMETRO_DB_USER", "MEGA_USER", "MYSQL_USER", "DB_USER", "DB_USUARIO") or "",
            "password": env_first("SEGUNDOMETRO_DB_PASSWORD", "MEGA_PASSWORD", "MYSQL_PASSWORD", "DB_PASSWORD", "DB_PASS") or "",
        }

    host = env_first("MEGA_HOST", "MYSQL_HOST", "DB_HOST", "DB_SERVIDOR")
    user = env_first("MEGA_USER", "MYSQL_USER", "DB_USER", "DB_USUARIO")
    password = env_first("MEGA_PASSWORD", "MYSQL_PASSWORD", "DB_PASSWORD", "DB_PASS") or ""
    database = env_first("MEGA_DATABASE", "MYSQL_DATABASE", "DB_NAME", "DB_ESQUEMA")
    port_s = env_first("MEGA_PORT", "MYSQL_PORT", "DB_PUERTO") or "3306"
    missing = []
    if not host:
        missing.append("MEGA_HOST/MYSQL_HOST/DB_HOST")
    if not user:
        missing.append("MEGA_USER/MYSQL_USER/DB_USER")
    if not database:
        missing.append("MEGA_DATABASE/MYSQL_DATABASE/DB_NAME")
    if missing:
        raise SystemExit("Faltan variables de conexion: " + ", ".join(missing) + " o usa --mega-php-defaults.")
    return {
        "host": host,
        "port": int(port_s),
        "database": database,
        "user": user,
        "password": password,
    }


def to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    text = str(value).strip()
    if text == "":
        return 0.0
    text = MONEY_RE.sub("", text)
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return 0.0


def to_int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError):
        return None


def clean_for_json(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(value, Decimal):
            out[key] = float(value)
        elif hasattr(value, "isoformat"):
            out[key] = value.isoformat()
        else:
            out[key] = value
    return out


def load_universe(cfg: dict[str, Any], limit: int | None) -> list[dict[str, Any]]:
    if not SQL_FILE.is_file():
        raise RuntimeError(f"No existe la consulta SQL: {SQL_FILE}")
    sql = SQL_FILE.read_text(encoding="utf-8-sig").strip().rstrip(";")
    if limit and limit > 0:
        sql += f"\nLIMIT {int(limit)}"

    cnx = mysql.connector.connect(
        host=cfg["host"],
        port=cfg["port"],
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        charset="utf8mb4",
        use_unicode=True,
        connection_timeout=10,
    )
    try:
        cur = cnx.cursor(dictionary=True)
        cur.execute(sql)
        rows = [clean_for_json(r) for r in cur.fetchall()]
        cur.close()
        return rows
    finally:
        cnx.close()


def cache_load(path: Path) -> dict[int, dict[str, Any]]:
    if not path.is_file():
        return {}
    cache: dict[int, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            cid = to_int_or_none(obj.get("Id_credito") or obj.get("id_credito"))
            if cid and int(obj.get("_cache_version") or 0) == CACHE_VERSION:
                cache[cid] = obj
    return cache


def cache_append(path: Path, item: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with PRINT_LOCK:
        with path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(item, ensure_ascii=False, separators=(",", ":")) + "\n")


def session() -> requests.Session:
    sess = getattr(THREAD_LOCAL, "session", None)
    if sess is None:
        sess = requests.Session()
        THREAD_LOCAL.session = sess
    return sess


def s2_request(id_credito: int, fecha_corte: str, timeout: int, retries: int) -> dict[str, Any]:
    payload = {"idCredito": int(id_credito), "fechaCorte": fecha_corte}
    headers = {"Token": os.getenv("S2_ESTADO_CUENTA_TOKEN") or S2_TOKEN, "Content-Type": "application/json"}
    last_error = ""

    for attempt in range(retries + 1):
        try:
            resp = session().post(S2_URL, json=payload, headers=headers, timeout=timeout)
            status = int(resp.status_code)
            try:
                data = resp.json()
            except ValueError:
                return {
                    "ok": False,
                    "status": status,
                    "error": "Respuesta S2 no es JSON valido",
                    "estadoCuenta": None,
                }
            if status != 200:
                msg = data.get("mensaje") if isinstance(data, dict) else None
                if isinstance(msg, list):
                    msg = msg[0] if msg else ""
                return {"ok": False, "status": status, "error": str(msg or "HTTP distinto de 200"), "estadoCuenta": data}
            if not isinstance(data, dict) or "estadoCuenta" not in data:
                return {"ok": False, "status": status, "error": "S2 no devolvio estadoCuenta", "estadoCuenta": data}
            return {"ok": True, "status": status, "error": "", "estadoCuenta": data["estadoCuenta"]}
        except requests.RequestException as exc:
            last_error = str(exc)
            if attempt < retries:
                time.sleep(0.5 * (attempt + 1))

    return {"ok": False, "status": 0, "error": last_error or "Error de conexion S2", "estadoCuenta": None}


def pick_numeric(*sources: dict[str, Any], keys: list[str]) -> float:
    for src in sources:
        if not isinstance(src, dict):
            continue
        for key in keys:
            if key in src and src[key] not in (None, ""):
                return to_float(src[key])
    return 0.0


def pick_value(*sources: dict[str, Any], keys: list[str]) -> Any:
    for src in sources:
        if not isinstance(src, dict):
            continue
        for key in keys:
            if key in src and src[key] not in (None, ""):
                return src[key]
    return None


def monday_for_fecha_corte(fecha_corte: str) -> date:
    try:
        d = date.fromisoformat(str(fecha_corte)[:10])
    except ValueError:
        d = date.today()
    return d - timedelta(days=d.weekday())


def pago_cuotas(pago: dict[str, Any]) -> list[int]:
    raw = str(pago.get("numeroCuotaSemanal") or "")
    return [int(x) for x in re.findall(r"\d+", raw)]


def calcular_saldo_sobrante_s2(ec: dict[str, Any], fecha_corte: str) -> float:
    """Replica el saldo a favor/sobrante operativo usado por reporte_cobranza.py."""
    saldos = ec.get("datosSaldos") if isinstance(ec.get("datosSaldos"), dict) else {}
    if to_float(saldos.get("saldoTotalVencido")) > 1:
        return 0.0

    lunes = monday_for_fecha_corte(fecha_corte)
    cargos = ec.get("datosCargos") if isinstance(ec.get("datosCargos"), list) else []
    if not cargos:
        return 0.0

    def cargo_id(cargo: dict[str, Any]) -> int:
        return int(to_float(cargo.get("idCargo")))

    cargos_sorted = sorted((c for c in cargos if isinstance(c, dict)), key=cargo_id)
    monto_por_idcargo: dict[int, float] = {}
    es_anticipo_por_idcargo: dict[int, bool] = {}
    target_idcargo: int | None = None
    target_monto = 0.0

    for cargo in cargos_sorted:
        idc = cargo_id(cargo)
        if idc <= 0:
            continue
        concepto = str(cargo.get("concepto") or "").upper()
        monto = round(to_float(cargo.get("monto")), 2)
        monto_por_idcargo[idc] = monto
        es_anticipo_por_idcargo[idc] = "ANTICIPO A CAPITAL" in concepto
        fecha_venc = str(cargo.get("fechaVencimiento") or "")[:10]
        if fecha_venc == str(lunes) and "CUOTA SEMANAL" in concepto and target_idcargo is None:
            target_idcargo = idc
            target_monto = monto

    if target_idcargo is None or target_monto <= 0:
        return 0.0

    pagos = ec.get("datosPagos") if isinstance(ec.get("datosPagos"), list) else []
    pagos_ordenados = sorted(
        (p for p in pagos if isinstance(p, dict)),
        key=lambda p: (str(p.get("fechaRegistro") or ""), int(to_float(p.get("idPago")))),
    )

    abonado_por_idcargo: dict[int, float] = {}
    tol = 0.02
    for pago in pagos_ordenados:
        cuotas = pago_cuotas(pago)
        if not cuotas:
            continue
        monto_real = round(to_float(pago.get("montoPago")) - to_float(pago.get("extemporaneos")), 2)
        if monto_real <= 0:
            continue
        pago_real_total = monto_real
        restante = monto_real
        for idc in cuotas:
            if restante <= 0:
                break
            if idc not in monto_por_idcargo:
                continue
            monto_cargo = monto_por_idcargo[idc]
            ya = round(abonado_por_idcargo.get(idc, 0.0), 2)
            falta = round(monto_cargo - ya, 2)
            if falta <= 0:
                continue
            es_sobrante_remaining = (restante + tol) < pago_real_total
            if es_anticipo_por_idcargo.get(idc, False) and es_sobrante_remaining:
                continue
            aplicar = round(min(restante, falta), 2)
            if aplicar <= 0:
                continue
            abonado_por_idcargo[idc] = round(ya + aplicar, 2)
            restante = round(restante - aplicar, 2)
        if restante > 0:
            sig = max(cuotas) + 1
            while es_anticipo_por_idcargo.get(sig, False):
                sig += 1
            abonado_por_idcargo[sig] = round(abonado_por_idcargo.get(sig, 0.0) + restante, 2)

    abonado_actual = round(abonado_por_idcargo.get(target_idcargo, 0.0), 2)
    if abonado_actual + 0.009 < target_monto:
        return 0.0

    total = round(max(abonado_actual - target_monto, 0.0), 2)
    for idc, abonado in abonado_por_idcargo.items():
        if int(idc) > int(target_idcargo):
            total += to_float(abonado)
    return round(total, 2)


def summarize_s2(id_credito: int, fecha_corte: str, result: dict[str, Any]) -> dict[str, Any]:
    base = {
        "_cache_version": CACHE_VERSION,
        "Id_credito": id_credito,
        "fecha_corte_s2": fecha_corte,
        "s2_ok": bool(result.get("ok")),
        "s2_status": result.get("status"),
        "s2_error": result.get("error") or "",
    }
    if not result.get("ok") or not isinstance(result.get("estadoCuenta"), dict):
        return base

    ec = result["estadoCuenta"]
    ds = ec.get("datosSaldos") if isinstance(ec.get("datosSaldos"), dict) else {}
    dc = ec.get("datosCliente") if isinstance(ec.get("datosCliente"), dict) else {}
    dcr = ec.get("datosCredito") if isinstance(ec.get("datosCredito"), dict) else {}

    saldo_total_vigente = pick_numeric(ds, ec, keys=["saldoTotalVigente", "CapitalPendientePago", "capitalPendientePago"])
    saldo_total_vencido = pick_numeric(ds, ec, keys=["saldoTotalVencido"])
    adeudo_total = pick_numeric(ds, ec, keys=["adeudoTotal", "montoTotalAdeudado", "saldo", "saldoTotal"])
    if adeudo_total <= 0:
        adeudo_total = saldo_total_vigente + saldo_total_vencido

    saldo_capital = pick_numeric(ds, ec, keys=["saldoCapital", "saldoTotalCapital"])
    if saldo_capital <= 0:
        saldo_capital = saldo_total_vigente if saldo_total_vigente > 0 else saldo_total_vigente + saldo_total_vencido

    cuotas_pagadas = to_int_or_none(
        pick_value(ds, dc, dcr, ec, keys=["cuotasPagadas", "Num_cuotas_pagadas", "num_cuotas_pagadas"])
    )
    cuotas_contratadas = to_int_or_none(
        pick_value(
            ds,
            dc,
            dcr,
            ec,
            keys=[
                "cuotasContratadas",
                "Numero_amortizaciones",
                "numeroAmortizaciones",
                "plazoTotal",
                "totalCuotas",
            ],
        )
    )
    cuotas_devengadas = to_int_or_none(
        pick_value(ds, dc, dcr, ec, keys=["cuotasDevengadas", "cuotas_devengadas", "Cuotas_devengadas"])
    )
    if cuotas_devengadas is None:
        cuotas_devengadas = to_int_or_none(pick_value(ds, dc, dcr, ec, keys=["cuotasVencidas", "Cuotas_vencidas"]))
    cuotas_anticipadas = 0
    if cuotas_pagadas is not None and cuotas_devengadas is not None:
        cuotas_anticipadas = max(cuotas_pagadas - cuotas_devengadas, 0)

    base.update(
        {
            "s2_id_credito": pick_value(dc, ec, keys=["idCredito", "Id_credito"]) or id_credito,
            "s2_id_cliente": pick_value(dc, ec, keys=["idCliente", "Id_cliente"]),
            "s2_nombre_cliente": pick_value(dc, ec, keys=["nombreCliente", "nombre", "Nombre_cliente"]),
            "s2_adeudo_total": round(adeudo_total, 2),
            "s2_saldo_capital": round(saldo_capital, 2),
            "s2_saldo_total_vigente": round(saldo_total_vigente, 2),
            "s2_saldo_total_vencido": round(saldo_total_vencido, 2),
            "s2_dias_mora_maximo": to_int_or_none(pick_value(ds, ec, keys=["diasMoraMaximo", "Dias_mora_max"])),
            "s2_saldo_sobrante": calcular_saldo_sobrante_s2(ec, fecha_corte),
            "s2_cuotas_devengadas": cuotas_devengadas,
            "s2_cuotas_pagadas": cuotas_pagadas,
            "s2_cuotas_contratadas": cuotas_contratadas,
            "s2_cuotas_anticipadas": cuotas_anticipadas,
            "s2_total_pagos": len(ec.get("datosPagos") or []) if isinstance(ec.get("datosPagos"), list) else None,
            "s2_total_cargos": len(ec.get("datosCargos") or []) if isinstance(ec.get("datosCargos"), list) else None,
        }
    )
    return base


def process_one(row: dict[str, Any], fecha_corte: str, timeout: int, retries: int, cache_path: Path) -> dict[str, Any]:
    id_credito = int(row["Id_credito"])
    result = s2_request(id_credito, fecha_corte, timeout, retries)
    summary = summarize_s2(id_credito, fecha_corte, result)
    cache_append(cache_path, summary)
    return summary


def enrich_with_s2(
    rows: list[dict[str, Any]],
    fecha_corte: str,
    timeout: int,
    retries: int,
    workers: int,
    cache_path: Path,
    use_cache: bool,
) -> list[dict[str, Any]]:
    cache = cache_load(cache_path) if use_cache else {}
    out_by_id: dict[int, dict[str, Any]] = {}
    pending: list[dict[str, Any]] = []
    for row in rows:
        cid = int(row["Id_credito"])
        if cid in cache:
            out_by_id[cid] = cache[cid]
        else:
            pending.append(row)

    total = len(rows)
    done = len(out_by_id)
    if done:
        print(f"[INFO] Respuestas S2 en cache: {done}")
    print(f"[INFO] Consultas S2 pendientes: {len(pending)}")

    if pending:
        with ThreadPoolExecutor(max_workers=max(1, min(workers, 32))) as executor:
            futures = {
                executor.submit(process_one, row, fecha_corte, timeout, retries, cache_path): int(row["Id_credito"])
                for row in pending
            }
            for future in as_completed(futures):
                cid = futures[future]
                try:
                    out_by_id[cid] = future.result()
                except Exception as exc:
                    item = {
                        "Id_credito": cid,
                        "fecha_corte_s2": fecha_corte,
                        "s2_ok": False,
                        "s2_status": 0,
                        "s2_error": f"Error inesperado: {exc}",
                    }
                    out_by_id[cid] = item
                    cache_append(cache_path, item)
                done += 1
                if done % 100 == 0 or done == total:
                    print(f"[INFO] S2 procesados: {done}/{total}", flush=True)

    return [out_by_id[int(row["Id_credito"])] for row in rows if int(row["Id_credito"]) in out_by_id]


def autosize(ws, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row_cells:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    add_table(ws, table_name)
    autosize(ws)


def build_excel(
    universe: list[dict[str, Any]],
    s2_rows: list[dict[str, Any]],
    saldo_field: str,
    output_path: Path,
    fecha_corte: str,
) -> dict[str, int]:
    s2_by_id = {int(r["Id_credito"]): r for r in s2_rows}
    ok_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for row in universe:
        cid = int(row["Id_credito"])
        s2 = s2_by_id.get(cid, {"s2_ok": False, "s2_error": "Sin respuesta S2"})
        merged = {**row, **s2}
        if not merged.get("s2_ok"):
            error_rows.append(merged)
            continue
        if to_float(merged.get(saldo_field)) > 0:
            ok_rows.append(merged)

    ok_rows.sort(key=lambda r: (to_float(r.get(saldo_field)) * -1, to_float(r.get("saldo_gc_pendiente")) * -1, int(r["Id_credito"])))
    error_rows.sort(key=lambda r: int(r["Id_credito"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Saldo S2 mayor 0"

    headers = [
        "Id_credito",
        "Id_cliente",
        "Nombre_cliente",
        "Celular",
        "Sucursal",
        "Status_credito",
        "Cierre_Actual",
        "Bucket_Morosidad_Real",
        "Dias_mora",
        "Dias_mora_ajustado",
        "Saldo_vencido_inicio",
        "gastos_pendientes",
        "saldo_gc_pendiente",
        "monto_gc_original",
        "monto_condonado_parcial",
        "monto_pagado_parcial",
        "primer_periodo_gc",
        "ultimo_periodo_gc",
        "s2_adeudo_total",
        "s2_saldo_capital",
        "s2_saldo_sobrante",
        "s2_saldo_total_vigente",
        "s2_saldo_total_vencido",
        "s2_dias_mora_maximo",
        "s2_cuotas_devengadas",
        "s2_cuotas_pagadas",
        "s2_cuotas_anticipadas",
        "s2_cuotas_contratadas",
        "s2_total_pagos",
        "s2_total_cargos",
        "fecha_corte_s2",
        "s2_nombre_cliente",
        "ids_gastos_cobranza",
    ]
    write_rows(ws, headers, ok_rows, "TablaSaldoS2Mayor0")

    money_cols = {
        "Saldo_vencido_inicio",
        "saldo_gc_pendiente",
        "monto_gc_original",
        "monto_condonado_parcial",
        "monto_pagado_parcial",
        "s2_adeudo_total",
        "s2_saldo_capital",
        "s2_saldo_sobrante",
        "s2_saldo_total_vigente",
        "s2_saldo_total_vencido",
    }
    for idx, header in enumerate(headers, start=1):
        if header in money_cols:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = '$#,##0.00'

    ws_err = wb.create_sheet("Errores S2")
    err_headers = [
        "Id_credito",
        "Id_cliente",
        "Nombre_cliente",
        "s2_status",
        "s2_error",
        "saldo_gc_pendiente",
        "gastos_pendientes",
        "ids_gastos_cobranza",
    ]
    write_rows(ws_err, err_headers, error_rows, "TablaErroresS2")

    ws_res = wb.create_sheet("Resumen")
    resumen_rows = [
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Fecha corte S2", fecha_corte),
        ("Universo SQL", len(universe)),
        ("Consultas S2 OK", sum(1 for r in s2_rows if r.get("s2_ok"))),
        ("Errores S2", len(error_rows)),
        ("Clientes con saldo S2 > 0", len(ok_rows)),
        ("Campo de saldo usado", saldo_field),
        ("Current", "tbl_segundometro_semana.Cierre_Actual = 'a) Current'"),
        ("Gasto pendiente", "condonado = 0 AND estatus_pago IN (0, 1)"),
    ]
    for row in resumen_rows:
        ws_res.append(row)
    ws_res["A1"].font = Font(bold=True)
    ws_res["B1"].font = Font(bold=True)
    autosize(ws_res)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"universo": len(universe), "ok_s2": sum(1 for r in s2_rows if r.get("s2_ok")), "errores_s2": len(error_rows), "saldo_mayor_0": len(ok_rows)}


def main() -> int:
    parser = argparse.ArgumentParser(description="Reporte current + GC pendiente + saldo S2 > 0.")
    parser.add_argument(
        "--env-db",
        action="store_true",
        help="Usa variables MEGA_*/MYSQL_*/DB_* en lugar de la conexion __SPARTA_SECRET_REDACTED__ del proyecto.",
    )
    parser.add_argument("--fecha-corte", default=datetime.now().strftime("%Y-%m-%d"), help="Fecha de corte S2 YYYY-MM-DD.")
    parser.add_argument("--limit", type=int, default=0, help="Limita creditos para prueba.")
    parser.add_argument("--workers", type=int, default=8, help="Consultas S2 en paralelo.")
    parser.add_argument("--timeout", type=int, default=20, help="Timeout por consulta S2.")
    parser.add_argument("--retries", type=int, default=1, help="Reintentos por consulta S2.")
    parser.add_argument(
        "--saldo-field",
        choices=["s2_adeudo_total", "s2_saldo_capital", "s2_saldo_total_vencido", "s2_saldo_total_vigente"],
        default="s2_adeudo_total",
        help="Campo S2 usado para filtrar saldo > 0.",
    )
    parser.add_argument("--no-cache", action="store_true", help="Ignora cache previo.")
    parser.add_argument("--output", type=Path, default=None, help="Ruta .xlsx de salida.")
    args = parser.parse_args()

    out = args.output
    if out is None:
        suffix = f"limit{args.limit}_" if args.limit and args.limit > 0 else ""
        out = OUTPUT_DIR / f"current_gc_pendiente_saldo_s2_{suffix}{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    cache_path = out.with_suffix(".s2_cache.jsonl")

    print("[INFO] Cargando universo SQL...")
    universe = load_universe(db_config(not args.env_db), args.limit if args.limit > 0 else None)
    print(f"[INFO] Universo SQL: {len(universe)} creditos")
    if not universe:
        print("[INFO] No hay datos para consultar.")
        return 0

    s2_rows = enrich_with_s2(
        universe,
        args.fecha_corte,
        timeout=max(2, args.timeout),
        retries=max(0, args.retries),
        workers=max(1, args.workers),
        cache_path=cache_path,
        use_cache=not args.no_cache,
    )
    stats = build_excel(universe, s2_rows, args.saldo_field, out, args.fecha_corte)
    print("[OK] Excel generado:", out)
    print("[OK] Resumen:", json.dumps(stats, ensure_ascii=False))
    print("[INFO] Cache S2:", cache_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
