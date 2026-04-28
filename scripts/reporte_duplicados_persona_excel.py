#!/usr/bin/env python3
"""
Genera un Excel de posibles duplicados en __SPARTA_SECRET_REDACTED__.persona.

Uso:
  python scripts/reporte_duplicados_persona_excel.py
  python scripts/reporte_duplicados_persona_excel.py --sim-threshold 88 --out scripts/reporte_duplicados_persona.xlsx
  python scripts/reporte_duplicados_persona_excel.py --limit 8000
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
from collections import defaultdict
from typing import Any

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def normalize_text(value: Any) -> str:
    s = str(value or "").strip().lower()
    if not s:
        return ""
    # Normalizacion simple sin dependencias externas
    repl = str.maketrans(
        {
            "á": "a",
            "à": "a",
            "ä": "a",
            "â": "a",
            "ã": "a",
            "é": "e",
            "è": "e",
            "ë": "e",
            "ê": "e",
            "í": "i",
            "ì": "i",
            "ï": "i",
            "î": "i",
            "ó": "o",
            "ò": "o",
            "ö": "o",
            "ô": "o",
            "õ": "o",
            "ú": "u",
            "ù": "u",
            "ü": "u",
            "û": "u",
            "ñ": "n",
        }
    )
    s = s.translate(repl)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def only_digits(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def similarity_ratio(a: str, b: str) -> float:
    # Similaridad por secuencias comunes sin dependencias.
    if not a or not b:
        return 0.0
    la = len(a)
    lb = len(b)
    if la == 0 or lb == 0:
        return 0.0
    # LCS dinamico para tener una medida robusta.
    dp = [0] * (lb + 1)
    best = 0
    for i in range(1, la + 1):
        prev = 0
        for j in range(1, lb + 1):
            tmp = dp[j]
            if a[i - 1] == b[j - 1]:
                dp[j] = prev + 1
            else:
                dp[j] = max(dp[j], dp[j - 1])
            prev = tmp
            if dp[j] > best:
                best = dp[j]
    return round((2.0 * best / (la + lb)) * 100.0, 1)


def pick_first(cols: list[str], candidates: list[str]) -> str | None:
    for c in candidates:
        if c in cols:
            return c
    return None


def build_full_name(row: dict[str, Any], name_col: str | None, second_name_col: str | None, ap1_col: str | None, ap2_col: str | None) -> str:
    parts: list[str] = []
    for c in [name_col, second_name_col, ap1_col, ap2_col]:
        if c:
            v = str(row.get(c, "") or "").strip()
            if v:
                parts.append(v)
    return " ".join(parts).strip()


def style_sheet(ws) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col:
            value = "" if c.value is None else str(c.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 50)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0, help="Limite de filas (0 = sin limite)")
    parser.add_argument("--sim-threshold", type=float, default=88.0, help="Umbral similitud 0..100")
    parser.add_argument("--out", type=str, default="", help="Ruta de salida .xlsx")
    args = parser.parse_args()

    db_host = os.getenv("DB_HOST", "__SPARTA_HOST_REDACTED__")
    db_port = int(os.getenv("DB_PUERTO", "3306"))
    db_name = os.getenv("DB_NAME", "__SPARTA_SECRET_REDACTED__")
    db_user = os.getenv("DB_USER", "__SPARTA_SECRET_REDACTED__")
    db_pass = os.getenv("DB_PASSWORD", "__SPARTA_PASSWORD_REDACTED__")

    conn = pymysql.connect(
        host=db_host,
        port=db_port,
        user=db_user,
        password=db_pass,
        database=db_name,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=10,
        read_timeout=120,
        write_timeout=120,
    )

    with conn.cursor() as cur:
        cur.execute("SHOW COLUMNS FROM persona")
        cols = [r["Field"] for r in cur.fetchall()]

    id_col = pick_first(cols, ["id", "id_persona"])
    name_col = pick_first(cols, ["nombre", "nombres"])
    ap1_col = pick_first(cols, ["apellido_paterno", "primer_apellido", "apellido1", "apellido", "apellidop", "ap_paterno"])
    ap2_col = pick_first(cols, ["apellido_materno", "segundo_apellido", "apellido2", "apellidom", "ap_materno"])
    second_name_col = pick_first(cols, ["segundo_nombre", "nombre2", "segundo_nombre_persona"])
    birth_col = pick_first(cols, ["fecha_nacimiento", "f_nacimiento", "nacimiento"])
    doc_col = pick_first(cols, ["dpi", "curp", "cedula", "identificacion", "documento", "no_documento"])
    mail_col = pick_first(cols, ["email", "correo", "correo_electronico", "mail"])
    phone_col = pick_first(cols, ["telefono", "telefono1", "celular", "movil"])
    status_col = pick_first(cols, ["estatus", "status"])

    required_missing = [n for n, v in [("id", id_col), ("nombre", name_col), ("apellido1", ap1_col)] if v is None]
    if required_missing:
        raise RuntimeError(f"Faltan columnas minimas en persona: {', '.join(required_missing)}")

    sql = "SELECT * FROM persona"
    if args.limit and args.limit > 0:
        sql += f" LIMIT {int(args.limit)}"
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()

    print(f"Filas leidas: {len(rows)}")
    print(f"Umbral similitud: {args.sim_threshold}%")

    lite_rows: list[dict[str, Any]] = []
    rows_by_id: dict[str, dict[str, Any]] = {}
    for r in rows:
        rid = str(r.get(id_col, "")).strip()
        rows_by_id[rid] = r
        full_name = build_full_name(r, name_col, second_name_col, ap1_col, ap2_col)
        lite_rows.append(
            {
                "id": rid,
                "nombre": full_name,
                "ap1": str(r.get(ap1_col, "") or "").strip(),
                "ap2": str(r.get(ap2_col, "") or "").strip() if ap2_col else "",
                "dob": str(r.get(birth_col, "") or "").strip()[:10] if birth_col else "",
                "doc": str(r.get(doc_col, "") or "").strip() if doc_col else "",
                "mail": str(r.get(mail_col, "") or "").strip() if mail_col else "",
                "phone": str(r.get(phone_col, "") or "").strip() if phone_col else "",
                "estatus": str(r.get(status_col, "") or "").strip() if status_col else "",
            }
        )

    # Coincidencias exactas por llaves fuertes.
    exact_key_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in lite_rows:
        doc = normalize_text(r["doc"])
        mail = normalize_text(r["mail"])
        phone = only_digits(r["phone"])
        full_name = normalize_text(f"{r['nombre']} {r['ap1']} {r['ap2']}")
        if doc and len(doc) >= 6:
            exact_key_map[f"DOC|{doc}"].append(r)
        if mail and "@" in mail:
            exact_key_map[f"MAIL|{mail}"].append(r)
        if phone and len(phone) >= 8:
            exact_key_map[f"PHONE|{phone}"].append(r)
        if r["dob"] and full_name:
            exact_key_map[f"NAME_DOB|{full_name}|{r['dob']}"].append(r)

    exact_groups: list[dict[str, Any]] = []
    for key, members in exact_key_map.items():
        if len(members) < 2:
            continue
        kind = key.split("|", 1)[0]
        reason = {
            "DOC": "Documento repetido",
            "MAIL": "Correo repetido",
            "PHONE": "Telefono repetido",
            "NAME_DOB": "Nombre completo + fecha nacimiento repetido",
        }.get(kind, "Coincidencia exacta")
        exact_groups.append({"reason": reason, "key": key, "members": members})
    exact_groups.sort(key=lambda x: len(x["members"]), reverse=True)

    # Coincidencias por similitud.
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in lite_rows:
        ap1 = normalize_text(r["ap1"])
        ap2 = normalize_text(r["ap2"])
        name = normalize_text(r["nombre"])
        if not ap1 or not name:
            continue
        first_name = name.split(" ")[0] if name else ""
        bucket_key = f"{ap1}|{ap2}|{first_name[:1]}"
        buckets[bucket_key].append(r)

    pair_rows: list[dict[str, Any]] = []
    seen_pairs: set[str] = set()
    for bucket_members in buckets.values():
        n = len(bucket_members)
        if n < 2:
            continue
        if n > 140:
            continue
        for i in range(n):
            for j in range(i + 1, n):
                a = bucket_members[i]
                b = bucket_members[j]
                if a["id"] == b["id"]:
                    continue
                key = "|".join(sorted([a["id"], b["id"]]))
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                full_a = normalize_text(f"{a['nombre']} {a['ap1']} {a['ap2']}")
                full_b = normalize_text(f"{b['nombre']} {b['ap1']} {b['ap2']}")
                sim = similarity_ratio(full_a, full_b)

                same_ap1 = normalize_text(a["ap1"]) and normalize_text(a["ap1"]) == normalize_text(b["ap1"])
                same_ap2 = normalize_text(a["ap2"]) and normalize_text(a["ap2"]) == normalize_text(b["ap2"])
                same_dob = a["dob"] and a["dob"] == b["dob"]

                reasons: list[str] = []
                if sim >= args.sim_threshold and (same_ap1 or same_ap2):
                    reasons.append("Nombre parecido + apellido coincidente")
                if same_ap1 and same_ap2:
                    reasons.append("Dos apellidos iguales")
                if not reasons:
                    continue
                if same_dob:
                    reasons.append("Misma fecha nacimiento")

                risk_score = sim + (10 if same_dob else 0) + (8 if same_ap1 and same_ap2 else 0)
                risk_level = "ALTO" if risk_score >= 95 else ("MEDIO" if risk_score >= 85 else "BAJO")

                pair_rows.append(
                    {
                        "tipo": "posible",
                        "reglas": " | ".join(reasons),
                        "similitud_pct": sim,
                        "riesgo": risk_level,
                        "score": round(risk_score, 1),
                        "id_a": a["id"],
                        "nombre_a": a["nombre"],
                        "id_b": b["id"],
                        "nombre_b": b["nombre"],
                        "estatus_a": a["estatus"],
                        "estatus_b": b["estatus"],
                        "doc_a": a["doc"],
                        "doc_b": b["doc"],
                        "mail_a": a["mail"],
                        "mail_b": b["mail"],
                        "phone_a": a["phone"],
                        "phone_b": b["phone"],
                        "dob_a": a["dob"],
                        "dob_b": b["dob"],
                        "coincide_ap1": "SI" if same_ap1 else "NO",
                        "coincide_ap2": "SI" if same_ap2 else "NO",
                        "coincide_dob": "SI" if same_dob else "NO",
                    }
                )

    # Convertir exactas a pares tambien para vista unificada.
    for g in exact_groups:
        members = g["members"]
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                a = members[i]
                b = members[j]
                key = "|".join(sorted([a["id"], b["id"]]))
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                pair_rows.append(
                    {
                        "tipo": "exacta",
                        "reglas": g["reason"],
                        "similitud_pct": similarity_ratio(
                            normalize_text(f"{a['nombre']} {a['ap1']} {a['ap2']}"),
                            normalize_text(f"{b['nombre']} {b['ap1']} {b['ap2']}"),
                        ),
                        "riesgo": "ALTO",
                        "score": 100.0,
                        "id_a": a["id"],
                        "nombre_a": a["nombre"],
                        "id_b": b["id"],
                        "nombre_b": b["nombre"],
                        "estatus_a": a["estatus"],
                        "estatus_b": b["estatus"],
                        "doc_a": a["doc"],
                        "doc_b": b["doc"],
                        "mail_a": a["mail"],
                        "mail_b": b["mail"],
                        "phone_a": a["phone"],
                        "phone_b": b["phone"],
                        "dob_a": a["dob"],
                        "dob_b": b["dob"],
                        "coincide_ap1": "SI" if normalize_text(a["ap1"]) == normalize_text(b["ap1"]) else "NO",
                        "coincide_ap2": "SI" if normalize_text(a["ap2"]) == normalize_text(b["ap2"]) else "NO",
                        "coincide_dob": "SI" if a["dob"] and a["dob"] == b["dob"] else "NO",
                    }
                )

    pair_rows.sort(key=lambda x: (x["tipo"] != "exacta", -float(x["score"]), str(x["id_a"]), str(x["id_b"])))
    involved_ids = set()
    for p in pair_rows:
        involved_ids.add(str(p["id_a"]))
        involved_ids.add(str(p["id_b"]))

    # Preparar salida
    if args.out:
        out_path = args.out
        if not os.path.isabs(out_path):
            out_path = os.path.join(os.getcwd(), out_path)
    else:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(os.getcwd(), f"scripts/reporte_duplicados_persona_{ts}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws_pairs = wb.active
    ws_pairs.title = "Coincidencias"
    pair_headers = [
        "similitud_pct",
        "id_a",
        "nombre_a",
        "estatus_a",
        "id_b",
        "nombre_b",
        "estatus_b",
    ]
    ws_pairs.append(pair_headers)
    for p in pair_rows:
        ws_pairs.append([p.get(h, "") for h in pair_headers])
    style_sheet(ws_pairs)

    wb.save(out_path)
    conn.close()

    print(f"Coincidencias detectadas (pares): {len(pair_rows)}")
    print(f"Personas involucradas: {len(involved_ids)}")
    print(f"Excel generado: {out_path}")


if __name__ == "__main__":
    main()

