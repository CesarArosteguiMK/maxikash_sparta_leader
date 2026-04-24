# Genera mapeo Pantalla / Campo / Fuente / Etiqueta (UTF-8)
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROWS = [
    ("Principal", "Nombre del cliente", "S2Credit", '"nombreCliente"'),
    ("Principal", "Vencido/Al corriente", "S2Credit", '"bandera"'),
    ("Principal", "Total a pagar", "resumenCondonaciones", '"totalAPagar"'),
    ("Principal", "Número de cuotas", "resumenCondonaciones", '"numeroCuotasCredito"'),
    ("Principal", "Saldo vencido", "resumenCondonaciones", '"saldoVencidoCredito"'),
    ("Principal", "Cargo por pago tardio", "resumenCondonaciones", '"totalCargosPagosTardio"'),
    ("Tu progreso", "Tu progreso", "S2Credit", '"cuotasPagadas" / "cuotasContratadas"'),
    ("Tu progreso", "Pagos semanales de", "S2Credit", '"cuota"'),
    ("Tu progreso", "Por pagar", "S2Credit", '"adeudoTotal"'),
    (
        "Últimos movimientos",
        "Últimos movimientos",
        "S2Credit",
        '"fechaDeposito", "montoPago", "moratorios"',
    ),
    (
        "Perfil",
        "Total a pagar",
        "S2Credit",
        '("cuotasContratadas")*("cuota")',
    ),
    ("Perfil", "Celular", "S2Credit", '"celular"'),
    ("Perfil", "Inicio de financiamiento", "S2Credit", '"primerVencimiento"'),
    ("Perfil", "Fin de financiamiento", "S2Credit", '"ultimoVencimiento"'),
    (
        "Datos para la transferencia",
        "Cuenta clabe personalizada",
        "S2Credit",
        '"referenciaSTP"',
    ),
    ("Datos para la transferencia", "Banco destino", "MaxiApp", '"banco"'),
    (
        "Datos para la transferencia",
        "Nombre del destinatario",
        "S2Credit",
        '"nombreCliente"',
    ),
    (
        "Detalles del financiamiento",
        "Fecha de abono inicial",
        "S2Credit",
        '"primerVencimiento"',
    ),
    ("Detalles del financiamiento", "Forma de Pago", "MaxiApp", '"metodoPago"'),
    ("Detalles del financiamiento", "Agencia", "MaxiApp", '"nombreSucursal"'),
    ("Próximo pago", "Próximo pago", "S2Credit", '"fechaSiguientePago"'),
]


def main() -> int:
    out = Path(__file__).resolve().parent.parent / "mapeo_pantallas_fuente_etiqueta.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapeo"

    header = ("Pantalla", "Campo", "Fuente", "Etiqueta")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    yellow = PatternFill("solid", fgColor="FFFFCC")

    for col, h in enumerate(header, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for r, row in enumerate(ROWS, 2):
        is_last = r == len(ROWS) + 1
        for col, val in enumerate(row, 1):
            c = ws.cell(r, col, val)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if is_last:
                c.fill = yellow

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 48

    wb.save(out)
    print(str(out))
    return 0


if __name__ == "__main__":
    sys.exit(main())
