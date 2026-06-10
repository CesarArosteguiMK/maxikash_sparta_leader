import datetime as dt
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def txt(value):
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if text.startswith("="):
        return ""
    return re.sub(r"\s+", " ", text).strip()


def keyn(value):
    text = txt(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.upper().replace("&", "Y")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean(value, max_len=255):
    text = txt(value)
    if keyn(text) in {"", "N A", "NA", "S N", "SN", "NO APLICA", "NULL", "NONE"}:
        return ""
    return text[:max_len]


def org_text(value):
    text = clean(value, 180)
    return "COBRANZA CORPORATIVO" if keyn(text) == "COBRANZA COPORATIVO" else text


def fecha(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = clean(value, 40)
    if not text:
        return None
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})$", text)
    if m:
        year = int(m.group(3))
        if year < 100:
            year += 1900 if year > 40 else 2000
        return f"{year:04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def decimal(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value, 40)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    return float(match.group(0)) if match else None


def split_nombre(nombres):
    return nombres.strip(), ""


def header_map(row):
    headers = {}
    for idx, value in enumerate(row, start=1):
        key = keyn(value)
        if key:
            headers[key] = idx
    return headers


def val(row, headers, names):
    if isinstance(names, str):
        names = [names]
    for name in names:
        idx = headers.get(keyn(name))
        if idx and idx <= len(row):
            return row[idx - 1]
    return None


def cval(row, col):
    return row[col - 1] if col <= len(row) else None


def make_record(row, headers, sheet_name, row_number, id_pais):
    es_guatemala = "GUATEMALA" in sheet_name
    nombres_excel = clean(val(row, headers, "NOMBRE (S)"), 120)
    nombres, segundo = split_nombre(nombres_excel)
    apellidop = clean(val(row, headers, "A. PATERNO"), 120)
    apellidom = clean(val(row, headers, "A. MATERNO"), 120)
    nombre_completo = clean(val(row, headers, "NOMBRE/APELLIDOS"), 260)
    if not nombre_completo:
        nombre_completo = f"{nombres_excel} {apellidop} {apellidom}".strip()

    sexo_key = keyn(val(row, headers, "SEXO"))
    sexo = "Femenino" if sexo_key in {"F", "FEMENINO"} else "Masculino" if sexo_key in {"M", "MASCULINO"} else clean(val(row, headers, "SEXO"), 20)

    return {
        "sheet": sheet_name,
        "row": row_number,
        "id_pais": id_pais,
        "nombre_completo": nombre_completo,
        "nombres": nombres,
        "segundo_nombre": segundo,
        "apellidop": apellidop,
        "apellidom": apellidom,
        "fecha_ingreso": fecha(val(row, headers, "FECHA DE INGRESO")),
        "fecha_contpaq": fecha(val(row, headers, "FECHA CONTPAC")),
        "fecha_imss_alta": fecha(val(row, headers, "FECHA IMSS ALTA")),
        "puesto": org_text(val(row, headers, "PUESTO")),
        "departamento": org_text(val(row, headers, "DEPARTAMENTO")),
        "area": org_text(val(row, headers, "AREA")),
        "direccion": org_text(val(row, headers, ["DIRECCION ORGANIZACIONAL", "DIRECCION"])),
        "ubicacion_laboral": clean(val(row, headers, "UBICACION LABORAL"), 180),
        "municipio_laboral": clean(val(row, headers, "MUNICIPIO"), 180),
        "jefe_directo_texto": clean(val(row, headers, "JEFE DIRECTO"), 220),
        "sueldo_neto": decimal(val(row, headers, "SUELDO NETO")),
        "sueldo_quincenal": decimal(val(row, headers, "SUELDO QUINCENAL")),
        "sueldo_bruto": decimal(val(row, headers, ["SUELDO BRUTO", "SUELDO BASE (BRUTO)"])),
        "salario_diario": decimal(val(row, headers, "SALARIO DIARIO")),
        "sbc": decimal(val(row, headers, "SBC")),
        "curp": clean(val(row, headers, "CURP"), 18).upper(),
        "nss": clean(val(row, headers, ["NSS", "IGSS"]), 20),
        "rfc": clean(val(row, headers, ["RFC", "RTU"]), 20).upper(),
        "entidad_federativa_rfc": clean(val(row, headers, "ENTIDAD FEDERATIVA / RFC"), 120),
        "codigo_postal": clean(val(row, headers, "CP"), 12),
        "fecha_nacimiento": fecha(val(row, headers, "FECHA DE NACIMIENTO")),
        "sexo": sexo,
        "telefono": clean(cval(row, 22 if es_guatemala else 36), 30),
        "correo": clean(cval(row, 23 if es_guatemala else 37), 160),
        "domicilio": clean(cval(row, 24 if es_guatemala else 38), 500),
        "registro_patronal": clean(val(row, headers, "REGISTRO PATRONAL"), 120),
        "codigo_contpaq": clean(val(row, headers, "CODIGO CONTPAC"), 80),
        "carta_no_credito": clean(val(row, headers, 'CARTA DE "NO CREDITO"'), 120),
        "credito_infonavit_fonacot": clean(val(row, headers, "CREDITO INFONAVIT/ FONACOT"), 80),
        "no_credito": clean(val(row, headers, "NO. DE CREDITO"), 80),
        "monto_descontar": decimal(val(row, headers, "MONTO A DESCONTAR")),
        "carta_no_nomina_bbva": clean(val(row, headers, 'CARTA "NO NOMINA EN BBVA"'), 120),
        "id_banco": clean(cval(row, 25 if es_guatemala else 44), 20),
        "nombre_banco": clean(cval(row, 26 if es_guatemala else 45), 120),
        "numero_cuenta": clean(cval(row, 27 if es_guatemala else 46), 40),
        "clabe": clean(cval(row, 28 if es_guatemala else 47), 30),
        "contacto1": "" if es_guatemala else clean(cval(row, 48), 220),
        "parentesco1": "" if es_guatemala else clean(cval(row, 49), 80),
        "telefono_contacto1": "" if es_guatemala else clean(cval(row, 50), 30),
        "contacto2": "" if es_guatemala else clean(cval(row, 51), 220),
        "parentesco2": "" if es_guatemala else clean(cval(row, 52), 80),
        "telefono_contacto2": "" if es_guatemala else clean(cval(row, 53), 30),
        "sueldo_bruto_letra": clean(val(row, headers, "SUELDO BRUTO-LETRA"), 255),
        "observaciones": clean(val(row, headers, "OBSERVACIONES"), 5000),
    }


def main():
    if len(sys.argv) < 3:
        print("Uso: python scripts/extraer_control_empleados_2026.py <archivo.xlsx> <salida.json>", file=sys.stderr)
        return 1
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = load_workbook(input_path, read_only=True, data_only=True)
    sheets = [
        ("ACTIVOS MAXI 2026", 11, 1),
        ("ACTIVOS GUATEMALA 2026", 12, 2),
        ("ACTIVOS FURIA 26", 11, 1),
    ]
    records = []
    for sheet_name, header_row, id_pais in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = None
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row, max_col=61, values_only=True), start=header_row):
            if row_number == header_row:
                header = header_map(row)
                continue
            record = make_record(row, header or {}, sheet_name, row_number, id_pais)
            if not record["nombre_completo"] and not record["puesto"] and not record["departamento"]:
                continue
            records.append(record)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(records, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"records": len(records), "output": str(output_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
